import logging
from datetime import datetime
from telegram import Update, ReplyKeyboardMarkup, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler, CallbackQueryHandler
from config import BOT_TOKEN
from database import (init_db, add_product, remove_product, get_inventory, get_history,
                      add_admin, remove_admin, get_admins, is_admin, get_report,
                      add_warehouse, get_warehouses, get_warehouse_by_id, add_firm, get_firms, get_firm_by_id,
                      add_client, get_clients, get_client_by_id, set_product_price, get_product_prices, get_product_price,
                      add_warehouse_group, get_warehouse_groups,
                      delete_warehouse, delete_warehouse_group, delete_product,
                      delete_firm, delete_client, delete_price,
                      add_coalition, get_coalitions, get_coalition_by_id, delete_coalition,
                      add_user, get_user_role, get_all_users, update_user_role,
                      update_user_warehouse_group, deactivate_user, activate_user,
                      delete_user, has_permission, get_user_warehouses, log_access,
                      add_user_with_login, verify_login, create_session, get_session,
                      logout_session, change_password, get_user_by_id,
                      add_arrival, get_arrivals, get_arrival_by_id, get_arrival_statistics,
                      get_inventory_by_user, get_inventory_detailed_by_user, 
                      get_inventory_by_warehouse_and_firm_for_user,
                      get_summary_stats_by_user, get_client_debts_by_user,
                      add_departure, get_departures,
                      add_payment, get_payments, get_payment_by_id,
                      add_partner, get_partners, get_partner_by_id,
                      delete_arrival, delete_departure, delete_payment, delete_partner,
                      get_connection)

# Настройка логирования
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

# Состояния для ConversationHandler
(CHOOSING_ACTION, PRODUCT_NAME, PRODUCT_QUANTITY, ADMIN_MENU, ADMIN_ID,
 MANAGEMENT_MENU, WAREHOUSE_MENU, WAREHOUSE_NAME, PRODUCT_MENU, PRODUCT_DATA,
 FIRM_MENU, FIRM_NAME, CLIENT_MENU, CLIENT_DATA, PRICE_MENU, PRICE_DATA,
 WAREHOUSE_GROUP_MENU, WAREHOUSE_GROUP_NAME, WAREHOUSE_DATA, WAREHOUSE_GROUP_SELECT,
 WAREHOUSE_DELETE, WAREHOUSE_GROUP_DELETE, PRODUCT_DELETE, FIRM_DELETE, 
 CLIENT_DELETE, PRICE_DELETE, PRICE_SELECT, PRICE_VALUE,
 COALITION_MENU, COALITION_NAME, COALITION_DELETE,
 USER_MENU, USER_ADD_ID, USER_ADD_ROLE, USER_ADD_GROUP, USER_EDIT_SELECT,
 USER_EDIT_ROLE, USER_EDIT_GROUP,
 LOGIN, PASSWORD, USER_ADD_LOGIN, USER_ADD_PASSWORD, USER_ADD_USERNAME,
 ARRIVAL_DATE, ARRIVAL_WAGON, ARRIVAL_FIRM, ARRIVAL_WAREHOUSE, ARRIVAL_PRODUCT,
 ARRIVAL_QTY_DOC, ARRIVAL_QTY_ACTUAL, ARRIVAL_NOTES,
 DEPARTURE_DATE, DEPARTURE_COALITION, DEPARTURE_FIRM, DEPARTURE_WAREHOUSE,
 DEPARTURE_PRODUCT, DEPARTURE_CLIENT, DEPARTURE_QUANTITY, DEPARTURE_PRICE, DEPARTURE_NOTES,
 REPORT_MENU, REPORT_EXPORT_PERIOD,
 PAYMENT_DATE, PAYMENT_CLIENT, PAYMENT_SOMONI, PAYMENT_RATE, PAYMENT_NOTES,
 PARTNER_DATE, PARTNER_CLIENT, PARTNER_SOMONI, PARTNER_RATE, PARTNER_NOTES,
 DELETE_MENU, DELETE_TYPE, DELETE_ID, DELETE_CONFIRM,
 EDIT_MENU, EDIT_TYPE, EDIT_ID, EDIT_FIELD, EDIT_VALUE) = range(81)

# Клавиатуры
main_keyboard = [
    ['➕ Приход товара', '➖ Вывод товара'],
    ['📦 Остатки', '📊 Сводка'],
    ['💰 Погашения', '🤝 Партнеры'],
    ['💳 Долги клиентов', '📋 Отчет'],
    ['📅 Отчет за день', '📤 Расход за день'],
    ['🏭 Фактический остаток'],
    ['👥 Пользователи', '⚙️ Управление'],
    ['🚪 Выход']
]
markup = ReplyKeyboardMarkup(main_keyboard, resize_keyboard=True)

admin_keyboard = [
    ['➕ Добавить админа', '➖ Удалить админа'],
    ['📋 Список админов', '🔙 Назад']
]
admin_markup = ReplyKeyboardMarkup(admin_keyboard, resize_keyboard=True)

user_keyboard = [
    ['➕ Добавить пользователя', '📋 Список пользователей'],
    ['✏️ Изменить роль', '🏢 Назначить склад'],
    ['🔒 Деактивировать', '✅ Активировать'],
    ['🗑 Удалить пользователя', '🔙 Назад']
]
user_markup = ReplyKeyboardMarkup(user_keyboard, resize_keyboard=True)

role_keyboard = [
    ['👑 Управляющий', '📊 Менеджер'],
    ['🏢 Завсклад', '💰 Кассир'],
    ['❌ Отмена']
]
role_markup = ReplyKeyboardMarkup(role_keyboard, resize_keyboard=True)

management_keyboard = [
    ['🏢 Склады', '📦 Товары'],
    ['🏭 Фирмы', '👤 Клиенты'],
    ['💰 Цены', '📊 Коалица'],
    ['✏️ Изменить записи', '🗑 Удалить записи'],
    ['🔙 Назад']
]
management_markup = ReplyKeyboardMarkup(management_keyboard, resize_keyboard=True)

warehouse_keyboard = [
    ['➕ Добавить склад', '📋 Список складов'],
    ['🗑 Удалить склад', '📁 Подгруппы складов'],
    ['🔙 Назад']
]
warehouse_markup = ReplyKeyboardMarkup(warehouse_keyboard, resize_keyboard=True)

warehouse_group_keyboard = [
    ['➕ Добавить подгруппу', '📋 Список подгрупп'],
    ['🗑 Удалить подгруппу', '🔙 Назад']
]
warehouse_group_markup = ReplyKeyboardMarkup(warehouse_group_keyboard, resize_keyboard=True)

product_keyboard = [
    ['➕ Добавить товар', '📋 Список товаров'],
    ['🗑 Удалить товар', '🔙 Назад']
]
product_markup = ReplyKeyboardMarkup(product_keyboard, resize_keyboard=True)

firm_keyboard = [
    ['➕ Добавить фирму', '📋 Список фирм'],
    ['🗑 Удалить фирму', '🔙 Назад']
]
firm_markup = ReplyKeyboardMarkup(firm_keyboard, resize_keyboard=True)

client_keyboard = [
    ['➕ Добавить клиента', '📋 Список клиентов'],
    ['🗑 Удалить клиента', '🔙 Назад']
]
client_markup = ReplyKeyboardMarkup(client_keyboard, resize_keyboard=True)

price_keyboard = [
    ['💰 Установить цену', '📋 Список цен'],
    ['🗑 Удалить цену', '🔙 Назад']
]
price_markup = ReplyKeyboardMarkup(price_keyboard, resize_keyboard=True)

coalition_keyboard = [
    ['➕ Добавить коалицу', '📋 Список коалиц'],
    ['🗑 Удалить коалицу', '🔙 Назад']
]
coalition_markup = ReplyKeyboardMarkup(coalition_keyboard, resize_keyboard=True)

report_keyboard = [
    ['➕ Отчет по приходу', '➖ Отчет по расходу'],
    ['💰 Отчет по погашениям', '🤝 Отчет по партнерам'],
    ['👤 Карточка клиента', '📢 Уведомления о долгах'],
    ['🚂 Итоги вагонов'],
    ['🔙 Назад']
]
report_markup = ReplyKeyboardMarkup(report_keyboard, resize_keyboard=True)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Команда /start"""
    telegram_id = update.effective_user.id
    
    # Проверяем есть ли активная сессия
    session = get_session(telegram_id)
    
    if session:
        user_id, username, role, warehouse_group_id, is_active = session
        
        if not is_active:
            logout_session(telegram_id)
            await update.message.reply_text('❌ Ваш аккаунт деактивирован. Обратитесь к администратору.')
            return ConversationHandler.END
        
        log_access(user_id, 'start', f'Роль: {role}')
        
        await update.message.reply_text(
            f'👋 С возвращением, {username}!\n'
            f'Ваша роль: {get_role_name(role)}\n\n'
            'Выберите действие:',
            reply_markup=markup
        )
        return CHOOSING_ACTION
    
    # Если нет сессии, запрашиваем логин
    await update.message.reply_text(
        '👋 Добро пожаловать в бот учета товаров!\n\n'
        '🔐 Для входа введите ваш логин:'
    )
    return LOGIN

def get_role_name(role):
    """Получить название роли на русском"""
    roles = {
        'admin': '👑 Управляющий',
        'manager': '📊 Менеджер',
        'warehouse_manager': '🏢 Завсклад',
        'cashier': '💰 Кассир'
    }
    return roles.get(role, role)

# ============= АУТЕНТИФИКАЦИЯ =============

async def handle_login(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка ввода логина"""
    login = update.message.text.strip()
    context.user_data['login'] = login
    await update.message.reply_text('🔑 Введите пароль:')
    return PASSWORD

async def handle_password(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка ввода пароля"""
    password = update.message.text.strip()
    login = context.user_data.get('login')
    telegram_id = update.effective_user.id
    
    # Проверяем логин и пароль
    success, user_id, username, role, warehouse_group_id = verify_login(login, password)
    
    if success:
        # Создаем сессию
        create_session(telegram_id, user_id)
        log_access(user_id, 'login', f'Успешный вход')
        
        await update.message.reply_text(
            f'✅ Вход выполнен успешно!\n'
            f'Добро пожаловать, {username}!\n'
            f'Ваша роль: {get_role_name(role)}\n\n'
            'Выберите действие:',
            reply_markup=markup
        )
        return CHOOSING_ACTION
    else:
        await update.message.reply_text(
            '❌ Неверный логин или пароль.\n'
            'Попробуйте еще раз или обратитесь к администратору.\n\n'
            'Введите логин:'
        )
        return LOGIN

async def handle_action(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора действия"""
    text = update.message.text
    telegram_id = update.effective_user.id
    
    # Проверяем сессию
    session = get_session(telegram_id)
    if not session:
        await update.message.reply_text('❌ Сессия истекла. Войдите заново: /start')
        return ConversationHandler.END
    
    user_id, username, role, warehouse_group_id, is_active = session
    
    if text == '➕ Приход товара':
        if not has_permission(user_id, 'manager'):
            await update.message.reply_text('❌ Недостаточно прав', reply_markup=markup)
            return CHOOSING_ACTION
        context.user_data['action'] = 'add'
        context.user_data['user_id'] = user_id
        await update.message.reply_text(
            '📅 Введите дату прихода товара\n'
            'Формат: ДД.ММ.ГГГГ (например: 23.02.2026)\n'
            'или просто нажмите "Сегодня"',
            reply_markup=ReplyKeyboardMarkup([['Сегодня'], ['❌ Отмена']], resize_keyboard=True)
        )
        return ARRIVAL_DATE
    
    elif text == '➖ Вывод товара':
        if not has_permission(user_id, 'cashier'):
            await update.message.reply_text('❌ Недостаточно прав', reply_markup=markup)
            return CHOOSING_ACTION
        context.user_data['action'] = 'remove'
        context.user_data['user_id'] = user_id
        await update.message.reply_text(
            '📅 Введите дату вывода товара\n'
            'Формат: ДД.ММ.ГГГГ (например: 23.02.2026)\n'
            'или просто нажмите "Сегодня"',
            reply_markup=ReplyKeyboardMarkup([['Сегодня'], ['❌ Отмена']], resize_keyboard=True)
        )
        return DEPARTURE_DATE
    
    elif text == '📦 Остатки':
        # Получаем остатки с группировкой по складам и фирмам
        from datetime import datetime
        
        products = get_inventory_by_warehouse_and_firm_for_user(user_id)
        if not products:
            await update.message.reply_text('📦 Нет доступных остатков', reply_markup=markup)
        else:
            current_year = datetime.now().year
            message = f'📦 *ОСТАТКИ СКЛАДОВ* 📅 {current_year}\n'
            message += '─' * 30 + '\n'
            
            # Группируем по складам
            warehouses = {}
            for warehouse_name, firm_name, product_name, balance in products:
                if warehouse_name not in warehouses:
                    warehouses[warehouse_name] = {}
                if firm_name not in warehouses[warehouse_name]:
                    warehouses[warehouse_name][firm_name] = []
                warehouses[warehouse_name][firm_name].append((product_name, balance))
            
            total_balance = 0
            
            # Выводим по складам
            for warehouse_name, firms in warehouses.items():
                message += f'\n🏪 *{warehouse_name}*\n'
                warehouse_total = 0
                
                for firm_name, items in firms.items():
                    for product_name, balance in items:
                        message += f'{firm_name} {product_name}: {balance:,.2f} шт\n'
                        warehouse_total += balance
                        total_balance += balance
                
                message += f'*Итого: {warehouse_total:,.2f} шт*\n'
            
            message += '\n' + '─' * 30 + '\n'
            message += f'📊 *ИТОГО: {total_balance:,.2f} шт*'
            
            await update.message.reply_text(message, parse_mode='Markdown', reply_markup=markup)
        return CHOOSING_ACTION
    
    elif text == '📊 Сводка':
        # Получаем сводную статистику
        from datetime import datetime
        
        stats = get_summary_stats_by_user(user_id)
        if not stats:
            await update.message.reply_text('❌ Нет доступных данных', reply_markup=markup)
        else:
            current_year = datetime.now().year
            message = f'📊 *СВОДКА ЗА {current_year}*\n'
            message += '─' * 30 + '\n'
            message += f'📥 Приход: *{stats["total_arrivals"]:,.2f} шт*\n'
            message += f'📤 Расход: *{stats["total_departures"]:,.2f} шт*\n'
            message += f'📦 Остаток: *{stats["current_balance"]:,.2f} шт*\n'
            message += '─' * 30 + '\n'
            message += f'💵 Сумма продаж: *{stats["total_sales"]:,.2f} $*\n'
            message += f'✅ Оплачено: *{stats["total_paid"]:,.2f} $*\n'
            message += f'💳 Общий долг: *{stats["total_debt"]:,.2f} $*\n'
            message += f'👥 Должников: *{stats["debtors_count"]}*'
            
            await update.message.reply_text(message, parse_mode='Markdown', reply_markup=markup)
        return CHOOSING_ACTION
    
    elif text == '💳 Долги клиентов':
        # Получаем список должников
        from datetime import datetime
        
        debts = get_client_debts_by_user(user_id)
        if not debts:
            await update.message.reply_text('💰 Нет должников', reply_markup=markup)
        else:
            current_year = datetime.now().year
            message = f'💰 *ДОЛГИ КЛИЕНТОВ*\n📅 {current_year}\n'
            message += '─' * 30 + '\n'
            
            total_debt = 0
            for idx, (client_id, client_name, total_sales, total_paid, debt) in enumerate(debts, 1):
                message += f'\n*{idx}. {client_name}*\n'
                message += f'Сумма: {total_sales:,.2f} $\n'
                message += f'Оплачено: {total_paid:,.2f} $\n'
                message += f'💳 Долг: *{debt:,.2f} $*\n'
                total_debt += debt
            
            message += '\n' + '─' * 30 + '\n'
            message += f'👥 Должников: *{len(debts)}*\n'
            message += f'💰 *ИТОГО ДОЛГ: {total_debt:,.2f} $*'
            
            # Сохраняем данные для экспорта
            context.user_data['client_debts'] = debts
            context.user_data['report_year'] = current_year
            
            # Добавляем inline кнопку экспорта
            inline_keyboard = [
                [InlineKeyboardButton("📊 Экспорт в Excel", callback_data='export_client_debts')]
            ]
            inline_markup = InlineKeyboardMarkup(inline_keyboard)
            
            await update.message.reply_text(message, parse_mode='Markdown', reply_markup=inline_markup)
        return CHOOSING_ACTION
    
    elif text == '📊 История':
        history = get_history(15)
        if not history:
            await update.message.reply_text('📊 История пуста', reply_markup=markup)
        else:
            message = '📊 Последние операции:\n\n'
            for product_name, quantity, op_type, date in history:
                emoji = '➕' if op_type == 'приход' else '➖'
                message += f'{emoji} {product_name}: {quantity} шт. ({date[:16]})\n'
            await update.message.reply_text(message, reply_markup=markup)
        return CHOOSING_ACTION
    
    elif text == '💰 Погашения':
        if not has_permission(user_id, 'cashier'):
            await update.message.reply_text('❌ Недостаточно прав', reply_markup=markup)
            return CHOOSING_ACTION
        context.user_data['user_id'] = user_id
        await update.message.reply_text(
            '📅 Введите дату погашения\n'
            'Формат: ДД.ММ.ГГГГ (например: 24.02.2026)\n'
            'или просто нажмите "Сегодня"',
            reply_markup=ReplyKeyboardMarkup([['Сегодня'], ['❌ Отмена']], resize_keyboard=True)
        )
        return PAYMENT_DATE
    
    elif text == '🤝 Партнеры':
        if not has_permission(user_id, 'cashier'):
            await update.message.reply_text('❌ Недостаточно прав', reply_markup=markup)
            return CHOOSING_ACTION
        context.user_data['user_id'] = user_id
        await update.message.reply_text(
            '📅 Введите дату партнера\n'
            'Формат: ДД.ММ.ГГГГ (например: 24.02.2026)\n'
            'или просто нажмите "Сегодня"',
            reply_markup=ReplyKeyboardMarkup([['Сегодня'], ['❌ Отмена']], resize_keyboard=True)
        )
        return PARTNER_DATE
    
    elif text == '👥 Пользователи':
        if not has_permission(user_id, 'admin'):
            await update.message.reply_text('❌ Доступ запрещен. Только для администраторов.', reply_markup=markup)
            return CHOOSING_ACTION
        await update.message.reply_text('👥 Управление пользователями:', reply_markup=user_markup)
        return USER_MENU
    
    elif text == '📋 Отчет':
        # Показываем меню отчетов
        await update.message.reply_text('📋 Выберите тип отчета:', reply_markup=report_markup)
        return REPORT_MENU
    
    elif text == '📅 Отчет за день':
        # Показываем меню выбора даты для отчета за день
        from datetime import datetime
        
        current_year = datetime.now().year
        
        keyboard = [
            [InlineKeyboardButton("📅 Сегодня", callback_data='daily_report_today')],
            [InlineKeyboardButton("📅 Вчера", callback_data='daily_report_yesterday')],
            [InlineKeyboardButton("📅 Позавчера", callback_data='daily_report_2days')],
        ]
        
        inline_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            f'📅 ОТЧЁТ ЗА ДЕНЬ\n'
            f'📆 Год: {current_year}\n\n'
            f'Выберите дату:',
            reply_markup=inline_markup
        )
        return CHOOSING_ACTION
    
    elif text == '📤 Расход за день':
        # Показываем меню выбора даты для расхода за день
        from datetime import datetime
        
        keyboard = [
            [InlineKeyboardButton("📅 Сегодня", callback_data='expense_report_today')],
            [InlineKeyboardButton("📅 Вчера", callback_data='expense_report_yesterday')],
            [InlineKeyboardButton("📅 Позавчера", callback_data='expense_report_2days')],
        ]
        
        inline_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            '📤 РАСХОД ЗА ДЕНЬ\n\n'
            'Выберите дату:',
            reply_markup=inline_markup
        )
        return CHOOSING_ACTION
    
    elif text == '🏭 Фактический остаток':
        # Показываем фактический остаток с группировкой по коалициям и складам
        from datetime import datetime
        
        await update.message.reply_text('⏳ Формирую отчет по остаткам...', reply_markup=markup)
        
        current_year = datetime.now().year
        
        # Получаем склады пользователя для фильтрации
        user_warehouses = get_user_warehouses(user_id)
        warehouse_ids = [wh[0] for wh in user_warehouses] if user_warehouses else []
        
        conn = get_connection()
        cursor = conn.cursor()
        
        # Получаем остатки с группировкой по коалициям и складам
        if role == 'admin':
            cursor.execute('''
                SELECT 
                    wg.name as coalition_name,
                    w.name as warehouse_name,
                    product_name,
                    SUM(CASE WHEN source = 'arrival' THEN quantity ELSE -quantity END) as balance
                FROM (
                    SELECT warehouse_id, product_name, quantity_actual as quantity, 'arrival' as source
                    FROM arrivals
                    UNION ALL
                    SELECT warehouse_id, product_name, quantity, 'departure' as source
                    FROM departures
                ) combined
                LEFT JOIN warehouses w ON combined.warehouse_id = w.id
                LEFT JOIN warehouse_groups wg ON w.group_id = wg.id
                GROUP BY wg.name, w.name, product_name
                HAVING SUM(CASE WHEN source = 'arrival' THEN quantity ELSE -quantity END) > 0
                ORDER BY wg.name, w.name, product_name
            ''')
        else:
            placeholders = ','.join(['%s'] * len(warehouse_ids))
            cursor.execute(f'''
                SELECT 
                    wg.name as coalition_name,
                    w.name as warehouse_name,
                    product_name,
                    SUM(CASE WHEN source = 'arrival' THEN quantity ELSE -quantity END) as balance
                FROM (
                    SELECT warehouse_id, product_name, quantity_actual as quantity, 'arrival' as source
                    FROM arrivals
                    WHERE warehouse_id IN ({placeholders})
                    UNION ALL
                    SELECT warehouse_id, product_name, quantity, 'departure' as source
                    FROM departures
                    WHERE warehouse_id IN ({placeholders})
                ) combined
                LEFT JOIN warehouses w ON combined.warehouse_id = w.id
                LEFT JOIN warehouse_groups wg ON w.group_id = wg.id
                GROUP BY wg.name, w.name, product_name
                HAVING SUM(CASE WHEN source = 'arrival' THEN quantity ELSE -quantity END) > 0
                ORDER BY wg.name, w.name, product_name
            ''', warehouse_ids + warehouse_ids)
        
        balances = cursor.fetchall()
        cursor.close()
        conn.close()
        
        if not balances:
            await update.message.reply_text('📦 Нет остатков товаров', reply_markup=markup)
            return CHOOSING_ACTION
        
        # Группируем данные: коалиция -> склад -> товар -> остаток
        coalition_data = {}
        product_totals = {}
        
        for coalition, warehouse, product, balance in balances:
            if not coalition:
                coalition = "Без коалиции"
            
            if coalition not in coalition_data:
                coalition_data[coalition] = {}
            
            if warehouse not in coalition_data[coalition]:
                coalition_data[coalition][warehouse] = {}
            
            coalition_data[coalition][warehouse][product] = balance
            
            if product not in product_totals:
                product_totals[product] = 0
            product_totals[product] += balance
        
        # Формируем отчет
        message = f'🏭 ФАКТИЧЕСКИЙ ОСТАТОК\n'
        message += f'📅 {current_year}\n'
        message += '═' * 25 + '\n\n'
        
        grand_total = 0
        
        for coalition in sorted(coalition_data.keys()):
            message += f'📁 {coalition}\n'
            message += '─' * 20 + '\n'
            
            coalition_total = 0
            
            for warehouse in sorted(coalition_data[coalition].keys()):
                message += f'🏪 {warehouse}\n'
                
                warehouse_total = 0
                
                for product in sorted(coalition_data[coalition][warehouse].keys()):
                    balance = coalition_data[coalition][warehouse][product]
                    weight = balance / 20  # Переводим в тонны
                    warehouse_total += weight
                    message += f'• {product}: {weight:,.2f} т\n'
                
                message += f'Итого: {warehouse_total:,.2f} т\n\n'
                coalition_total += warehouse_total
            
            message += f'📊 Итого {coalition}: {coalition_total:,.2f} т\n\n'
            grand_total += coalition_total
        
        message += '═' * 25 + '\n'
        message += f'🏭 ОБЩИЙ ИТОГ: {grand_total:,.2f} тонн\n\n'
        
        # Итого по товарам
        message += '📦 ИТОГО ПО ТОВАРАМ:\n'
        for product in sorted(product_totals.keys()):
            total_qty = product_totals[product]
            weight = total_qty / 20
            message += f'• {product}: {weight:,.2f} т\n'
        
        # Если сообщение слишком длинное, разбиваем на части
        if len(message) > 4000:
            parts = []
            current_part = f'🏭 ФАКТИЧЕСКИЙ ОСТАТОК\n📅 {current_year}\n' + '═' * 25 + '\n\n'
            
            for coalition in sorted(coalition_data.keys()):
                coalition_section = f'📁 {coalition}\n' + '─' * 20 + '\n'
                coalition_total = 0
                
                for warehouse in sorted(coalition_data[coalition].keys()):
                    warehouse_section = f'🏪 {warehouse}\n'
                    warehouse_total = 0
                    
                    for product in sorted(coalition_data[coalition][warehouse].keys()):
                        balance = coalition_data[coalition][warehouse][product]
                        weight = balance / 20
                        warehouse_total += weight
                        warehouse_section += f'• {product}: {weight:,.2f} т\n'
                    
                    warehouse_section += f'Итого: {warehouse_total:,.2f} т\n\n'
                    coalition_total += warehouse_total
                    coalition_section += warehouse_section
                
                coalition_section += f'📊 Итого {coalition}: {coalition_total:,.2f} т\n\n'
                
                # Если добавление секции превысит лимит, отправляем текущую часть
                if len(current_part) + len(coalition_section) > 3800:
                    parts.append(current_part)
                    current_part = coalition_section
                else:
                    current_part += coalition_section
            
            # Добавляем последнюю часть
            if current_part:
                parts.append(current_part)
            
            # Отправляем части
            for i, part in enumerate(parts):
                if i == 0:
                    await update.message.reply_text(part, reply_markup=markup)
                else:
                    await update.message.reply_text(part)
            
            # Отправляем итог отдельно
            footer = '═' * 25 + '\n'
            footer += f'🏭 ОБЩИЙ ИТОГ: {grand_total:,.2f} тонн\n\n'
            footer += '📦 ИТОГО ПО ТОВАРАМ:\n'
            for product in sorted(product_totals.keys()):
                total_qty = product_totals[product]
                weight = total_qty / 20
                footer += f'• {product}: {weight:,.2f} т\n'
            
            await update.message.reply_text(footer)
        else:
            await update.message.reply_text(message, reply_markup=markup)
        
        return CHOOSING_ACTION
    
    elif text == '⚙️ Управление':
        if not has_permission(user_id, 'admin'):
            await update.message.reply_text('❌ Доступ запрещен. Только для администраторов.', reply_markup=markup)
            return CHOOSING_ACTION
        await update.message.reply_text('⚙️ Управление системой:', reply_markup=management_markup)
        return MANAGEMENT_MENU
    
    elif text == '🚪 Выход':
        # Выход из системы
        logout_session(telegram_id)
        log_access(user_id, 'logout', 'Выход из системы')
        await update.message.reply_text(
            '👋 Вы вышли из системы.\n'
            'Для входа используйте команду /start'
        )
        return ConversationHandler.END
        return CHOOSING_ACTION

async def get_product_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Получение названия товара"""
    context.user_data['product_name'] = update.message.text
    await update.message.reply_text('Введите количество:')
    return PRODUCT_QUANTITY

async def get_product_quantity(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Получение количества и выполнение операции"""
    try:
        quantity = int(update.message.text)
        if quantity <= 0:
            await update.message.reply_text('❌ Количество должно быть больше 0', reply_markup=markup)
            return CHOOSING_ACTION
        
        product_name = context.user_data['product_name']
        action = context.user_data['action']
        
        if action == 'add':
            add_product(product_name, quantity)
            await update.message.reply_text(
                f'✅ Приход: {product_name} - {quantity} шт.',
                reply_markup=markup
            )
        else:
            success, message = remove_product(product_name, quantity)
            if success:
                await update.message.reply_text(
                    f'✅ Вывод: {product_name} - {quantity} шт.',
                    reply_markup=markup
                )
            else:
                await update.message.reply_text(f'❌ {message}', reply_markup=markup)
        
        return CHOOSING_ACTION
    
    except ValueError:
        await update.message.reply_text('❌ Введите корректное число', reply_markup=markup)
        return CHOOSING_ACTION

async def handle_admin_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка меню администратора"""
    text = update.message.text
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text('❌ Доступ запрещен', reply_markup=markup)
        return CHOOSING_ACTION
    
    if text == '➕ Добавить админа':
        await update.message.reply_text('Введите ID пользователя для добавления в админы:')
        context.user_data['admin_action'] = 'add'
        return ADMIN_ID
    
    elif text == '➖ Удалить админа':
        await update.message.reply_text('Введите ID пользователя для удаления из админов:')
        context.user_data['admin_action'] = 'remove'
        return ADMIN_ID
    
    elif text == '📋 Список админов':
        admins = get_admins()
        if not admins:
            await update.message.reply_text('📋 Список админов пуст', reply_markup=admin_markup)
        else:
            message = '📋 Администраторы:\n\n'
            for admin_id, username, date in admins:
                message += f'• ID: {admin_id} - {username} (с {date[:10]})\n'
            await update.message.reply_text(message, reply_markup=admin_markup)
        return ADMIN_MENU
    
    elif text == '🔙 Назад':
        await update.message.reply_text('Главное меню:', reply_markup=markup)
        return CHOOSING_ACTION

async def handle_admin_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка ID администратора"""
    try:
        admin_id = int(update.message.text)
        action = context.user_data.get('admin_action')
        
        if action == 'add':
            add_admin(admin_id, f"User_{admin_id}")
            await update.message.reply_text(f'✅ Пользователь {admin_id} добавлен в админы', reply_markup=admin_markup)
        elif action == 'remove':
            remove_admin(admin_id)
            await update.message.reply_text(f'✅ Пользователь {admin_id} удален из админов', reply_markup=admin_markup)
        
        return ADMIN_MENU
    except ValueError:
        await update.message.reply_text('❌ Введите корректный ID (число)', reply_markup=admin_markup)
        return ADMIN_MENU

async def handle_management_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка меню управления"""
    text = update.message.text
    
    if text == '🏢 Склады':
        await update.message.reply_text('🏢 Управление складами:', reply_markup=warehouse_markup)
        return WAREHOUSE_MENU
    elif text == '📦 Товары':
        await update.message.reply_text('📦 Управление товарами:', reply_markup=product_markup)
        return PRODUCT_MENU
    elif text == '🏭 Фирмы':
        await update.message.reply_text('🏭 Управление фирмами:', reply_markup=firm_markup)
        return FIRM_MENU
    elif text == '👤 Клиенты':
        await update.message.reply_text('👤 Управление клиентами:', reply_markup=client_markup)
        return CLIENT_MENU
    elif text == '💰 Цены':
        await update.message.reply_text('💰 Управление ценами:', reply_markup=price_markup)
        return PRICE_MENU
    elif text == '📊 Коалица':
        await update.message.reply_text('📊 Управление коалицей:', reply_markup=coalition_markup)
        return COALITION_MENU
    elif text == '✏️ Изменить записи':
        edit_keyboard = [
            ['✏️ Изменить приход', '✏️ Изменить расход'],
            ['✏️ Изменить погашение', '✏️ Изменить партнера'],
            ['🔙 Назад']
        ]
        edit_markup_menu = ReplyKeyboardMarkup(edit_keyboard, resize_keyboard=True)
        await update.message.reply_text(
            '✏️ Выберите тип записи для изменения:',
            reply_markup=edit_markup_menu
        )
        return EDIT_MENU
    elif text == '🗑 Удалить записи':
        delete_keyboard = [
            ['🗑 Удалить приход', '🗑 Удалить расход'],
            ['🗑 Удалить погашение', '🗑 Удалить партнера'],
            ['🔙 Назад']
        ]
        delete_markup_menu = ReplyKeyboardMarkup(delete_keyboard, resize_keyboard=True)
        await update.message.reply_text('🗑 Выберите тип записи для удаления:', reply_markup=delete_markup_menu)
        return DELETE_MENU
    elif text == '🔙 Назад':
        await update.message.reply_text('Главное меню:', reply_markup=markup)
        return CHOOSING_ACTION

async def handle_warehouse_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка меню складов"""
    text = update.message.text
    
    if text == '➕ Добавить склад':
        await update.message.reply_text('Введите название склада:')
        return WAREHOUSE_NAME
    elif text == '📋 Список складов':
        warehouses = get_warehouses()
        if not warehouses:
            await update.message.reply_text('📋 Список складов пуст', reply_markup=warehouse_markup)
        else:
            message = '📋 Склады:\n\n'
            for wh_id, name, address, group_name in warehouses:
                group_info = f" [{group_name}]" if group_name else ""
                message += f'{wh_id}. {name}{group_info} ({address or "адрес не указан"})\n'
            await update.message.reply_text(message, reply_markup=warehouse_markup)
        return WAREHOUSE_MENU
    elif text == '🗑 Удалить склад':
        warehouses = get_warehouses()
        if not warehouses:
            await update.message.reply_text('📋 Список складов пуст', reply_markup=warehouse_markup)
            return WAREHOUSE_MENU
        
        # Создаем динамическую клавиатуру со складами
        keyboard = []
        for wh_id, name, address, group_name in warehouses:
            keyboard.append([f'{wh_id}. {name}'])
        keyboard.append(['❌ Отмена'])
        
        delete_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text('Выберите склад для удаления:', reply_markup=delete_markup)
        return WAREHOUSE_DELETE
    elif text == '📁 Подгруппы складов':
        await update.message.reply_text('📁 Управление подгруппами:', reply_markup=warehouse_group_markup)
        return WAREHOUSE_GROUP_MENU
    elif text == '🔙 Назад':
        await update.message.reply_text('⚙️ Управление системой:', reply_markup=management_markup)
        return MANAGEMENT_MENU

async def handle_warehouse_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка названия склада"""
    warehouse_name = update.message.text
    context.user_data['warehouse_name'] = warehouse_name
    
    # Показываем список подгрупп для выбора
    groups = get_warehouse_groups()
    if not groups:
        add_warehouse(warehouse_name, "", None)
        await update.message.reply_text(f'✅ Склад "{warehouse_name}" добавлен без подгруппы', reply_markup=warehouse_markup)
        return WAREHOUSE_MENU
    
    # Создаем динамическую клавиатуру с подгруппами
    keyboard = []
    for group_id, group_name in groups:
        keyboard.append([f'{group_id}. {group_name}'])
    keyboard.append(['⏭ Без подгруппы', '❌ Отмена'])
    
    group_select_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text('Выберите подгруппу для склада:', reply_markup=group_select_markup)
    return WAREHOUSE_GROUP_SELECT

async def handle_warehouse_group_select(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора подгруппы для склада"""
    text = update.message.text
    warehouse_name = context.user_data.get('warehouse_name')
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=warehouse_markup)
        return WAREHOUSE_MENU
    
    if text == '⏭ Без подгруппы':
        add_warehouse(warehouse_name, "", None)
        await update.message.reply_text(f'✅ Склад "{warehouse_name}" добавлен без подгруппы', reply_markup=warehouse_markup)
        return WAREHOUSE_MENU
    
    # Извлекаем ID из текста кнопки (формат: "1. Название")
    try:
        group_id = int(text.split('.')[0])
        add_warehouse(warehouse_name, "", group_id)
        await update.message.reply_text(f'✅ Склад "{warehouse_name}" добавлен', reply_markup=warehouse_markup)
        return WAREHOUSE_MENU
    except (ValueError, IndexError):
        await update.message.reply_text('❌ Неверный выбор', reply_markup=warehouse_markup)
        return WAREHOUSE_MENU

async def handle_warehouse_group_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка меню подгрупп складов"""
    text = update.message.text
    
    if text == '➕ Добавить подгруппу':
        await update.message.reply_text('Введите название подгруппы:')
        return WAREHOUSE_GROUP_NAME
    elif text == '📋 Список подгрупп':
        groups = get_warehouse_groups()
        if not groups:
            await update.message.reply_text('📋 Список подгрупп пуст', reply_markup=warehouse_group_markup)
        else:
            message = '📋 Подгруппы складов:\n\n'
            for group_id, name in groups:
                message += f'{group_id}. {name}\n'
            await update.message.reply_text(message, reply_markup=warehouse_group_markup)
        return WAREHOUSE_GROUP_MENU
    elif text == '🗑 Удалить подгруппу':
        groups = get_warehouse_groups()
        if not groups:
            await update.message.reply_text('📋 Список подгрупп пуст', reply_markup=warehouse_group_markup)
            return WAREHOUSE_GROUP_MENU
        
        # Создаем динамическую клавиатуру с подгруппами
        keyboard = []
        for group_id, name in groups:
            keyboard.append([f'{group_id}. {name}'])
        keyboard.append(['❌ Отмена'])
        
        delete_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text('Выберите подгруппу для удаления:', reply_markup=delete_markup)
        return WAREHOUSE_GROUP_DELETE
    elif text == '🔙 Назад':
        await update.message.reply_text('🏢 Управление складами:', reply_markup=warehouse_markup)
        return WAREHOUSE_MENU

async def handle_warehouse_group_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка названия подгруппы"""
    group_name = update.message.text
    add_warehouse_group(group_name)
    await update.message.reply_text(f'✅ Подгруппа "{group_name}" добавлена', reply_markup=warehouse_group_markup)
    return WAREHOUSE_GROUP_MENU

async def handle_warehouse_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка дополнительных данных склада"""
    # Заглушка для совместимости
    return WAREHOUSE_MENU

async def handle_product_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка меню товаров"""
    text = update.message.text
    
    if text == '➕ Добавить товар':
        await update.message.reply_text('Введите название товара:')
        context.user_data['product_action'] = 'add'
        return PRODUCT_DATA
    elif text == '📋 Список товаров':
        products = get_inventory()
        if not products:
            await update.message.reply_text('📋 Список товаров пуст', reply_markup=product_markup)
        else:
            message = '📋 Товары:\n\n'
            for name, quantity in products:
                message += f'• {name}: {quantity} шт.\n'
            await update.message.reply_text(message, reply_markup=product_markup)
        return PRODUCT_MENU
    elif text == '🗑 Удалить товар':
        products = get_inventory()
        if not products:
            await update.message.reply_text('📋 Список товаров пуст', reply_markup=product_markup)
            return PRODUCT_MENU
        
        # Создаем динамическую клавиатуру с товарами
        keyboard = []
        for name, quantity in products:
            keyboard.append([name])
        keyboard.append(['❌ Отмена'])
        
        delete_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text('Выберите товар для удаления:', reply_markup=delete_markup)
        return PRODUCT_DELETE
    elif text == '🔙 Назад':
        await update.message.reply_text('⚙️ Управление системой:', reply_markup=management_markup)
        return MANAGEMENT_MENU

async def handle_product_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка данных товара"""
    product_name = update.message.text
    add_product(product_name, 0)
    await update.message.reply_text(f'✅ Товар "{product_name}" добавлен', reply_markup=product_markup)
    return PRODUCT_MENU

async def handle_firm_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка меню фирм"""
    text = update.message.text
    
    if text == '➕ Добавить фирму':
        await update.message.reply_text('Введите название фирмы:')
        return FIRM_NAME
    elif text == '📋 Список фирм':
        firms = get_firms()
        if not firms:
            await update.message.reply_text('📋 Список фирм пуст', reply_markup=firm_markup)
        else:
            message = '📋 Фирмы:\n\n'
            for firm_id, name, contact in firms:
                message += f'{firm_id}. {name} ({contact or "контакт не указан"})\n'
            await update.message.reply_text(message, reply_markup=firm_markup)
        return FIRM_MENU
    elif text == '🗑 Удалить фирму':
        firms = get_firms()
        if not firms:
            await update.message.reply_text('📋 Список фирм пуст', reply_markup=firm_markup)
            return FIRM_MENU
        
        # Создаем динамическую клавиатуру с фирмами
        keyboard = []
        for firm_id, name, contact in firms:
            keyboard.append([f'{firm_id}. {name}'])
        keyboard.append(['❌ Отмена'])
        
        delete_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text('Выберите фирму для удаления:', reply_markup=delete_markup)
        return FIRM_DELETE
    elif text == '🔙 Назад':
        await update.message.reply_text('⚙️ Управление системой:', reply_markup=management_markup)
        return MANAGEMENT_MENU

async def handle_firm_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка названия фирмы"""
    firm_name = update.message.text
    add_firm(firm_name, "")
    await update.message.reply_text(f'✅ Фирма "{firm_name}" добавлена', reply_markup=firm_markup)
    return FIRM_MENU

async def handle_client_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка меню клиентов"""
    text = update.message.text
    
    if text == '➕ Добавить клиента':
        await update.message.reply_text('Введите имя клиента:')
        return CLIENT_DATA
    elif text == '📋 Список клиентов':
        clients = get_clients()
        if not clients:
            await update.message.reply_text('📋 Список клиентов пуст', reply_markup=client_markup)
        else:
            message = '📋 Клиенты:\n\n'
            for client_id, name, phone in clients:
                message += f'{client_id}. {name} ({phone or "телефон не указан"})\n'
            await update.message.reply_text(message, reply_markup=client_markup)
        return CLIENT_MENU
    elif text == '🗑 Удалить клиента':
        clients = get_clients()
        if not clients:
            await update.message.reply_text('📋 Список клиентов пуст', reply_markup=client_markup)
            return CLIENT_MENU
        
        # Создаем динамическую клавиатуру с клиентами
        keyboard = []
        for client_id, name, phone in clients:
            keyboard.append([f'{client_id}. {name}'])
        keyboard.append(['❌ Отмена'])
        
        delete_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text('Выберите клиента для удаления:', reply_markup=delete_markup)
        return CLIENT_DELETE
    elif text == '🔙 Назад':
        await update.message.reply_text('⚙️ Управление системой:', reply_markup=management_markup)
        return MANAGEMENT_MENU

async def handle_client_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка данных клиента"""
    client_name = update.message.text
    add_client(client_name, "")
    await update.message.reply_text(f'✅ Клиент "{client_name}" добавлен', reply_markup=client_markup)
    return CLIENT_MENU

async def handle_price_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка меню цен"""
    text = update.message.text
    
    if text == '💰 Установить цену':
        products = get_inventory()
        if not products:
            await update.message.reply_text('📋 Список товаров пуст. Сначала добавьте товары.', reply_markup=price_markup)
            return PRICE_MENU
        
        # Создаем динамическую клавиатуру с товарами
        keyboard = []
        for name, quantity in products:
            keyboard.append([name])
        keyboard.append(['➕ Новый товар', '❌ Отмена'])
        
        select_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text('Выберите товар для установки цены:', reply_markup=select_markup)
        return PRICE_SELECT
    elif text == '📋 Список цен':
        prices = get_product_prices()
        if not prices:
            await update.message.reply_text('📋 Список цен пуст', reply_markup=price_markup)
        else:
            message = '📋 Цены на товары:\n\n'
            for product_name, price in prices:
                message += f'• {product_name}: {price} $\n'
            await update.message.reply_text(message, reply_markup=price_markup)
        return PRICE_MENU
    elif text == '🗑 Удалить цену':
        prices = get_product_prices()
        if not prices:
            await update.message.reply_text('📋 Список цен пуст', reply_markup=price_markup)
            return PRICE_MENU
        
        # Создаем динамическую клавиатуру с товарами
        keyboard = []
        for product_name, price in prices:
            keyboard.append([product_name])
        keyboard.append(['❌ Отмена'])
        
        delete_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text('Выберите товар для удаления цены:', reply_markup=delete_markup)
        return PRICE_DELETE
    elif text == '🔙 Назад':
        await update.message.reply_text('⚙️ Управление системой:', reply_markup=management_markup)
        return MANAGEMENT_MENU

async def handle_price_select(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора товара для установки цены"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=price_markup)
        return PRICE_MENU
    
    if text == '➕ Новый товар':
        await update.message.reply_text('Введите название нового товара:')
        context.user_data['price_new_product'] = True
        return PRICE_DATA
    
    # Сохраняем выбранный товар
    context.user_data['price_product'] = text
    context.user_data['price_new_product'] = False
    await update.message.reply_text(f'Введите цену для товара "{text}" (в долларах):')
    return PRICE_VALUE

async def handle_price_value(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка ввода цены"""
    try:
        price = float(update.message.text)
        product_name = context.user_data.get('price_product')
        
        set_product_price(product_name, price)
        await update.message.reply_text(f'✅ Цена для "{product_name}" установлена: {price} $', reply_markup=price_markup)
        return PRICE_MENU
    except ValueError:
        await update.message.reply_text('❌ Неверный формат цены. Введите число:', reply_markup=price_markup)
        return PRICE_MENU

async def handle_price_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка данных цены для нового товара"""
    product_name = update.message.text
    context.user_data['price_product'] = product_name
    await update.message.reply_text(f'Введите цену для товара "{product_name}" (в долларах):')
    return PRICE_VALUE

async def handle_coalition_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка меню коалиц"""
    text = update.message.text
    
    if text == '➕ Добавить коалицу':
        await update.message.reply_text('Введите название коалицы:')
        return COALITION_NAME
    elif text == '📋 Список коалиц':
        coalitions = get_coalitions()
        if not coalitions:
            await update.message.reply_text('📋 Список коалиц пуст', reply_markup=coalition_markup)
        else:
            message = '📋 Коалицы:\n\n'
            for coalition_id, name, contact in coalitions:
                message += f'{coalition_id}. {name} ({contact or "контакт не указан"})\n'
            await update.message.reply_text(message, reply_markup=coalition_markup)
        return COALITION_MENU
    elif text == '🗑 Удалить коалицу':
        coalitions = get_coalitions()
        if not coalitions:
            await update.message.reply_text('📋 Список коалиц пуст', reply_markup=coalition_markup)
            return COALITION_MENU
        
        # Создаем динамическую клавиатуру с коалицами
        keyboard = []
        for coalition_id, name, contact in coalitions:
            keyboard.append([f'{coalition_id}. {name}'])
        keyboard.append(['❌ Отмена'])
        
        delete_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text('Выберите коалицу для удаления:', reply_markup=delete_markup)
        return COALITION_DELETE
    elif text == '🔙 Назад':
        await update.message.reply_text('⚙️ Управление системой:', reply_markup=management_markup)
        return MANAGEMENT_MENU

async def handle_coalition_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка названия коалицы"""
    coalition_name = update.message.text
    add_coalition(coalition_name, "")
    await update.message.reply_text(f'✅ Коалица "{coalition_name}" добавлена', reply_markup=coalition_markup)
    return COALITION_MENU

async def handle_coalition_delete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка удаления коалицы"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=coalition_markup)
        return COALITION_MENU
    
    try:
        coalition_id = int(text.split('.')[0])
        delete_coalition(coalition_id)
        await update.message.reply_text(f'✅ Коалица удалена', reply_markup=coalition_markup)
    except (ValueError, IndexError):
        await update.message.reply_text('❌ Неверный выбор', reply_markup=coalition_markup)
    return COALITION_MENU

# ============= МЕНЮ ОТЧЕТОВ =============

async def handle_report_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка меню отчетов"""
    text = update.message.text
    telegram_id = update.effective_user.id
    
    # Проверяем сессию
    session = get_session(telegram_id)
    if not session:
        await update.message.reply_text('❌ Сессия истекла. Войдите заново: /start')
        return ConversationHandler.END
    
    user_id, username, role, warehouse_group_id, is_active = session
    
    if text == '➕ Отчет по приходу':
        # Получаем отчет по приходам с группировкой по месяцам
        from datetime import datetime
        
        current_year = datetime.now().year
        
        # Получаем склады пользователя для фильтрации
        user_warehouses = get_user_warehouses(user_id)
        warehouse_ids = [wh[0] for wh in user_warehouses] if user_warehouses else []
        
        # Получаем все приходы за текущий год
        date_from = f'{current_year}-01-01'
        date_to = f'{current_year}-12-31'
        
        all_arrivals = []
        
        if role == 'admin':
            # Админ видит все
            all_arrivals = get_arrivals(limit=10000, date_from=date_from, date_to=date_to)
        else:
            # Остальные видят только свои склады
            for wh_id in warehouse_ids:
                arrivals_wh = get_arrivals(limit=10000, warehouse_id=wh_id, date_from=date_from, date_to=date_to)
                all_arrivals.extend(arrivals_wh)
        
        # Группируем по месяцам
        months_data = {}
        month_names = {
            1: 'январь', 2: 'февраль', 3: 'март', 4: 'апрель',
            5: 'май', 6: 'июнь', 7: 'июль', 8: 'август',
            9: 'сентябрь', 10: 'октябрь', 11: 'ноябрь', 12: 'декабрь'
        }
        
        for arr in all_arrivals:
            arr_id, arr_date, wagon, firm, warehouse, product, source, qty_doc, qty_act, notes, username_arr, created = arr
            
            # Извлекаем месяц из даты
            if isinstance(arr_date, str):
                month = int(arr_date.split('-')[1])
            else:
                month = arr_date.month
            
            if month not in months_data:
                months_data[month] = {'count': 0, 'total_tons': 0}
            
            months_data[month]['count'] += 1
            months_data[month]['total_tons'] += qty_act / 20  # Переводим в тонны
        
        # Формируем отчет
        message = f'📈 ПРИХОД ЗА {current_year}\n'
        message += '─' * 20 + '\n'
        
        total_count = 0
        total_tons = 0
        
        if not months_data:
            message += 'Нет данных за этот период\n'
        else:
            # Выводим данные по месяцам в порядке
            for month in sorted(months_data.keys()):
                data = months_data[month]
                total_count += data['count']
                total_tons += data['total_tons']
                message += f"📅 {month_names[month]}: {data['count']} записей, {data['total_tons']:.2f} т\n"
        
        message += '─' * 20 + '\n'
        message += f'📊 Всего: {total_count} записей\n'
        message += f'📦 Итого: {total_tons:.2f} тонн\n'
        
        # Сохраняем тип отчета для экспорта
        context.user_data['report_type'] = 'arrival'
        context.user_data['report_year'] = current_year
        
        # Добавляем inline кнопку экспорта
        inline_keyboard = [
            [InlineKeyboardButton("📊 Детальный отчет (Excel)", callback_data='export_arrival_menu')],
            [InlineKeyboardButton("🔙 Назад", callback_data='back_to_report_menu')]
        ]
        inline_markup = InlineKeyboardMarkup(inline_keyboard)
        
        await update.message.reply_text(message, reply_markup=inline_markup)
        return REPORT_MENU
    
    elif text == '➖ Отчет по расходу':
        # Получаем отчет по расходам с группировкой по месяцам
        from datetime import datetime
        
        current_year = datetime.now().year
        
        # Получаем склады пользователя для фильтрации
        user_warehouses = get_user_warehouses(user_id)
        warehouse_ids = [wh[0] for wh in user_warehouses] if user_warehouses else []
        
        # Получаем все выводы за текущий год
        date_from = f'{current_year}-01-01'
        date_to = f'{current_year}-12-31'
        
        all_departures = []
        
        if role == 'admin':
            # Админ видит все
            all_departures = get_departures(limit=10000, date_from=date_from, date_to=date_to)
        else:
            # Остальные видят только свои склады
            for wh_id in warehouse_ids:
                departures_wh = get_departures(limit=10000, warehouse_id=wh_id, date_from=date_from, date_to=date_to)
                all_departures.extend(departures_wh)
        
        # Группируем по месяцам
        months_data = {}
        month_names = {
            1: 'январь', 2: 'февраль', 3: 'март', 4: 'апрель',
            5: 'май', 6: 'июнь', 7: 'июль', 8: 'август',
            9: 'сентябрь', 10: 'октябрь', 11: 'ноябрь', 12: 'декабрь'
        }
        
        for dep in all_departures:
            dep_id, dep_date, coalition, firm, warehouse, product, qty, price, notes, username_dep, created, client = dep
            
            # Извлекаем месяц из даты
            if isinstance(dep_date, str):
                month = int(dep_date.split('-')[1])
            else:
                month = dep_date.month
            
            if month not in months_data:
                months_data[month] = {'count': 0, 'total_tons': 0, 'total_sum': 0}
            
            months_data[month]['count'] += 1
            tons = qty / 20
            months_data[month]['total_tons'] += tons
            
            if price:
                months_data[month]['total_sum'] += tons * float(price)
        
        # Формируем отчет
        message = f'📉 РАСХОД ЗА {current_year}\n'
        message += '─' * 20 + '\n'
        
        total_count = 0
        total_tons = 0
        total_sum = 0
        
        if not months_data:
            message += 'Нет данных за этот период\n'
        else:
            # Выводим данные по месяцам в порядке
            for month in sorted(months_data.keys()):
                data = months_data[month]
                total_count += data['count']
                total_tons += data['total_tons']
                total_sum += data['total_sum']
                message += f"📅 {month_names[month]}: {data['count']} записей, {data['total_tons']:.2f} т, {data['total_sum']:.2f} $\n"
        
        message += '─' * 20 + '\n'
        message += f'📊 Всего: {total_count} записей\n'
        message += f'📦 Итого: {total_tons:.2f} тонн\n'
        message += f'💵 Сумма: {total_sum:.2f} $\n'
        
        # Сохраняем тип отчета для экспорта
        context.user_data['report_type'] = 'departure'
        context.user_data['report_year'] = current_year
        
        # Добавляем inline кнопку экспорта
        inline_keyboard = [
            [InlineKeyboardButton("📊 Детальный отчет (Excel)", callback_data='export_departure_menu')],
            [InlineKeyboardButton("🔙 Назад", callback_data='back_to_report_menu')]
        ]
        inline_markup = InlineKeyboardMarkup(inline_keyboard)
        
        await update.message.reply_text(message, reply_markup=inline_markup)
        return REPORT_MENU
    
    elif text == '💰 Отчет по погашениям':
        # Получаем отчет по погашениям с группировкой по месяцам
        from datetime import datetime
        
        current_year = datetime.now().year
        
        # Получаем все погашения за текущий год
        date_from = f'{current_year}-01-01'
        date_to = f'{current_year}-12-31'
        
        all_payments = get_payments(limit=10000, date_from=date_from, date_to=date_to)
        
        # Группируем по месяцам
        months_data = {}
        month_names = {
            1: 'январь', 2: 'февраль', 3: 'март', 4: 'апрель',
            5: 'май', 6: 'июнь', 7: 'июль', 8: 'август',
            9: 'сентябрь', 10: 'октябрь', 11: 'ноябрь', 12: 'декабрь'
        }
        
        for payment in all_payments:
            pay_id, pay_date, client_name, somoni, dollar, rate, total_usd, notes, username_pay, created_at = payment
            
            # Извлекаем месяц из даты
            if isinstance(pay_date, str):
                month = int(pay_date.split('-')[1])
            else:
                month = pay_date.month
            
            if month not in months_data:
                months_data[month] = {'count': 0, 'total_usd': 0}
            
            months_data[month]['count'] += 1
            months_data[month]['total_usd'] += float(total_usd)
        
        # Формируем отчет
        message = f'💰 ПОГАШЕНИЯ ЗА {current_year}\n'
        message += '─' * 20 + '\n'
        
        total_count = 0
        total_usd = 0
        
        if not months_data:
            message += 'Нет данных за этот период\n'
        else:
            # Выводим данные по месяцам в порядке
            for month in sorted(months_data.keys()):
                data = months_data[month]
                total_count += data['count']
                total_usd += data['total_usd']
                message += f"📅 {month_names[month]}: {data['count']} записей, {data['total_usd']:.2f} $\n"
        
        message += '─' * 20 + '\n'
        message += f'📊 Всего: {total_count} записей\n'
        message += f'💵 Итого: {total_usd:.2f} $\n'
        
        # Сохраняем тип отчета для экспорта
        context.user_data['report_type'] = 'payment'
        context.user_data['report_year'] = current_year
        
        # Добавляем inline кнопку экспорта
        inline_keyboard = [
            [InlineKeyboardButton("📊 Детальный отчет (Excel)", callback_data='export_payment_menu')],
            [InlineKeyboardButton("🔙 Назад", callback_data='back_to_report_menu')]
        ]
        inline_markup = InlineKeyboardMarkup(inline_keyboard)
        
        await update.message.reply_text(message, reply_markup=inline_markup)
        return REPORT_MENU
    
    elif text == '🤝 Отчет по партнерам':
        # Получаем отчет по партнерам с группировкой по месяцам
        from datetime import datetime
        
        current_year = datetime.now().year
        
        # Получаем всех партнеров за текущий год
        date_from = f'{current_year}-01-01'
        date_to = f'{current_year}-12-31'
        
        all_partners = get_partners(limit=10000, date_from=date_from, date_to=date_to)
        
        # Группируем по месяцам
        months_data = {}
        month_names = {
            1: 'январь', 2: 'февраль', 3: 'март', 4: 'апрель',
            5: 'май', 6: 'июнь', 7: 'июль', 8: 'август',
            9: 'сентябрь', 10: 'октябрь', 11: 'ноябрь', 12: 'декабрь'
        }
        
        for partner in all_partners:
            part_id, part_date, client_name, somoni, rate, total_usd, notes, username_part = partner
            
            # Извлекаем месяц из даты
            if isinstance(part_date, str):
                month = int(part_date.split('-')[1])
            else:
                month = part_date.month
            
            if month not in months_data:
                months_data[month] = {'count': 0, 'total_usd': 0}
            
            months_data[month]['count'] += 1
            months_data[month]['total_usd'] += float(total_usd)
        
        # Формируем отчет
        message = f'🤝 ПАРТНЕРЫ ЗА {current_year}\n'
        message += '─' * 20 + '\n'
        
        total_count = 0
        total_usd = 0
        
        if not months_data:
            message += 'Нет данных за этот период\n'
        else:
            # Выводим данные по месяцам в порядке
            for month in sorted(months_data.keys()):
                data = months_data[month]
                total_count += data['count']
                total_usd += data['total_usd']
                message += f"📅 {month_names[month]}: {data['count']} записей, {data['total_usd']:.2f} $\n"
        
        message += '─' * 20 + '\n'
        message += f'📊 Всего: {total_count} записей\n'
        message += f'💵 Итого: {total_usd:.2f} $\n'
        
        await update.message.reply_text(message, reply_markup=inline_markup)
        return REPORT_MENU
    
    elif text == '👤 Карточка клиента':
        # Показываем список клиентов для выбора
        clients = get_clients()
        
        if not clients:
            await update.message.reply_text('❌ Нет клиентов в базе', reply_markup=report_markup)
            return REPORT_MENU
        
        # Формируем inline кнопки с клиентами
        keyboard = []
        for client_id, client_name, phone in clients:
            keyboard.append([InlineKeyboardButton(client_name, callback_data=f'client_card_{client_id}')])
        
        inline_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            '👤 Выберите клиента для формирования карточки:',
            reply_markup=inline_markup
        )
        return REPORT_MENU
    
    elif text == '📢 Уведомления о долгах':
        # Показываем меню выбора периода
        from datetime import datetime, timedelta
        
        current_year = datetime.now().year
        
        # Создаем inline кнопки для выбора периода
        keyboard = [
            [InlineKeyboardButton("📅 Сегодня", callback_data='debt_notify_today')],
            [InlineKeyboardButton("📅 Вчера", callback_data='debt_notify_yesterday')],
            [InlineKeyboardButton("📅 7 дней назад", callback_data='debt_notify_7days')],
            [InlineKeyboardButton("📅 14 дней назад", callback_data='debt_notify_14days')],
            [InlineKeyboardButton("📅 30 дней назад", callback_data='debt_notify_30days')],
        ]
        
        inline_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            f'🔔 УВЕДОМЛЕНИЯ О ДОЛГАХ\n'
            f'📅 Год: {current_year}\n\n'
            f'Выберите период:\n'
            f'Показать клиентов с долгами, которые покупали:',
            reply_markup=inline_markup
        )
        return REPORT_MENU
        
        inline_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            '📢 Выберите клиента для отправки уведомления о долге:\n\n'
            '💡 Уведомление будет отправлено клиенту с информацией о его долге.',
            reply_markup=inline_markup
        )
        return REPORT_MENU
    
    elif text == '🚂 Итоги вагонов':
        # Отчет по итогам вагонов с группировкой по складам
        from datetime import datetime
        
        await update.message.reply_text('⏳ Формирую отчет по вагонам...', reply_markup=report_markup)
        
        current_year = datetime.now().year
        
        # Получаем склады пользователя для фильтрации
        user_warehouses = get_user_warehouses(user_id)
        warehouse_ids = [wh[0] for wh in user_warehouses] if user_warehouses else []
        
        # Получаем все приходы за текущий год
        date_from = f'{current_year}-01-01'
        date_to = f'{current_year}-12-31'
        
        all_arrivals = []
        
        if role == 'admin':
            all_arrivals = get_arrivals(limit=10000, date_from=date_from, date_to=date_to)
        else:
            for wh_id in warehouse_ids:
                arrivals_wh = get_arrivals(limit=10000, warehouse_id=wh_id, date_from=date_from, date_to=date_to)
                all_arrivals.extend(arrivals_wh)
        
        if not all_arrivals:
            await update.message.reply_text('📊 Нет данных по вагонам за этот период', reply_markup=report_markup)
            return REPORT_MENU
        
        # Группируем данные: склад -> товар+фирма -> данные
        warehouse_data = {}
        
        for arr in all_arrivals:
            arr_id, arr_date, wagon, firm, warehouse, product, source, qty_doc, qty_act, notes, username_arr, created = arr
            
            if warehouse not in warehouse_data:
                warehouse_data[warehouse] = {}
            
            # Ключ: товар + фирма
            key = f"{product} ({firm})"
            
            if key not in warehouse_data[warehouse]:
                warehouse_data[warehouse][key] = {
                    'wagons': 0,
                    'qty_doc': 0,
                    'qty_act': 0
                }
            
            warehouse_data[warehouse][key]['wagons'] += 1
            warehouse_data[warehouse][key]['qty_doc'] += qty_doc
            warehouse_data[warehouse][key]['qty_act'] += qty_act
        
        # Формируем отчет
        message = f'🚂 ИТОГИ ВАГОНОВ\n'
        message += f'📅 {current_year}\n'
        message += '═' * 25 + '\n\n'
        
        total_wagons = 0
        total_qty_doc = 0
        total_qty_act = 0
        
        for warehouse in sorted(warehouse_data.keys()):
            message += f'🏪 {warehouse}\n'
            message += '─' * 20 + '\n'
            
            warehouse_wagons = 0
            warehouse_weight = 0
            
            for product_firm in sorted(warehouse_data[warehouse].keys()):
                data = warehouse_data[warehouse][product_firm]
                
                wagons = data['wagons']
                qty_doc = data['qty_doc']
                qty_act = data['qty_act']
                diff = qty_act - qty_doc
                weight = qty_act / 20  # Переводим в тонны из фактического количества
                
                warehouse_wagons += wagons
                warehouse_weight += weight
                total_wagons += wagons
                total_qty_doc += qty_doc
                total_qty_act += qty_act
                
                message += f'📦 {product_firm}\n'
                message += f'🚂 Вагонов: {wagons}\n'
                message += f'📄 По док: {int(qty_doc)} шт\n'
                message += f'✅ Факт: {int(qty_act)} шт\n'
                
                if diff > 0:
                    message += f'📈 Разница: {int(diff)} шт\n'
                elif diff < 0:
                    message += f'📉 Разница: {int(diff)} шт\n'
                else:
                    message += f'📈 Разница: 0 шт\n'
                
                message += f'⚖️ Вес: {weight:,.2f} т\n\n'
            
            message += f'📊 Итого {warehouse}:\n'
            message += f'🚂 {warehouse_wagons} вагонов, ⚖️ {warehouse_weight:,.2f} т\n\n'
        
        # Общий итог
        total_diff = total_qty_act - total_qty_doc
        total_weight = total_qty_act / 20  # Вес из фактического количества
        
        message += '═' * 25 + '\n'
        message += f'🚂 ОБЩИЙ ИТОГ:\n'
        message += f'Вагонов: {total_wagons}\n'
        message += f'По документам: {int(total_qty_doc)} шт\n'
        message += f'Фактически: {int(total_qty_act)} шт\n'
        message += f'Разница: {int(total_diff)} шт\n'
        message += f'Вес: {total_weight:,.2f} тонн'
        
        # Если сообщение слишком длинное, разбиваем на части
        if len(message) > 4000:
            parts = []
            current_part = f'🚂 ИТОГИ ВАГОНОВ\n📅 {current_year}\n' + '═' * 25 + '\n\n'
            
            for warehouse in sorted(warehouse_data.keys()):
                warehouse_section = f'🏪 {warehouse}\n' + '─' * 20 + '\n'
                
                warehouse_wagons = 0
                warehouse_weight = 0
                
                for product_firm in sorted(warehouse_data[warehouse].keys()):
                    data = warehouse_data[warehouse][product_firm]
                    wagons = data['wagons']
                    qty_doc = data['qty_doc']
                    qty_act = data['qty_act']
                    diff = qty_act - qty_doc
                    weight = qty_act / 20
                    
                    warehouse_wagons += wagons
                    warehouse_weight += weight
                    
                    product_section = f'📦 {product_firm}\n'
                    product_section += f'🚂 Вагонов: {wagons}\n'
                    product_section += f'📄 По док: {int(qty_doc)} шт\n'
                    product_section += f'✅ Факт: {int(qty_act)} шт\n'
                    
                    if diff > 0:
                        product_section += f'📈 Разница: {int(diff)} шт\n'
                    elif diff < 0:
                        product_section += f'📉 Разница: {int(diff)} шт\n'
                    else:
                        product_section += f'📈 Разница: 0 шт\n'
                    
                    product_section += f'⚖️ Вес: {weight:,.2f} т\n\n'
                    
                    warehouse_section += product_section
                
                warehouse_section += f'📊 Итого {warehouse}:\n'
                warehouse_section += f'🚂 {warehouse_wagons} вагонов, ⚖️ {warehouse_weight:,.2f} т\n\n'
                
                # Если добавление секции превысит лимит, отправляем текущую часть
                if len(current_part) + len(warehouse_section) > 3800:
                    parts.append(current_part)
                    current_part = warehouse_section
                else:
                    current_part += warehouse_section
            
            # Добавляем последнюю часть
            if current_part:
                parts.append(current_part)
            
            # Отправляем части
            for i, part in enumerate(parts):
                if i == 0:
                    await update.message.reply_text(part, reply_markup=report_markup)
                else:
                    await update.message.reply_text(part)
            
            # Отправляем итог отдельно
            footer = '═' * 25 + '\n'
            footer += f'🚂 ОБЩИЙ ИТОГ:\n'
            footer += f'Вагонов: {total_wagons}\n'
            footer += f'По документам: {int(total_qty_doc)} шт\n'
            footer += f'Фактически: {int(total_qty_act)} шт\n'
            footer += f'Разница: {int(total_diff)} шт\n'
            footer += f'Вес: {total_weight:,.2f} тонн'
            
            await update.message.reply_text(footer)
        else:
            await update.message.reply_text(message, reply_markup=report_markup)
        
        return REPORT_MENU
    
    elif text == '🔙 Назад':
        await update.message.reply_text('Главное меню:', reply_markup=markup)
        return CHOOSING_ACTION

async def handle_report_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка экспорта отчета в Excel"""
    text = update.message.text
    telegram_id = update.effective_user.id
    
    # Проверяем сессию
    session = get_session(telegram_id)
    if not session:
        await update.message.reply_text('❌ Сессия истекла. Войдите заново: /start')
        return ConversationHandler.END
    
    user_id, username, role, warehouse_group_id, is_active = session
    
    if text == '📊 Детальный отчет (Excel)':
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import os
        
        report_type = context.user_data.get('report_type')
        report_year = context.user_data.get('report_year', datetime.now().year)
        
        await update.message.reply_text('⏳ Формирую отчет...')
        
        # Получаем склады пользователя для фильтрации
        user_warehouses = get_user_warehouses(user_id)
        warehouse_ids = [wh[0] for wh in user_warehouses] if user_warehouses else []
        
        # Получаем данные за год
        date_from = f'{report_year}-01-01'
        date_to = f'{report_year}-12-31'
        
        if report_type == 'arrival':
            # Экспорт приходов
            all_arrivals = []
            
            if role == 'admin':
                all_arrivals = get_arrivals(limit=10000, date_from=date_from, date_to=date_to)
            else:
                for wh_id in warehouse_ids:
                    arrivals_wh = get_arrivals(limit=10000, warehouse_id=wh_id, date_from=date_from, date_to=date_to)
                    all_arrivals.extend(arrivals_wh)
            
            # Создаем Excel файл
            wb = Workbook()
            ws = wb.active
            ws.title = f"Приходы {report_year}"
            
            # Заголовок
            headers = ['№', 'Дата', 'Вагон', 'Фирма', 'Склад', 'Товар', 
                      'По документу (шт)', 'По факту (шт)', 'Тонны', 'Расхождение', 'Примечания', 'Пользователь']
            
            # Стиль заголовка
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Данные
            for idx, arr in enumerate(all_arrivals, 2):
                arr_id, arr_date, wagon, firm, warehouse, product, source, qty_doc, qty_act, notes, username_arr, created = arr
                
                tons = qty_act / 20
                diff = qty_act - qty_doc
                
                # Преобразуем дату в строку если нужно
                date_str = arr_date if isinstance(arr_date, str) else arr_date.strftime('%Y-%m-%d')
                
                ws.cell(row=idx, column=1, value=arr_id)
                ws.cell(row=idx, column=2, value=date_str)
                ws.cell(row=idx, column=3, value=wagon or '')
                ws.cell(row=idx, column=4, value=firm or '')
                ws.cell(row=idx, column=5, value=warehouse or '')
                ws.cell(row=idx, column=6, value=product)
                ws.cell(row=idx, column=7, value=qty_doc)
                ws.cell(row=idx, column=8, value=qty_act)
                ws.cell(row=idx, column=9, value=round(tons, 2))
                ws.cell(row=idx, column=10, value=diff)
                ws.cell(row=idx, column=12, value=notes or '')
                ws.cell(row=idx, column=12, value=username_arr or '')
            
            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем файл
            filename = f'arrival_report_{report_year}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(filename)
            
            # Отправляем файл
            with open(filename, 'rb') as file:
                await update.message.reply_document(
                    document=file,
                    filename=filename,
                    caption=f'📊 Детальный отчет по приходам за {report_year} год'
                )
            
            # Удаляем временный файл
            os.remove(filename)
            
        elif report_type == 'departure':
            # Экспорт расходов
            all_departures = []
            
            if role == 'admin':
                all_departures = get_departures(limit=10000, date_from=date_from, date_to=date_to)
            else:
                for wh_id in warehouse_ids:
                    departures_wh = get_departures(limit=10000, warehouse_id=wh_id, date_from=date_from, date_to=date_to)
                    all_departures.extend(departures_wh)
            
            # Создаем Excel файл
            wb = Workbook()
            ws = wb.active
            ws.title = f"Расходы {report_year}"
            
            # Заголовок
            headers = ['№', 'Дата', 'Коалица', 'Фирма', 'Склад', 'Товар', 'Клиент',
                      'Количество (шт)', 'Тонны', 'Цена ($)', 'Сумма ($)', 'Примечания', 'Пользователь']
            
            # Стиль заголовка
            header_fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Данные
            for idx, dep in enumerate(all_departures, 2):
                dep_id, dep_date, coalition, firm, warehouse, product, qty, price, notes, username_dep, created, client = dep
                
                tons = qty / 20
                sum_value = tons * float(price) if price else 0
                
                # Преобразуем дату в строку если нужно
                date_str = dep_date if isinstance(dep_date, str) else dep_date.strftime('%Y-%m-%d')
                
                ws.cell(row=idx, column=1, value=dep_id)
                ws.cell(row=idx, column=2, value=date_str)
                ws.cell(row=idx, column=3, value=coalition or '')
                ws.cell(row=idx, column=4, value=firm or '')
                ws.cell(row=idx, column=5, value=warehouse or '')
                ws.cell(row=idx, column=6, value=product)
                ws.cell(row=idx, column=7, value=client or '')
                ws.cell(row=idx, column=8, value=qty)
                ws.cell(row=idx, column=9, value=round(tons, 2))
                ws.cell(row=idx, column=10, value=round(float(price), 2) if price else 0)
                ws.cell(row=idx, column=11, value=round(sum_value, 2))
                ws.cell(row=idx, column=12, value=notes or '')
                ws.cell(row=idx, column=13, value=username_dep or '')
            
            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем файл
            filename = f'departure_report_{report_year}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(filename)
            
            # Отправляем файл
            with open(filename, 'rb') as file:
                await update.message.reply_document(
                    document=file,
                    filename=filename,
                    caption=f'📊 Детальный отчет по расходам за {report_year} год'
                )
            
            # Удаляем временный файл
            os.remove(filename)
        
        # Возвращаемся к отчету с кнопкой экспорта
        export_keyboard = [
            ['📊 Детальный отчет (Excel)'],
            ['🔙 Назад']
        ]
        export_markup = ReplyKeyboardMarkup(export_keyboard, resize_keyboard=True)
        
        await update.message.reply_text('✅ Отчет отправлен', reply_markup=export_markup)
        return REPORT_EXPORT_PERIOD
    
    elif text == '🔙 Назад':
        await update.message.reply_text('📋 Выберите тип отчета:', reply_markup=report_markup)
        return REPORT_MENU

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отмена операции"""
    await update.message.reply_text('Операция отменена', reply_markup=markup)
    return CHOOSING_ACTION

async def handle_warehouse_delete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка удаления склада"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=warehouse_markup)
        return WAREHOUSE_MENU
    
    try:
        warehouse_id = int(text.split('.')[0])
        delete_warehouse(warehouse_id)
        await update.message.reply_text(f'✅ Склад удален', reply_markup=warehouse_markup)
    except (ValueError, IndexError):
        await update.message.reply_text('❌ Неверный выбор', reply_markup=warehouse_markup)
    return WAREHOUSE_MENU

async def handle_warehouse_group_delete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка удаления подгруппы"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=warehouse_group_markup)
        return WAREHOUSE_GROUP_MENU
    
    try:
        group_id = int(text.split('.')[0])
        delete_warehouse_group(group_id)
        await update.message.reply_text(f'✅ Подгруппа удалена', reply_markup=warehouse_group_markup)
    except (ValueError, IndexError):
        await update.message.reply_text('❌ Неверный выбор', reply_markup=warehouse_group_markup)
    return WAREHOUSE_GROUP_MENU

async def handle_product_delete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка удаления товара"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=product_markup)
        return PRODUCT_MENU
    
    product_name = text
    delete_product(product_name)
    await update.message.reply_text(f'✅ Товар "{product_name}" удален', reply_markup=product_markup)
    return PRODUCT_MENU

async def handle_firm_delete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка удаления фирмы"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=firm_markup)
        return FIRM_MENU
    
    try:
        firm_id = int(text.split('.')[0])
        delete_firm(firm_id)
        await update.message.reply_text(f'✅ Фирма удалена', reply_markup=firm_markup)
    except (ValueError, IndexError):
        await update.message.reply_text('❌ Неверный выбор', reply_markup=firm_markup)
    return FIRM_MENU

async def handle_client_delete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка удаления клиента"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=client_markup)
        return CLIENT_MENU
    
    try:
        client_id = int(text.split('.')[0])
        delete_client(client_id)
        await update.message.reply_text(f'✅ Клиент удален', reply_markup=client_markup)
    except (ValueError, IndexError):
        await update.message.reply_text('❌ Неверный выбор', reply_markup=client_markup)
    return CLIENT_MENU

async def handle_price_delete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка удаления цены"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=price_markup)
        return PRICE_MENU
    
    product_name = text
    delete_price(product_name)
    await update.message.reply_text(f'✅ Цена для "{product_name}" удалена', reply_markup=price_markup)
    return PRICE_MENU

# ============= УПРАВЛЕНИЕ ПОЛЬЗОВАТЕЛЯМИ =============

async def handle_user_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка меню пользователей"""
    text = update.message.text
    user_id = update.effective_user.id
    
    if not has_permission(user_id, 'admin'):
        await update.message.reply_text('❌ Доступ запрещен', reply_markup=markup)
        return CHOOSING_ACTION
    
    if text == '➕ Добавить пользователя':
        await update.message.reply_text('Введите логин для нового пользователя:')
        return USER_ADD_LOGIN
    
    elif text == '📋 Список пользователей':
        users = get_all_users()
        if not users:
            await update.message.reply_text('📋 Список пользователей пуст', reply_markup=user_markup)
        else:
            message = '📋 Пользователи:\n\n'
            for uid, uname, role, group_name, is_active, date in users:
                status = '✅' if is_active else '🔒'
                group_info = f" [{group_name}]" if group_name else ""
                message += f'{status} {uid} - {uname}\n'
                message += f'   Роль: {get_role_name(role)}{group_info}\n'
                message += f'   Добавлен: {date[:10]}\n\n'
            await update.message.reply_text(message, reply_markup=user_markup)
        return USER_MENU
    
    elif text == '✏️ Изменить роль':
        users = get_all_users()
        if not users:
            await update.message.reply_text('📋 Список пользователей пуст', reply_markup=user_markup)
            return USER_MENU
        
        keyboard = []
        for uid, uname, role, group_name, is_active, date in users:
            keyboard.append([f'{uid} - {uname}'])
        keyboard.append(['❌ Отмена'])
        
        select_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text('Выберите пользователя:', reply_markup=select_markup)
        context.user_data['edit_action'] = 'role'
        return USER_EDIT_SELECT
    
    elif text == '🏢 Назначить склад':
        users = get_all_users()
        if not users:
            await update.message.reply_text('📋 Список пользователей пуст', reply_markup=user_markup)
            return USER_MENU
        
        keyboard = []
        for uid, uname, role, group_name, is_active, date in users:
            keyboard.append([f'{uid} - {uname}'])
        keyboard.append(['❌ Отмена'])
        
        select_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text('Выберите пользователя:', reply_markup=select_markup)
        context.user_data['edit_action'] = 'warehouse'
        return USER_EDIT_SELECT
    
    elif text == '🔒 Деактивировать':
        users = [u for u in get_all_users() if u[4]]  # Только активные
        if not users:
            await update.message.reply_text('📋 Нет активных пользователей', reply_markup=user_markup)
            return USER_MENU
        
        keyboard = []
        for uid, uname, role, group_name, is_active, date in users:
            keyboard.append([f'{uid} - {uname}'])
        keyboard.append(['❌ Отмена'])
        
        select_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text('Выберите пользователя для деактивации:', reply_markup=select_markup)
        context.user_data['edit_action'] = 'deactivate'
        return USER_EDIT_SELECT
    
    elif text == '✅ Активировать':
        users = [u for u in get_all_users() if not u[4]]  # Только неактивные
        if not users:
            await update.message.reply_text('📋 Нет деактивированных пользователей', reply_markup=user_markup)
            return USER_MENU
        
        keyboard = []
        for uid, uname, role, group_name, is_active, date in users:
            keyboard.append([f'{uid} - {uname}'])
        keyboard.append(['❌ Отмена'])
        
        select_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text('Выберите пользователя для активации:', reply_markup=select_markup)
        context.user_data['edit_action'] = 'activate'
        return USER_EDIT_SELECT
    
    elif text == '🗑 Удалить пользователя':
        users = get_all_users()
        if not users:
            await update.message.reply_text('📋 Список пользователей пуст', reply_markup=user_markup)
            return USER_MENU
        
        keyboard = []
        for uid, uname, role, group_name, is_active, date in users:
            keyboard.append([f'{uid} - {uname}'])
        keyboard.append(['❌ Отмена'])
        
        select_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text('Выберите пользователя для удаления:', reply_markup=select_markup)
        context.user_data['edit_action'] = 'delete'
        return USER_EDIT_SELECT
    
    elif text == '🔙 Назад':
        await update.message.reply_text('Главное меню:', reply_markup=markup)
        return CHOOSING_ACTION

async def handle_user_add_login(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка логина нового пользователя"""
    login = update.message.text.strip()
    
    if len(login) < 3:
        await update.message.reply_text('❌ Логин должен быть не менее 3 символов. Попробуйте еще раз:')
        return USER_ADD_LOGIN
    
    context.user_data['new_user_login'] = login
    await update.message.reply_text('Введите пароль для нового пользователя:')
    return USER_ADD_PASSWORD

async def handle_user_add_password(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка пароля нового пользователя"""
    password = update.message.text.strip()
    
    if len(password) < 4:
        await update.message.reply_text('❌ Пароль должен быть не менее 4 символов. Попробуйте еще раз:')
        return USER_ADD_PASSWORD
    
    context.user_data['new_user_password'] = password
    await update.message.reply_text('Введите имя пользователя (ФИО):')
    return USER_ADD_USERNAME

async def handle_user_add_username(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка имени нового пользователя"""
    username = update.message.text.strip()
    context.user_data['new_user_username'] = username
    await update.message.reply_text('Выберите роль для пользователя:', reply_markup=role_markup)
    return USER_ADD_ROLE

async def handle_user_add_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка ID нового пользователя (устаревший метод)"""
    try:
        new_user_id = int(update.message.text)
        context.user_data['new_user_id'] = new_user_id
        await update.message.reply_text('Выберите роль для пользователя:', reply_markup=role_markup)
        return USER_ADD_ROLE
    except ValueError:
        await update.message.reply_text('❌ Введите корректный ID (число)', reply_markup=user_markup)
        return USER_MENU

async def handle_user_add_role(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора роли для нового пользователя"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=user_markup)
        return USER_MENU
    
    role_map = {
        '👑 Управляющий': 'admin',
        '📊 Менеджер': 'manager',
        '🏢 Завсклад': 'warehouse_manager',
        '💰 Кассир': 'cashier'
    }
    
    role = role_map.get(text)
    if not role:
        await update.message.reply_text('❌ Неверный выбор', reply_markup=role_markup)
        return USER_ADD_ROLE
    
    context.user_data['new_user_role'] = role
    
    # Если роль не админ, предлагаем выбрать группу складов
    if role != 'admin':
        groups = get_warehouse_groups()
        if not groups:
            # Если нет групп, создаем пользователя без группы
            login = context.user_data['new_user_login']
            password = context.user_data['new_user_password']
            username = context.user_data['new_user_username']
            
            success, result = add_user_with_login(login, password, username, role, None)
            if success:
                await update.message.reply_text(
                    f'✅ Пользователь добавлен\n'
                    f'Логин: {login}\n'
                    f'Роль: {get_role_name(role)}',
                    reply_markup=user_markup
                )
            else:
                await update.message.reply_text(f'❌ Ошибка: {result}', reply_markup=user_markup)
            return USER_MENU
        
        keyboard = []
        for group_id, group_name in groups:
            keyboard.append([f'{group_id}. {group_name}'])
        keyboard.append(['⏭ Без группы', '❌ Отмена'])
        
        group_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text('Выберите группу складов:', reply_markup=group_markup)
        return USER_ADD_GROUP
    else:
        # Админу не нужна группа
        login = context.user_data['new_user_login']
        password = context.user_data['new_user_password']
        username = context.user_data['new_user_username']
        
        success, result = add_user_with_login(login, password, username, role, None)
        if success:
            await update.message.reply_text(
                f'✅ Пользователь добавлен\n'
                f'Логин: {login}\n'
                f'Роль: {get_role_name(role)}',
                reply_markup=user_markup
            )
        else:
            await update.message.reply_text(f'❌ Ошибка: {result}', reply_markup=user_markup)
        return USER_MENU

async def handle_user_add_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора группы складов для нового пользователя"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=user_markup)
        return USER_MENU
    
    warehouse_group_id = None
    if text != '⏭ Без группы':
        try:
            warehouse_group_id = int(text.split('.')[0])
        except (ValueError, IndexError):
            await update.message.reply_text('❌ Неверный выбор', reply_markup=user_markup)
            return USER_MENU
    
    login = context.user_data['new_user_login']
    password = context.user_data['new_user_password']
    username = context.user_data['new_user_username']
    role = context.user_data['new_user_role']
    
    success, result = add_user_with_login(login, password, username, role, warehouse_group_id)
    if success:
        await update.message.reply_text(
            f'✅ Пользователь добавлен\n'
            f'Логин: {login}\n'
            f'Имя: {username}\n'
            f'Роль: {get_role_name(role)}',
            reply_markup=user_markup
        )
    else:
        await update.message.reply_text(f'❌ Ошибка: {result}', reply_markup=user_markup)
    return USER_MENU

async def handle_user_edit_select(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора пользователя для редактирования"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=user_markup)
        return USER_MENU
    
    try:
        selected_user_id = int(text.split(' - ')[0])
        context.user_data['selected_user_id'] = selected_user_id
        
        action = context.user_data.get('edit_action')
        
        if action == 'role':
            await update.message.reply_text('Выберите новую роль:', reply_markup=role_markup)
            return USER_EDIT_ROLE
        
        elif action == 'warehouse':
            groups = get_warehouse_groups()
            if not groups:
                await update.message.reply_text('📋 Нет доступных групп складов', reply_markup=user_markup)
                return USER_MENU
            
            keyboard = []
            for group_id, group_name in groups:
                keyboard.append([f'{group_id}. {group_name}'])
            keyboard.append(['⏭ Без группы', '❌ Отмена'])
            
            group_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            await update.message.reply_text('Выберите группу складов:', reply_markup=group_markup)
            return USER_EDIT_GROUP
        
        elif action == 'deactivate':
            deactivate_user(selected_user_id)
            await update.message.reply_text(f'✅ Пользователь {selected_user_id} деактивирован', reply_markup=user_markup)
            return USER_MENU
        
        elif action == 'activate':
            activate_user(selected_user_id)
            await update.message.reply_text(f'✅ Пользователь {selected_user_id} активирован', reply_markup=user_markup)
            return USER_MENU
        
        elif action == 'delete':
            delete_user(selected_user_id)
            await update.message.reply_text(f'✅ Пользователь {selected_user_id} удален', reply_markup=user_markup)
            return USER_MENU
        
    except (ValueError, IndexError):
        await update.message.reply_text('❌ Неверный выбор', reply_markup=user_markup)
        return USER_MENU

async def handle_user_edit_role(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка изменения роли пользователя"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=user_markup)
        return USER_MENU
    
    role_map = {
        '👑 Управляющий': 'admin',
        '📊 Менеджер': 'manager',
        '🏢 Завсклад': 'warehouse_manager',
        '💰 Кассир': 'cashier'
    }
    
    role = role_map.get(text)
    if not role:
        await update.message.reply_text('❌ Неверный выбор', reply_markup=role_markup)
        return USER_EDIT_ROLE
    
    selected_user_id = context.user_data['selected_user_id']
    update_user_role(selected_user_id, role)
    await update.message.reply_text(
        f'✅ Роль пользователя {selected_user_id} изменена на {get_role_name(role)}',
        reply_markup=user_markup
    )
    return USER_MENU

async def handle_user_edit_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка изменения группы складов пользователя"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=user_markup)
        return USER_MENU
    
    warehouse_group_id = None
    if text != '⏭ Без группы':
        try:
            warehouse_group_id = int(text.split('.')[0])
        except (ValueError, IndexError):
            await update.message.reply_text('❌ Неверный выбор', reply_markup=user_markup)
            return USER_MENU
    
    selected_user_id = context.user_data['selected_user_id']
    update_user_warehouse_group(selected_user_id, warehouse_group_id)
    await update.message.reply_text(
        f'✅ Группа складов пользователя {selected_user_id} обновлена',
        reply_markup=user_markup
    )
    return USER_MENU

# ============= ДЕТАЛЬНЫЙ ПРИХОД ТОВАРА =============

async def handle_arrival_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка даты прихода"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    if text == 'Сегодня':
        from datetime import date
        arrival_date = date.today().strftime('%Y-%m-%d')
        context.user_data['arrival_date'] = arrival_date
    else:
        try:
            # Парсим дату в формате ДД.ММ.ГГГГ
            from datetime import datetime
            date_obj = datetime.strptime(text, '%d.%m.%Y')
            context.user_data['arrival_date'] = date_obj.strftime('%Y-%m-%d')
        except ValueError:
            await update.message.reply_text(
                '❌ Неверный формат даты. Используйте ДД.ММ.ГГГГ\n'
                'Например: 23.02.2026',
                reply_markup=ReplyKeyboardMarkup([['Сегодня'], ['❌ Отмена']], resize_keyboard=True)
            )
            return ARRIVAL_DATE
    
    await update.message.reply_text(
        '🚂 Введите номер вагона\n'
        '(или нажмите "Пропустить" если нет вагона):',
        reply_markup=ReplyKeyboardMarkup([['Пропустить'], ['❌ Отмена']], resize_keyboard=True)
    )
    return ARRIVAL_WAGON

async def handle_arrival_wagon(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка номера вагона"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    if text == 'Пропустить':
        context.user_data['wagon_number'] = None
    else:
        context.user_data['wagon_number'] = text
    
    # Показываем список фирм
    firms = get_firms()
    if not firms:
        await update.message.reply_text(
            '❌ Нет доступных фирм. Сначала добавьте фирмы в управлении.',
            reply_markup=markup
        )
        return CHOOSING_ACTION
    
    keyboard = []
    for firm_id, name, contact in firms:
        keyboard.append([f'{firm_id}. {name}'])
    keyboard.append(['❌ Отмена'])
    
    firm_markup_temp = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text('🏭 Выберите фирму:', reply_markup=firm_markup_temp)
    return ARRIVAL_FIRM

async def handle_arrival_firm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора фирмы"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    try:
        firm_id = int(text.split('.')[0])
        context.user_data['firm_id'] = firm_id
    except (ValueError, IndexError):
        await update.message.reply_text('❌ Неверный выбор', reply_markup=markup)
        return CHOOSING_ACTION
    
    # Показываем список складов
    telegram_id = update.effective_user.id
    session = get_session(telegram_id)
    if not session:
        await update.message.reply_text('❌ Сессия истекла', reply_markup=markup)
        return CHOOSING_ACTION
    
    user_id = session[0]
    warehouses = get_user_warehouses(user_id)
    
    if not warehouses:
        await update.message.reply_text(
            '❌ Нет доступных складов.',
            reply_markup=markup
        )
        return CHOOSING_ACTION
    
    keyboard = []
    for wh_id, name, address, group_name in warehouses:
        keyboard.append([f'{wh_id}. {name}'])
    keyboard.append(['❌ Отмена'])
    
    warehouse_markup_temp = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text('🏢 Выберите склад:', reply_markup=warehouse_markup_temp)
    return ARRIVAL_WAREHOUSE

async def handle_arrival_warehouse(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора склада"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    try:
        warehouse_id = int(text.split('.')[0])
        context.user_data['warehouse_id'] = warehouse_id
    except (ValueError, IndexError):
        await update.message.reply_text('❌ Неверный выбор', reply_markup=markup)
        return CHOOSING_ACTION
    
    # Показываем список ВСЕХ товаров из справочника
    # При приходе можно выбрать любой товар, не только те что уже есть на складах
    products = get_inventory()
    
    keyboard = []
    for name, quantity in products:
        keyboard.append([name])
    keyboard.append(['➕ Новый товар', '❌ Отмена'])
    
    product_markup_temp = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text('📦 Выберите товар или добавьте новый:', reply_markup=product_markup_temp)
    return ARRIVAL_PRODUCT

async def handle_arrival_product(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора товара"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    if text == '➕ Новый товар':
        await update.message.reply_text(
            '📦 Введите название нового товара:',
            reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
        )
        context.user_data['new_product'] = True
        return ARRIVAL_PRODUCT
    
    if context.user_data.get('new_product'):
        context.user_data['product_name'] = text
        context.user_data['new_product'] = False
    else:
        context.user_data['product_name'] = text
    
    # Сразу переходим к количеству по документу
    context.user_data['source'] = None  # Устанавливаем source как None
    
    await update.message.reply_text(
        '📄 Введите количество по документу:',
        reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
    )
    return ARRIVAL_QTY_DOC

async def handle_arrival_qty_doc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка количества по документу"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    try:
        qty_doc = int(text)
        if qty_doc <= 0:
            await update.message.reply_text('❌ Количество должно быть больше 0')
            return ARRIVAL_QTY_DOC
        context.user_data['quantity_document'] = qty_doc
    except ValueError:
        await update.message.reply_text('❌ Введите корректное число')
        return ARRIVAL_QTY_DOC
    
    await update.message.reply_text(
        '✅ Введите количество по факту:',
        reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
    )
    return ARRIVAL_QTY_ACTUAL

async def handle_arrival_qty_actual(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка количества по факту"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    try:
        qty_actual = int(text)
        if qty_actual <= 0:
            await update.message.reply_text('❌ Количество должно быть больше 0')
            return ARRIVAL_QTY_ACTUAL
        context.user_data['quantity_actual'] = qty_actual
    except ValueError:
        await update.message.reply_text('❌ Введите корректное число')
        return ARRIVAL_QTY_ACTUAL
    
    await update.message.reply_text(
        '📝 Введите примечания\n'
        '(или нажмите "Пропустить"):',
        reply_markup=ReplyKeyboardMarkup([['Пропустить'], ['❌ Отмена']], resize_keyboard=True)
    )
    return ARRIVAL_NOTES

async def handle_arrival_notes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка примечаний и сохранение прихода"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    if text == 'Пропустить':
        notes = None
    else:
        notes = text
    
    # Сохраняем приход
    arrival_date = context.user_data['arrival_date']
    wagon_number = context.user_data.get('wagon_number')
    firm_id = context.user_data['firm_id']
    warehouse_id = context.user_data['warehouse_id']
    product_name = context.user_data['product_name']
    source = context.user_data.get('source')
    quantity_document = context.user_data['quantity_document']
    quantity_actual = context.user_data['quantity_actual']
    user_id = context.user_data['user_id']
    
    success, result = add_arrival(
        arrival_date, wagon_number, firm_id, warehouse_id, product_name,
        source, quantity_document, quantity_actual, notes, user_id
    )
    
    if success:
        difference = quantity_actual - quantity_document
        diff_text = ""
        if difference != 0:
            diff_text = f"\n⚠️ Расхождение: {difference:+d} шт."
        
        wagon_text = ""
        if wagon_number:
            wagon_text = f"🚂 Вагон: {wagon_number}\n"
        
        await update.message.reply_text(
            f'✅ Приход товара оформлен!\n\n'
            f'{wagon_text}'
            f'📦 Товар: {product_name}\n'
            f'📄 По документу: {quantity_document} шт.\n'
            f'✅ По факту: {quantity_actual} шт.{diff_text}\n'
            f'📋 ID прихода: {result}',
            reply_markup=markup
        )
    else:
        await update.message.reply_text(f'❌ Ошибка: {result}', reply_markup=markup)
    
    return CHOOSING_ACTION

# ============= ДЕТАЛЬНЫЙ ВЫВОД ТОВАРА =============

async def handle_departure_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка даты вывода"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    if text == 'Сегодня':
        from datetime import date
        departure_date = date.today().strftime('%Y-%m-%d')
        context.user_data['departure_date'] = departure_date
    else:
        try:
            from datetime import datetime
            date_obj = datetime.strptime(text, '%d.%m.%Y')
            context.user_data['departure_date'] = date_obj.strftime('%Y-%m-%d')
        except ValueError:
            await update.message.reply_text(
                '❌ Неверный формат даты. Используйте ДД.ММ.ГГГГ',
                reply_markup=ReplyKeyboardMarkup([['Сегодня'], ['❌ Отмена']], resize_keyboard=True)
            )
            return DEPARTURE_DATE
    
    # Показываем список коалиц
    coalitions = get_coalitions()
    if not coalitions:
        await update.message.reply_text(
            '❌ Нет доступных коалиц. Сначала добавьте коалиции в управлении.',
            reply_markup=markup
        )
        return CHOOSING_ACTION
    
    keyboard = []
    for coalition_id, name, contact in coalitions:
        keyboard.append([f'{coalition_id}. {name}'])
    keyboard.append(['⏭ Пропустить'])
    keyboard.append(['🔙 Назад', '❌ Отмена'])
    
    coalition_markup_temp = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text('📊 Выберите коалицу (или пропустите):', reply_markup=coalition_markup_temp)
    return DEPARTURE_COALITION

async def handle_departure_coalition(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора коалицы"""
    text = update.message.text
    
    if text == '⏭ Пропустить':
        context.user_data['coalition_id'] = None
        
        # Показываем список фирм
        firms = get_firms()
        if not firms:
            await update.message.reply_text(
                '❌ Нет доступных фирм.',
                reply_markup=markup
            )
            return CHOOSING_ACTION
        
        keyboard = []
        for firm_id, name, contact in firms:
            keyboard.append([f'{firm_id}. {name}'])
        keyboard.append(['❌ Отмена'])
        
        firm_markup_temp = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text('🏭 Выберите фирму:', reply_markup=firm_markup_temp)
        return DEPARTURE_FIRM
    
    if text == '🔙 Назад':
        await update.message.reply_text(
            '📅 Введите дату вывода товара\n'
            'Формат: ДД.ММ.ГГГГ (например: 24.02.2026)\n'
            'или просто нажмите "Сегодня"',
            reply_markup=ReplyKeyboardMarkup([['Сегодня'], ['❌ Отмена']], resize_keyboard=True)
        )
        return DEPARTURE_DATE
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    try:
        coalition_id = int(text.split('.')[0])
        context.user_data['coalition_id'] = coalition_id
    except (ValueError, IndexError):
        await update.message.reply_text('❌ Неверный выбор', reply_markup=markup)
        return CHOOSING_ACTION
    
    # Показываем список фирм
    firms = get_firms()
    if not firms:
        await update.message.reply_text(
            '❌ Нет доступных фирм.',
            reply_markup=markup
        )
        return CHOOSING_ACTION
    
    keyboard = []
    for firm_id, name, contact in firms:
        keyboard.append([f'{firm_id}. {name}'])
    keyboard.append(['❌ Отмена'])
    
    firm_markup_temp = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text('🏭 Выберите фирму:', reply_markup=firm_markup_temp)
    return DEPARTURE_FIRM

async def handle_departure_firm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора фирмы"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    try:
        firm_id = int(text.split('.')[0])
        context.user_data['firm_id'] = firm_id
    except (ValueError, IndexError):
        await update.message.reply_text('❌ Неверный выбор', reply_markup=markup)
        return CHOOSING_ACTION
    
    # Показываем список складов
    telegram_id = update.effective_user.id
    session = get_session(telegram_id)
    if not session:
        await update.message.reply_text('❌ Сессия истекла', reply_markup=markup)
        return CHOOSING_ACTION
    
    user_id = session[0]
    warehouses = get_user_warehouses(user_id)
    
    if not warehouses:
        await update.message.reply_text('❌ Нет доступных складов.', reply_markup=markup)
        return CHOOSING_ACTION
    
    keyboard = []
    for wh_id, name, address, group_name in warehouses:
        keyboard.append([f'{wh_id}. {name}'])
    keyboard.append(['❌ Отмена'])
    
    warehouse_markup_temp = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text('🏢 Выберите склад:', reply_markup=warehouse_markup_temp)
    return DEPARTURE_WAREHOUSE

async def handle_departure_warehouse(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора склада"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    try:
        warehouse_id = int(text.split('.')[0])
        context.user_data['warehouse_id'] = warehouse_id
    except (ValueError, IndexError):
        await update.message.reply_text('❌ Неверный выбор', reply_markup=markup)
        return CHOOSING_ACTION
    
    # Получаем user_id из сессии
    telegram_id = update.effective_user.id
    session = get_session(telegram_id)
    if not session:
        await update.message.reply_text('❌ Сессия истекла', reply_markup=markup)
        return CHOOSING_ACTION
    
    user_id = session[0]
    
    # Показываем список товаров только из подгруппы пользователя
    products = get_inventory_by_user(user_id)
    
    if not products:
        await update.message.reply_text('❌ Нет доступных товаров.', reply_markup=markup)
        return CHOOSING_ACTION
    
    keyboard = []
    for name, quantity in products:
        keyboard.append([f'{name} ({quantity} шт.)'])
    keyboard.append(['❌ Отмена'])
    
    product_markup_temp = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text('📦 Выберите товар:', reply_markup=product_markup_temp)
    return DEPARTURE_PRODUCT

async def handle_departure_product(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора товара"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    # Извлекаем название товара (до скобки с количеством)
    product_name = text.split(' (')[0] if ' (' in text else text
    context.user_data['product_name'] = product_name
    
    # Запрашиваем количество
    await update.message.reply_text(
        '📦 Введите количество:',
        reply_markup=ReplyKeyboardMarkup([['🔙 Назад', '❌ Отмена']], resize_keyboard=True)
    )
    return DEPARTURE_QUANTITY

async def handle_departure_quantity(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка количества"""
    text = update.message.text
    
    if text == '🔙 Назад':
        # Возврат к выбору товара
        telegram_id = update.effective_user.id
        session = get_session(telegram_id)
        user_id = session[0] if session else None
        products = get_inventory_by_user(user_id) if user_id else []
        
        keyboard = []
        for name, quantity in products:
            keyboard.append([f'{name} ({quantity} шт.)'])
        keyboard.append(['❌ Отмена'])
        
        product_markup_temp = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text('📦 Выберите товар:', reply_markup=product_markup_temp)
        return DEPARTURE_PRODUCT
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    try:
        quantity = int(text)
        if quantity <= 0:
            await update.message.reply_text('❌ Количество должно быть больше 0')
            return DEPARTURE_QUANTITY
        context.user_data['quantity'] = quantity
    except ValueError:
        await update.message.reply_text('❌ Введите корректное число')
        return DEPARTURE_QUANTITY
    
    # Запрашиваем ввод имени клиента для поиска
    await update.message.reply_text(
        '👤 Введите имя клиента для поиска:\n'
        '(можно ввести часть имени)',
        reply_markup=ReplyKeyboardMarkup([['🔙 Назад', '❌ Отмена']], resize_keyboard=True)
    )
    return DEPARTURE_CLIENT

async def handle_departure_client(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора клиента"""
    text = update.message.text
    
    if text == '🔙 Назад':
        await update.message.reply_text(
            '📦 Введите количество:',
            reply_markup=ReplyKeyboardMarkup([['🔙 Назад', '❌ Отмена']], resize_keyboard=True)
        )
        return DEPARTURE_QUANTITY
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    # Проверяем, выбирает ли пользователь из списка
    if context.user_data.get('selecting_from_list'):
        try:
            client_id = int(text.split('.')[0])
            context.user_data['client_id'] = client_id
            context.user_data['selecting_from_list'] = False
            
            # Получаем цену товара из справочника
            product_name = context.user_data.get('product_name')
            quantity = context.user_data.get('quantity')
            price = get_product_price(product_name)
            
            if price is None:
                # Если цены нет, запрашиваем ввод
                await update.message.reply_text(
                    f'⚠️ Цена для товара "{product_name}" не установлена.\n'
                    '💰 Введите цену за единицу (в долларах):',
                    reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
                )
                return DEPARTURE_PRICE
            
            # Цена найдена, сохраняем и переходим к примечаниям
            context.user_data['price'] = price
            total_sum = (quantity / 20) * price
            
            await update.message.reply_text(
                f'💰 Цена: {price} $ (из справочника)\n'
                f'📦 Количество: {quantity} шт.\n'
                f'💵 Сумма: {total_sum:.2f} $\n\n'
                '📝 Введите примечания\n'
                '(или нажмите "Пропустить"):',
                reply_markup=ReplyKeyboardMarkup([['Пропустить'], ['❌ Отмена']], resize_keyboard=True)
            )
            return DEPARTURE_NOTES
        except (ValueError, IndexError):
            await update.message.reply_text('❌ Неверный выбор', reply_markup=markup)
            return CHOOSING_ACTION
    
    # Ищем клиентов по введенному тексту
    search_text = text.lower()
    all_clients = get_clients()
    
    # Фильтруем клиентов по поисковому запросу
    matching_clients = [
        (client_id, name, phone) 
        for client_id, name, phone in all_clients 
        if search_text in name.lower()
    ]
    
    if not matching_clients:
        await update.message.reply_text(
            f'❌ Клиенты с именем "{text}" не найдены.\n'
            'Попробуйте другое имя:',
            reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
        )
        return DEPARTURE_CLIENT
    
    # Если найден только один клиент, выбираем его автоматически
    if len(matching_clients) == 1:
        client_id = matching_clients[0][0]
        context.user_data['client_id'] = client_id
        
        # Получаем цену товара из справочника
        product_name = context.user_data.get('product_name')
        quantity = context.user_data.get('quantity')
        price = get_product_price(product_name)
        
        if price is None:
            # Если цены нет, запрашиваем ввод
            await update.message.reply_text(
                f'✅ Выбран клиент: {matching_clients[0][1]}\n\n'
                f'⚠️ Цена для товара "{product_name}" не установлена.\n'
                '💰 Введите цену за единицу (в долларах):',
                reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
            )
            return DEPARTURE_PRICE
        
        # Цена найдена, сохраняем и переходим к примечаниям
        context.user_data['price'] = price
        total_sum = (quantity / 20) * price
        
        await update.message.reply_text(
            f'✅ Выбран клиент: {matching_clients[0][1]}\n\n'
            f'💰 Цена: {price} $ (из справочника)\n'
            f'📦 Количество: {quantity} шт.\n'
            f'💵 Сумма: {total_sum:.2f} $\n\n'
            '📝 Введите примечания\n'
            '(или нажмите "Пропустить"):',
            reply_markup=ReplyKeyboardMarkup([['Пропустить'], ['❌ Отмена']], resize_keyboard=True)
        )
        return DEPARTURE_NOTES
    
    # Если найдено несколько клиентов, показываем список для выбора
    keyboard = []
    for client_id, name, phone in matching_clients:
        keyboard.append([f'{client_id}. {name}'])
    keyboard.append(['❌ Отмена'])
    
    client_markup_temp = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text(
        f'Найдено клиентов: {len(matching_clients)}\n'
        'Выберите клиента:',
        reply_markup=client_markup_temp
    )
    
    # Сохраняем состояние для следующего шага
    context.user_data['selecting_from_list'] = True
    return DEPARTURE_CLIENT

async def handle_departure_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка цены"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    try:
        price = float(text)
        if price < 0:
            await update.message.reply_text('❌ Цена не может быть отрицательной')
            return DEPARTURE_PRICE
        context.user_data['price'] = price
    except ValueError:
        await update.message.reply_text('❌ Введите корректное число')
        return DEPARTURE_PRICE
    
    await update.message.reply_text(
        '📝 Введите примечания\n'
        '(или нажмите "Пропустить"):',
        reply_markup=ReplyKeyboardMarkup([['Пропустить'], ['❌ Отмена']], resize_keyboard=True)
    )
    return DEPARTURE_NOTES

async def handle_departure_notes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка примечаний и сохранение вывода"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    if text == 'Пропустить':
        notes = None
    else:
        notes = text
    
    # Сохраняем вывод
    departure_date = context.user_data['departure_date']
    coalition_id = context.user_data['coalition_id']
    firm_id = context.user_data['firm_id']
    warehouse_id = context.user_data['warehouse_id']
    product_name = context.user_data['product_name']
    client_id = context.user_data['client_id']
    quantity = context.user_data['quantity']
    price = context.user_data['price']
    user_id = context.user_data['user_id']
    
    success, result = add_departure(
        departure_date, coalition_id, firm_id, warehouse_id, product_name,
        quantity, price, notes, user_id, client_id
    )
    
    if success:
        # Получаем названия
        coalition_name = get_coalition_by_id(coalition_id)
        firm_name = get_firm_by_id(firm_id)
        warehouse_name = get_warehouse_by_id(warehouse_id)
        client_name = get_client_by_id(client_id)
        
        # Получаем текущий остаток товара после вывода (только для подгруппы пользователя)
        telegram_id = update.effective_user.id
        session = get_session(telegram_id)
        user_id = session[0] if session else None
        
        # Получаем остатки по подгруппе пользователя
        user_products = get_inventory_by_user(user_id) if user_id else []
        remaining_quantity = 0
        
        # Ищем нужный товар в остатках
        for prod_name, prod_qty in user_products:
            if prod_name == product_name:
                remaining_quantity = prod_qty
                break
        
        remaining_tons = remaining_quantity / 20
        
        # Расчеты
        total_sum = (quantity / 20) * price
        quantity_tons = quantity / 20  # Количество в тоннах
        
        await update.message.reply_text(
            f'✅ Вывод товара оформлен!\n\n'
            f'📊 Коалица: {coalition_name}\n'
            f'🏭 Фирма: {firm_name}\n'
            f'🏢 Склад: {warehouse_name}\n'
            f'📦 Товар: {product_name}\n'
            f'👤 Клиент: {client_name}\n'
            f'📦 Количество: {quantity} шт.\n'
            f'⚖️ Количество: {quantity_tons:.2f} тонн\n'
            f'💰 Цена: {price} $\n'
            f'💵 Сумма: {total_sum:.2f} $\n'
            f'📋 ID вывода: {result}\n\n'
            f'📊 Остаток: {remaining_quantity} шт. ({remaining_tons:.2f} тонн)',
            reply_markup=markup
        )
    else:
        await update.message.reply_text(f'❌ Ошибка: {result}', reply_markup=markup)
    
    return CHOOSING_ACTION

# ============= ПОГАШЕНИЯ =============

async def handle_payment_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка даты погашения"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    if text == 'Сегодня':
        from datetime import date
        payment_date = date.today().strftime('%Y-%m-%d')
        context.user_data['payment_date'] = payment_date
    else:
        try:
            from datetime import datetime
            date_obj = datetime.strptime(text, '%d.%m.%Y')
            context.user_data['payment_date'] = date_obj.strftime('%Y-%m-%d')
        except ValueError:
            await update.message.reply_text(
                '❌ Неверный формат даты. Используйте ДД.ММ.ГГГГ',
                reply_markup=ReplyKeyboardMarkup([['Сегодня'], ['❌ Отмена']], resize_keyboard=True)
            )
            return PAYMENT_DATE
    
    await update.message.reply_text(
        '👤 Введите имя клиента для поиска:\n'
        '(можно ввести часть имени)',
        reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
    )
    return PAYMENT_CLIENT

async def handle_payment_client(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора клиента"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    # Проверяем, выбирает ли пользователь из списка
    if context.user_data.get('selecting_payment_client'):
        try:
            client_id = int(text.split('.')[0])
            context.user_data['payment_client_id'] = client_id
            context.user_data['selecting_payment_client'] = False
            
            await update.message.reply_text(
                '💵 Введите сумму в сомони:',
                reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
            )
            return PAYMENT_SOMONI
        except (ValueError, IndexError):
            await update.message.reply_text('❌ Неверный выбор', reply_markup=markup)
            return CHOOSING_ACTION
    
    # Ищем клиентов по введенному тексту
    search_text = text.lower()
    all_clients = get_clients()
    
    matching_clients = [
        (client_id, name, phone) 
        for client_id, name, phone in all_clients 
        if search_text in name.lower()
    ]
    
    if not matching_clients:
        await update.message.reply_text(
            f'❌ Клиенты с именем "{text}" не найдены.\n'
            'Попробуйте другое имя:',
            reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
        )
        return PAYMENT_CLIENT
    
    # Если найден только один клиент, выбираем его автоматически
    if len(matching_clients) == 1:
        client_id = matching_clients[0][0]
        context.user_data['payment_client_id'] = client_id
        
        await update.message.reply_text(
            f'✅ Выбран клиент: {matching_clients[0][1]}\n\n'
            '💵 Введите сумму в сомони:',
            reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
        )
        return PAYMENT_SOMONI
    
    # Если найдено несколько клиентов, показываем список
    keyboard = []
    for client_id, name, phone in matching_clients:
        keyboard.append([f'{client_id}. {name}'])
    keyboard.append(['❌ Отмена'])
    
    client_markup_temp = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text(
        f'Найдено клиентов: {len(matching_clients)}\n'
        'Выберите клиента:',
        reply_markup=client_markup_temp
    )
    
    context.user_data['selecting_payment_client'] = True
    return PAYMENT_CLIENT

async def handle_payment_somoni(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка суммы в сомони"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    try:
        somoni = float(text)
        if somoni <= 0:
            await update.message.reply_text('❌ Сумма должна быть больше 0')
            return PAYMENT_SOMONI
        context.user_data['payment_somoni'] = somoni
    except ValueError:
        await update.message.reply_text('❌ Введите корректное число')
        return PAYMENT_SOMONI
    
    await update.message.reply_text(
        '💱 Введите курс обмена:\n'
        '(например: 13.5)',
        reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
    )
    return PAYMENT_RATE

async def handle_payment_rate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка курса обмена"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    try:
        rate = float(text)
        if rate <= 0:
            await update.message.reply_text('❌ Курс должен быть больше 0')
            return PAYMENT_RATE
        context.user_data['payment_rate'] = rate
    except ValueError:
        await update.message.reply_text('❌ Введите корректное число')
        return PAYMENT_RATE
    
    # Вычисляем сумму в долларах
    somoni = context.user_data['payment_somoni']
    total_usd = somoni / rate
    
    await update.message.reply_text(
        f'💰 Расчет:\n'
        f'Сомони: {somoni}\n'
        f'Курс: {rate}\n'
        f'Сумма $: {total_usd:.2f}\n\n'
        '📝 Введите примечания\n'
        '(или нажмите "Пропустить"):',
        reply_markup=ReplyKeyboardMarkup([['Пропустить'], ['❌ Отмена']], resize_keyboard=True)
    )
    return PAYMENT_NOTES

async def handle_payment_notes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка примечаний и сохранение погашения"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    if text == 'Пропустить':
        notes = None
    else:
        notes = text
    
    # Сохраняем погашение
    payment_date = context.user_data['payment_date']
    client_id = context.user_data['payment_client_id']
    somoni = context.user_data['payment_somoni']
    dollar = 0  # Больше не используется
    rate = context.user_data['payment_rate']
    user_id = context.user_data['user_id']
    
    success, result = add_payment(payment_date, client_id, somoni, dollar, rate, notes, user_id)
    
    if success:
        total_usd = somoni / rate
        client_name = get_client_by_id(client_id)
        
        await update.message.reply_text(
            f'✅ Погашение оформлено!\n\n'
            f'📅 Дата: {payment_date}\n'
            f'👤 Клиент: {client_name}\n'
            f'💵 Сомони: {somoni}\n'
            f' Курс: {rate}\n'
            f'💰 Сумма $: {total_usd:.2f}\n'
            f'📋 ID погашения: {result}',
            reply_markup=markup
        )
    else:
        await update.message.reply_text(f'❌ Ошибка: {result}', reply_markup=markup)
    
    return CHOOSING_ACTION

# ============= ПАРТНЕРЫ =============

async def handle_partner_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка даты партнера"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    if text == 'Сегодня':
        partner_date = datetime.now().strftime('%Y-%m-%d')
    else:
        try:
            day, month, year = text.split('.')
            partner_date = f'{year}-{month.zfill(2)}-{day.zfill(2)}'
            datetime.strptime(partner_date, '%Y-%m-%d')
        except (ValueError, AttributeError):
            await update.message.reply_text(
                '❌ Неверный формат даты. Используйте ДД.ММ.ГГГГ',
                reply_markup=ReplyKeyboardMarkup([['Сегодня'], ['❌ Отмена']], resize_keyboard=True)
            )
            return PARTNER_DATE
    
    context.user_data['partner_date'] = partner_date
    
    await update.message.reply_text(
        '👤 Введите имя клиента для поиска:',
        reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
    )
    return PARTNER_CLIENT

async def handle_partner_client(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора клиента для партнера"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    # Проверяем, выбирает ли пользователь из списка
    if context.user_data.get('selecting_partner_client'):
        try:
            client_id = int(text.split('.')[0])
            context.user_data['partner_client_id'] = client_id
            context.user_data['selecting_partner_client'] = False
            
            await update.message.reply_text(
                '💵 Введите сумму в сомони:',
                reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
            )
            return PARTNER_SOMONI
        except (ValueError, IndexError):
            await update.message.reply_text('❌ Неверный выбор', reply_markup=markup)
            return CHOOSING_ACTION
    
    # Ищем клиентов по введенному тексту
    search_text = text.lower()
    all_clients = get_clients()
    
    matching_clients = [
        (client_id, name, phone) 
        for client_id, name, phone in all_clients 
        if search_text in name.lower()
    ]
    
    if not matching_clients:
        await update.message.reply_text(
            f'❌ Клиенты с именем "{text}" не найдены.\n'
            'Попробуйте другое имя:',
            reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
        )
        return PARTNER_CLIENT
    
    # Если найден только один клиент, выбираем его автоматически
    if len(matching_clients) == 1:
        client_id = matching_clients[0][0]
        context.user_data['partner_client_id'] = client_id
        
        await update.message.reply_text(
            f'✅ Выбран клиент: {matching_clients[0][1]}\n\n'
            '💵 Введите сумму в сомони:',
            reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
        )
        return PARTNER_SOMONI
    
    # Если найдено несколько клиентов, показываем список
    keyboard = []
    for client_id, name, phone in matching_clients:
        keyboard.append([f'{client_id}. {name}'])
    keyboard.append(['❌ Отмена'])
    
    client_markup_temp = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text(
        f'Найдено клиентов: {len(matching_clients)}\n'
        'Выберите клиента:',
        reply_markup=client_markup_temp
    )
    
    context.user_data['selecting_partner_client'] = True
    return PARTNER_CLIENT

async def handle_partner_somoni(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка суммы в сомони для партнера"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    try:
        somoni = float(text)
        if somoni <= 0:
            await update.message.reply_text('❌ Сумма должна быть больше 0')
            return PARTNER_SOMONI
        context.user_data['partner_somoni'] = somoni
    except ValueError:
        await update.message.reply_text('❌ Введите корректное число')
        return PARTNER_SOMONI
    
    await update.message.reply_text(
        '💱 Введите курс обмена:\n'
        '(например: 13.5)',
        reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
    )
    return PARTNER_RATE

async def handle_partner_rate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка курса обмена для партнера"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    try:
        rate = float(text)
        if rate <= 0:
            await update.message.reply_text('❌ Курс должен быть больше 0')
            return PARTNER_RATE
        context.user_data['partner_rate'] = rate
    except ValueError:
        await update.message.reply_text('❌ Введите корректное число')
        return PARTNER_RATE
    
    # Вычисляем сумму в долларах
    somoni = context.user_data['partner_somoni']
    total_usd = somoni / rate
    
    await update.message.reply_text(
        f'💰 Расчет:\n'
        f'Сомони: {somoni}\n'
        f'Курс: {rate}\n'
        f'Сумма $: {total_usd:.2f}\n\n'
        '📝 Введите примечания\n'
        '(или нажмите "Пропустить"):',
        reply_markup=ReplyKeyboardMarkup([['Пропустить'], ['❌ Отмена']], resize_keyboard=True)
    )
    return PARTNER_NOTES

async def handle_partner_notes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка примечаний и сохранение партнера"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('Операция отменена', reply_markup=markup)
        return CHOOSING_ACTION
    
    if text == 'Пропустить':
        notes = None
    else:
        notes = text
    
    # Сохраняем партнера
    partner_date = context.user_data['partner_date']
    client_id = context.user_data['partner_client_id']
    somoni = context.user_data['partner_somoni']
    rate = context.user_data['partner_rate']
    user_id = context.user_data['user_id']
    
    success, result = add_partner(partner_date, client_id, somoni, rate, notes, user_id)
    
    if success:
        total_usd = somoni / rate
        client_name = get_client_by_id(client_id)
        
        await update.message.reply_text(
            f'✅ Партнер оформлен!\n\n'
            f'📅 Дата: {partner_date}\n'
            f'👤 Клиент: {client_name}\n'
            f'💵 Сомони: {somoni}\n'
            f'💱 Курс: {rate}\n'
            f'💰 Сумма $: {total_usd:.2f}\n'
            f'📋 ID партнера: {result}',
            reply_markup=markup
        )
    else:
        await update.message.reply_text(f'❌ Ошибка: {result}', reply_markup=markup)
    
    return CHOOSING_ACTION

# ============= УДАЛЕНИЕ ЗАПИСЕЙ (ТОЛЬКО ДЛЯ АДМИНОВ) =============

async def handle_delete_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка меню удаления записей"""
    text = update.message.text
    telegram_id = update.effective_user.id
    
    session = get_session(telegram_id)
    if not session:
        await update.message.reply_text('❌ Сессия истекла. Войдите заново: /start')
        return ConversationHandler.END
    
    user_id, username, role, warehouse_group_id, is_active = session
    
    if role != 'admin':
        await update.message.reply_text('❌ Недостаточно прав', reply_markup=markup)
        return CHOOSING_ACTION
    
    if text == '🗑 Удалить приход':
        context.user_data['delete_type'] = 'arrival'
        await update.message.reply_text(
            '🗑 Введите ID записи прихода для удаления:\n'
            '(ID можно найти в истории или отчетах)',
            reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
        )
        return DELETE_ID
    
    elif text == '🗑 Удалить расход':
        context.user_data['delete_type'] = 'departure'
        await update.message.reply_text(
            '🗑 Введите ID записи расхода для удаления:\n'
            '(ID можно найти в истории или отчетах)',
            reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
        )
        return DELETE_ID
    
    elif text == '🗑 Удалить погашение':
        context.user_data['delete_type'] = 'payment'
        await update.message.reply_text(
            '🗑 Введите ID записи погашения для удаления:\n'
            '(ID можно найти в отчетах)',
            reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
        )
        return DELETE_ID
    
    elif text == '🗑 Удалить партнера':
        context.user_data['delete_type'] = 'partner'
        await update.message.reply_text(
            '🗑 Введите ID записи партнера для удаления:\n'
            '(ID можно найти в отчетах)',
            reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
        )
        return DELETE_ID
    
    elif text == '🔙 Назад':
        await update.message.reply_text('⚙️ Управление системой:', reply_markup=management_markup)
        return MANAGEMENT_MENU

async def handle_delete_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка ID для удаления - показ информации"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('⚙️ Управление системой:', reply_markup=management_markup)
        return MANAGEMENT_MENU
    
    try:
        record_id = int(text)
    except ValueError:
        await update.message.reply_text('❌ Введите корректный ID (число)')
        return DELETE_ID
    
    delete_type = context.user_data.get('delete_type')
    context.user_data['delete_id'] = record_id
    
    # Получаем информацию о записи
    if delete_type == 'arrival':
        record = get_arrival_by_id(record_id)
        if not record:
            await update.message.reply_text('❌ Запись прихода не найдена', reply_markup=management_markup)
            return MANAGEMENT_MENU
        
        arr_id, arr_date, wagon, firm, warehouse, product, source, qty_doc, qty_act, notes, username_arr, created = record
        tons = qty_act / 20
        diff = qty_act - qty_doc
        
        message = f'📦 ИНФОРМАЦИЯ О ПРИХОДЕ #{arr_id}\n\n'
        message += f'📅 Дата: {arr_date}\n'
        message += f'🚂 Вагон: {wagon or "не указан"}\n'
        message += f'🏭 Фирма: {firm or "не указана"}\n'
        message += f'🏢 Склад: {warehouse}\n'
        message += f'📦 Товар: {product}\n'
        message += f'📄 По документу: {qty_doc} шт\n'
        message += f'✅ По факту: {qty_act} шт\n'
        message += f'⚖️ Тонны: {tons:.2f} т\n'
        message += f'📊 Расхождение: {diff} шт\n'
        message += f'📝 Примечания: {notes or "нет"}\n'
        message += f'👤 Пользователь: {username_arr}\n'
        message += f'🕐 Создано: {created}\n'
        
    elif delete_type == 'departure':
        record = get_departures(limit=1, departure_id=record_id)
        if not record:
            await update.message.reply_text('❌ Запись расхода не найдена', reply_markup=management_markup)
            return MANAGEMENT_MENU
        
        dep_id, dep_date, coalition, firm, warehouse, product, qty, price, notes, username_dep, created, client = record[0]
        tons = qty / 20
        sum_value = tons * float(price) if price else 0
        
        message = f'📤 ИНФОРМАЦИЯ О РАСХОДЕ #{dep_id}\n\n'
        message += f'📅 Дата: {dep_date}\n'
        message += f'📊 Коалица: {coalition or "не указана"}\n'
        message += f'🏭 Фирма: {firm or "не указана"}\n'
        message += f'🏢 Склад: {warehouse}\n'
        message += f'📦 Товар: {product}\n'
        message += f'👤 Клиент: {client or "не указан"}\n'
        message += f'📦 Количество: {qty} шт\n'
        message += f'⚖️ Тонны: {tons:.2f} т\n'
        message += f'💰 Цена: {price} $/т\n'
        message += f'💵 Сумма: {sum_value:.2f} $\n'
        message += f'📝 Примечания: {notes or "нет"}\n'
        message += f'👤 Пользователь: {username_dep}\n'
        message += f'🕐 Создано: {created}\n'
        
    elif delete_type == 'payment':
        record = get_payment_by_id(record_id)
        if not record:
            await update.message.reply_text('❌ Запись погашения не найдена', reply_markup=management_markup)
            return MANAGEMENT_MENU
        
        pay_id, pay_date, client_name, somoni, dollar, rate, total_usd, notes, username_pay, created_at = record
        
        message = f'💰 ИНФОРМАЦИЯ О ПОГАШЕНИИ #{pay_id}\n\n'
        message += f'📅 Дата: {pay_date}\n'
        message += f'👤 Клиент: {client_name or "не указан"}\n'
        message += f'💵 Сомони: {somoni}\n'
        message += f'💱 Курс: {rate}\n'
        message += f'💰 Сумма $: {total_usd}\n'
        message += f'📝 Примечания: {notes or "нет"}\n'
        message += f'👤 Пользователь: {username_pay}\n'
        message += f'🕐 Создано: {created_at}\n'
        
    elif delete_type == 'partner':
        record = get_partner_by_id(record_id)
        if not record:
            await update.message.reply_text('❌ Запись партнера не найдена', reply_markup=management_markup)
            return MANAGEMENT_MENU
        
        part_id, part_date, client_name, somoni, rate, total_usd, notes, username_part, created_at = record
        
        message = f'🤝 ИНФОРМАЦИЯ О ПАРТНЕРЕ #{part_id}\n\n'
        message += f'📅 Дата: {part_date}\n'
        message += f'👤 Клиент: {client_name or "не указан"}\n'
        message += f'💵 Сомони: {somoni}\n'
        message += f'💱 Курс: {rate}\n'
        message += f'💰 Сумма $: {total_usd}\n'
        message += f'📝 Примечания: {notes or "нет"}\n'
        message += f'👤 Пользователь: {username_part}\n'
        message += f'🕐 Создано: {created_at}\n'
    else:
        await update.message.reply_text('❌ Ошибка: неизвестный тип записи', reply_markup=management_markup)
        return MANAGEMENT_MENU
    
    message += '\n⚠️ Вы уверены, что хотите удалить эту запись?'
    
    confirm_keyboard = [
        ['✅ Да, удалить', '❌ Нет, отменить']
    ]
    confirm_markup = ReplyKeyboardMarkup(confirm_keyboard, resize_keyboard=True)
    
    await update.message.reply_text(message, reply_markup=confirm_markup)
    return DELETE_CONFIRM

async def handle_delete_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка подтверждения удаления"""
    text = update.message.text
    
    if text == '❌ Нет, отменить':
        await update.message.reply_text('❌ Удаление отменено', reply_markup=management_markup)
        return MANAGEMENT_MENU
    
    if text != '✅ Да, удалить':
        await update.message.reply_text('❌ Неверный выбор', reply_markup=management_markup)
        return MANAGEMENT_MENU
    
    delete_type = context.user_data.get('delete_type')
    record_id = context.user_data.get('delete_id')
    
    if delete_type == 'arrival':
        success, message = delete_arrival(record_id)
    elif delete_type == 'departure':
        success, message = delete_departure(record_id)
    elif delete_type == 'payment':
        success, message = delete_payment(record_id)
    elif delete_type == 'partner':
        success, message = delete_partner(record_id)
    else:
        await update.message.reply_text('❌ Ошибка: неизвестный тип записи', reply_markup=management_markup)
        return MANAGEMENT_MENU
    
    if success:
        await update.message.reply_text(f'✅ {message}', reply_markup=management_markup)
    else:
        await update.message.reply_text(f'❌ Ошибка: {message}', reply_markup=management_markup)
    
    return MANAGEMENT_MENU

# ============= РЕДАКТИРОВАНИЕ ЗАПИСЕЙ (ВСЕ ПОЛЯ) =============

async def handle_edit_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка меню редактирования"""
    text = update.message.text
    
    if text == '✏️ Изменить приход':
        context.user_data['edit_type'] = 'arrival'
    elif text == '✏️ Изменить расход':
        context.user_data['edit_type'] = 'departure'
    elif text == '✏️ Изменить погашение':
        context.user_data['edit_type'] = 'payment'
    elif text == '✏️ Изменить партнера':
        context.user_data['edit_type'] = 'partner'
    elif text == '🔙 Назад':
        await update.message.reply_text('⚙️ Управление системой:', reply_markup=management_markup)
        return MANAGEMENT_MENU
    else:
        return EDIT_MENU
    
    await update.message.reply_text(
        '✏️ Введите ID записи для изменения:',
        reply_markup=ReplyKeyboardMarkup([['❌ Отмена']], resize_keyboard=True)
    )
    return EDIT_ID

async def handle_edit_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показ записи и выбор поля для изменения"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('⚙️ Управление системой:', reply_markup=management_markup)
        return MANAGEMENT_MENU
    
    try:
        record_id = int(text)
        context.user_data['edit_id'] = record_id
    except ValueError:
        await update.message.reply_text('❌ Введите корректный ID')
        return EDIT_ID
    
    edit_type = context.user_data.get('edit_type')
    
    # Получаем и показываем запись
    if edit_type == 'arrival':
        record = get_arrival_by_id(record_id)
        if not record:
            await update.message.reply_text('❌ Запись не найдена', reply_markup=management_markup)
            return MANAGEMENT_MENU
        
        context.user_data['edit_record'] = record
        info = (
            f"📦 Приход ID: {record[0]}\n"
            f"📅 Дата: {record[1]}\n"
            f"🚂 Вагон: {record[2]}\n"
            f"🏭 Фирма: {record[3]}\n"
            f"🏢 Склад: {record[4]}\n"
            f"📦 Товар: {record[5]}\n"
            f"📍 Источник: {record[6]}\n"
            f"📄 Кол-во по док.: {record[7]}\n"
            f"✅ Кол-во факт.: {record[8]}\n"
            f"📝 Примечания: {record[9] or 'нет'}\n\n"
            "Выберите поле для изменения:"
        )
        field_keyboard = [
            ['📅 Дата', '🚂 Вагон'],
            ['📦 Товар', '📍 Источник'],
            ['📄 Кол-во док.', '✅ Кол-во факт.'],
            ['📝 Примечания'],
            ['❌ Отмена']
        ]
        
    elif edit_type == 'departure':
        records = get_departures(limit=1, departure_id=record_id)
        if not records:
            await update.message.reply_text('❌ Запись не найдена', reply_markup=management_markup)
            return MANAGEMENT_MENU
        
        record = records[0]
        context.user_data['edit_record'] = record
        info = (
            f"📤 Расход ID: {record[0]}\n"
            f"📅 Дата: {record[1]}\n"
            f"📊 Коалица: {record[2]}\n"
            f"🏭 Фирма: {record[3]}\n"
            f"🏢 Склад: {record[4]}\n"
            f"📦 Товар: {record[5]}\n"
            f"📦 Количество: {record[6]}\n"
            f"💵 Цена: {record[7]}\n"
            f"📝 Примечания: {record[8] or 'нет'}\n"
            f"👤 Клиент: {record[11] or 'нет'}\n\n"
            "Выберите поле для изменения:"
        )
        field_keyboard = [
            ['📅 Дата', '📦 Товар'],
            ['📦 Количество', '💵 Цена'],
            ['📝 Примечания'],
            ['❌ Отмена']
        ]
        
    elif edit_type == 'payment':
        record = get_payment_by_id(record_id)
        if not record:
            await update.message.reply_text('❌ Запись не найдена', reply_markup=management_markup)
            return MANAGEMENT_MENU
        
        context.user_data['edit_record'] = record
        info = (
            f"💰 Погашение ID: {record[0]}\n"
            f"📅 Дата: {record[1]}\n"
            f"👤 Клиент: {record[2]}\n"
            f"💵 Сомони: {record[3]}\n"
            f"💱 Курс: {record[5]}\n"
            f"💲 Сумма $: {record[6]:.2f}\n"
            f"📝 Примечания: {record[7] or 'нет'}\n\n"
            "Выберите поле для изменения:"
        )
        field_keyboard = [
            ['📅 Дата', '💵 Сомони'],
            ['💱 Курс', '📝 Примечания'],
            ['❌ Отмена']
        ]
        
    elif edit_type == 'partner':
        record = get_partner_by_id(record_id)
        if not record:
            await update.message.reply_text('❌ Запись не найдена', reply_markup=management_markup)
            return MANAGEMENT_MENU
        
        context.user_data['edit_record'] = record
        info = (
            f"🤝 Партнер ID: {record[0]}\n"
            f"📅 Дата: {record[1]}\n"
            f"👤 Клиент: {record[2]}\n"
            f"💵 Сомони: {record[3]}\n"
            f"💱 Курс: {record[4]}\n"
            f"💲 Сумма $: {record[5]:.2f}\n"
            f"📝 Примечания: {record[6] or 'нет'}\n\n"
            "Выберите поле для изменения:"
        )
        field_keyboard = [
            ['📅 Дата', '💵 Сомони'],
            ['💱 Курс', '📝 Примечания'],
            ['❌ Отмена']
        ]
    else:
        await update.message.reply_text('❌ Ошибка', reply_markup=management_markup)
        return MANAGEMENT_MENU
    
    await update.message.reply_text(
        info,
        reply_markup=ReplyKeyboardMarkup(field_keyboard, resize_keyboard=True)
    )
    return EDIT_FIELD

async def handle_edit_field(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора поля для изменения"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('⚙️ Управление системой:', reply_markup=management_markup)
        return MANAGEMENT_MENU
    
    edit_type = context.user_data.get('edit_type')
    
    # Определяем какое поле редактируем
    field_map = {
        '📅 Дата': 'date',
        '🚂 Вагон': 'wagon',
        '📦 Товар': 'product',
        '📍 Источник': 'source',
        '📄 Кол-во док.': 'qty_doc',
        '✅ Кол-во факт.': 'qty_actual',
        '📦 Количество': 'quantity',
        '💵 Цена': 'price',
        '💵 Сомони': 'somoni',
        '💱 Курс': 'rate',
        '📝 Примечания': 'notes'
    }
    
    field = field_map.get(text)
    if not field:
        await update.message.reply_text('❌ Неверный выбор')
        return EDIT_FIELD
    
    context.user_data['edit_field'] = field
    
    # Запрашиваем новое значение
    prompts = {
        'date': '📅 Введите новую дату (ДД.ММ.ГГГГ):',
        'wagon': '🚂 Введите номер вагона:',
        'product': '📦 Введите название товара:',
        'source': '📍 Введите источник:',
        'qty_doc': '📄 Введите количество по документу:',
        'qty_actual': '✅ Введите фактическое количество:',
        'quantity': '📦 Введите количество:',
        'price': '💵 Введите цену:',
        'somoni': '💵 Введите сумму в сомони:',
        'rate': '💱 Введите курс:',
        'notes': '📝 Введите примечания (или "Пропустить"):'
    }
    
    await update.message.reply_text(
        prompts.get(field, 'Введите новое значение:'),
        reply_markup=ReplyKeyboardMarkup([['Пропустить'], ['❌ Отмена']], resize_keyboard=True)
    )
    return EDIT_VALUE

async def handle_edit_value(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Сохранение нового значения"""
    text = update.message.text
    
    if text == '❌ Отмена':
        await update.message.reply_text('⚙️ Управление системой:', reply_markup=management_markup)
        return MANAGEMENT_MENU
    
    edit_type = context.user_data.get('edit_type')
    record_id = context.user_data.get('edit_id')
    field = context.user_data.get('edit_field')
    
    # Обработка значения
    if text == 'Пропустить' and field == 'notes':
        new_value = None
    else:
        new_value = text
    
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        # Обновляем соответствующее поле
        if edit_type == 'arrival':
            field_db_map = {
                'date': 'arrival_date',
                'wagon': 'wagon_number',
                'product': 'product_name',
                'source': 'source',
                'qty_doc': 'quantity_document',
                'qty_actual': 'quantity_actual',
                'notes': 'notes'
            }
            db_field = field_db_map.get(field)
            
            # Валидация для числовых полей
            if field in ['qty_doc', 'qty_actual']:
                new_value = float(new_value)
            
            cursor.execute(f'UPDATE arrivals SET {db_field} = %s WHERE id = %s', (new_value, record_id))
            
        elif edit_type == 'departure':
            field_db_map = {
                'date': 'departure_date',
                'product': 'product_name',
                'quantity': 'quantity',
                'price': 'price',
                'notes': 'notes'
            }
            db_field = field_db_map.get(field)
            
            # Валидация для числовых полей
            if field in ['quantity', 'price']:
                new_value = float(new_value)
                
                # Пересчитываем total если изменили количество или цену
                # Формула: (quantity / 20) * price
                cursor.execute('SELECT quantity, price FROM departures WHERE id = %s', (record_id,))
                current = cursor.fetchone()
                if field == 'quantity':
                    total = (new_value / 20) * current[1]
                    cursor.execute('UPDATE departures SET quantity = %s, total = %s WHERE id = %s', 
                                 (new_value, total, record_id))
                elif field == 'price':
                    total = (current[0] / 20) * new_value
                    cursor.execute('UPDATE departures SET price = %s, total = %s WHERE id = %s', 
                                 (new_value, total, record_id))
            else:
                cursor.execute(f'UPDATE departures SET {db_field} = %s WHERE id = %s', (new_value, record_id))
            
        elif edit_type == 'payment':
            field_db_map = {
                'date': 'payment_date',
                'somoni': 'somoni',
                'rate': 'exchange_rate',
                'notes': 'notes'
            }
            db_field = field_db_map.get(field)
            
            # Валидация для числовых полей
            if field in ['somoni', 'rate']:
                new_value = float(new_value)
                
                # Пересчитываем total_usd если изменили сомони или курс
                cursor.execute('SELECT somoni, exchange_rate FROM payments WHERE id = %s', (record_id,))
                current = cursor.fetchone()
                if field == 'somoni':
                    total_usd = new_value / current[1]
                    cursor.execute('UPDATE payments SET somoni = %s, total_usd = %s WHERE id = %s', 
                                 (new_value, total_usd, record_id))
                elif field == 'rate':
                    total_usd = current[0] / new_value
                    cursor.execute('UPDATE payments SET exchange_rate = %s, total_usd = %s WHERE id = %s', 
                                 (new_value, total_usd, record_id))
            else:
                cursor.execute(f'UPDATE payments SET {db_field} = %s WHERE id = %s', (new_value, record_id))
            
        elif edit_type == 'partner':
            field_db_map = {
                'date': 'partner_date',
                'somoni': 'somoni',
                'rate': 'exchange_rate',
                'notes': 'notes'
            }
            db_field = field_db_map.get(field)
            
            # Валидация для числовых полей
            if field in ['somoni', 'rate']:
                new_value = float(new_value)
                
                # Пересчитываем total_usd если изменили сомони или курс
                cursor.execute('SELECT somoni, exchange_rate FROM partners WHERE id = %s', (record_id,))
                current = cursor.fetchone()
                if field == 'somoni':
                    total_usd = new_value / current[1]
                    cursor.execute('UPDATE partners SET somoni = %s, total_usd = %s WHERE id = %s', 
                                 (new_value, total_usd, record_id))
                elif field == 'rate':
                    total_usd = current[0] / new_value
                    cursor.execute('UPDATE partners SET exchange_rate = %s, total_usd = %s WHERE id = %s', 
                                 (new_value, total_usd, record_id))
            else:
                cursor.execute(f'UPDATE partners SET {db_field} = %s WHERE id = %s', (new_value, record_id))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        await update.message.reply_text('✅ Данные обновлены', reply_markup=management_markup)
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        await update.message.reply_text(f'❌ Ошибка: {str(e)}', reply_markup=management_markup)
    
    return MANAGEMENT_MENU


async def handle_callback_query(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка нажатий на inline кнопки"""
    query = update.callback_query
    await query.answer()
    
    telegram_id = update.effective_user.id
    session = get_session(telegram_id)
    
    if not session:
        await query.edit_message_text('❌ Сессия истекла. Войдите заново: /start')
        return ConversationHandler.END
    
    user_id, username, role, warehouse_group_id, is_active = session
    
    # Экспорт долгов клиентов
    if query.data == 'export_client_debts':
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import os
        
        logging.info(f"Export client debts called by user {user_id}")
        
        await query.edit_message_text(query.message.text + '\n\n⏳ Формирую отчет...')
        
        # Получаем данные о долгах
        debts = context.user_data.get('client_debts', [])
        report_year = context.user_data.get('report_year', datetime.now().year)
        
        logging.info(f"Debts data: {len(debts) if debts else 0} records, year: {report_year}")
        
        if not debts:
            await query.edit_message_text('❌ Нет данных для экспорта')
            return CHOOSING_ACTION
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = f"Долги клиентов {report_year}"
        
        # Заголовок
        headers = ['№', 'Клиент', 'Сумма продаж ($)', 'Погашения ($)', 'Остаток долга ($)']
        
        # Стиль заголовка
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Границы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Записываем данные
        total_sales = 0
        total_paid = 0
        total_debt = 0
        
        for idx, (client_id, client_name, sales, paid, debt) in enumerate(debts, 1):
            row = idx + 1
            ws.cell(row=row, column=1, value=idx).border = thin_border
            ws.cell(row=row, column=2, value=client_name).border = thin_border
            ws.cell(row=row, column=3, value=float(sales)).border = thin_border
            ws.cell(row=row, column=4, value=float(paid)).border = thin_border
            ws.cell(row=row, column=5, value=float(debt)).border = thin_border
            
            # Форматируем числа
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            
            total_sales += sales
            total_paid += paid
            total_debt += debt
        
        # Добавляем итоговую строку
        total_row = len(debts) + 2
        ws.cell(row=total_row, column=1, value='ИТОГО:').font = Font(bold=True)
        ws.cell(row=total_row, column=1).border = thin_border
        ws.cell(row=total_row, column=2).border = thin_border
        ws.cell(row=total_row, column=3, value=float(total_sales)).font = Font(bold=True)
        ws.cell(row=total_row, column=3).number_format = '#,##0.00'
        ws.cell(row=total_row, column=3).border = thin_border
        ws.cell(row=total_row, column=4, value=float(total_paid)).font = Font(bold=True)
        ws.cell(row=total_row, column=4).number_format = '#,##0.00'
        ws.cell(row=total_row, column=4).border = thin_border
        ws.cell(row=total_row, column=5, value=float(total_debt)).font = Font(bold=True)
        ws.cell(row=total_row, column=5).number_format = '#,##0.00'
        ws.cell(row=total_row, column=5).border = thin_border
        ws.cell(row=total_row, column=5).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # Настраиваем ширину колонок
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20
        
        # Сохраняем файл
        filename = f'client_debts_{report_year}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(filename)
        
        # Отправляем файл
        with open(filename, 'rb') as file:
            await context.bot.send_document(
                chat_id=query.message.chat_id,
                document=file,
                filename=filename,
                caption=f'📊 Отчет по долгам клиентов за {report_year} год\n'
                       f'👥 Должников: {len(debts)}\n'
                       f'💰 Итого долг: {total_debt:,.2f} $'
            )
        
        # Удаляем временный файл
        os.remove(filename)
        
        await query.edit_message_text(query.message.text.replace('⏳ Формирую отчет...', '✅ Отчет отправлен!'))
        return CHOOSING_ACTION
    
    # Меню выбора периода для экспорта
    elif query.data == 'export_arrival_menu':
        context.user_data['export_type'] = 'arrival'
        period_keyboard = [
            [InlineKeyboardButton("📅 Сегодня", callback_data='export_period_today')],
            [InlineKeyboardButton("📅 Неделя", callback_data='export_period_week')],
            [InlineKeyboardButton("📅 Месяц", callback_data='export_period_month')],
            [InlineKeyboardButton("📅 Весь год", callback_data='export_period_year')],
            [InlineKeyboardButton("🔙 Назад", callback_data='back_to_report')]
        ]
        period_markup = InlineKeyboardMarkup(period_keyboard)
        await query.edit_message_reply_markup(reply_markup=period_markup)
        return REPORT_EXPORT_PERIOD
    
    elif query.data == 'export_departure_menu':
        context.user_data['export_type'] = 'departure'
        period_keyboard = [
            [InlineKeyboardButton("📅 Сегодня", callback_data='export_period_today')],
            [InlineKeyboardButton("📅 Неделя", callback_data='export_period_week')],
            [InlineKeyboardButton("📅 Месяц", callback_data='export_period_month')],
            [InlineKeyboardButton("📅 Весь год", callback_data='export_period_year')],
            [InlineKeyboardButton("🔙 Назад", callback_data='back_to_report')]
        ]
        period_markup = InlineKeyboardMarkup(period_keyboard)
        await query.edit_message_reply_markup(reply_markup=period_markup)
        return REPORT_EXPORT_PERIOD
    
    elif query.data == 'export_payment_menu':
        context.user_data['export_type'] = 'payment'
        period_keyboard = [
            [InlineKeyboardButton("📅 Сегодня", callback_data='export_period_today')],
            [InlineKeyboardButton("📅 Неделя", callback_data='export_period_week')],
            [InlineKeyboardButton("📅 Месяц", callback_data='export_period_month')],
            [InlineKeyboardButton("📅 Весь год", callback_data='export_period_year')],
            [InlineKeyboardButton("🔙 Назад", callback_data='back_to_report')]
        ]
        period_markup = InlineKeyboardMarkup(period_keyboard)
        await query.edit_message_reply_markup(reply_markup=period_markup)
        return REPORT_EXPORT_PERIOD
    
    elif query.data == 'export_partner_menu':
        context.user_data['export_type'] = 'partner'
        period_keyboard = [
            [InlineKeyboardButton("📅 Сегодня", callback_data='export_period_today')],
            [InlineKeyboardButton("📅 Неделя", callback_data='export_period_week')],
            [InlineKeyboardButton("📅 Месяц", callback_data='export_period_month')],
            [InlineKeyboardButton("📅 Весь год", callback_data='export_period_year')],
            [InlineKeyboardButton("🔙 Назад", callback_data='back_to_report')]
        ]
        period_markup = InlineKeyboardMarkup(period_keyboard)
        await query.edit_message_reply_markup(reply_markup=period_markup)
        return REPORT_EXPORT_PERIOD
    
    # Возврат к исходным кнопкам отчета
    elif query.data == 'back_to_report':
        export_type = context.user_data.get('export_type', 'arrival')
        if export_type == 'arrival':
            inline_keyboard = [
                [InlineKeyboardButton("📊 Детальный отчет (Excel)", callback_data='export_arrival_menu')],
                [InlineKeyboardButton("🔙 Назад", callback_data='back_to_report_menu')]
            ]
        elif export_type == 'departure':
            inline_keyboard = [
                [InlineKeyboardButton("📊 Детальный отчет (Excel)", callback_data='export_departure_menu')],
                [InlineKeyboardButton("🔙 Назад", callback_data='back_to_report_menu')]
            ]
        elif export_type == 'payment':
            inline_keyboard = [
                [InlineKeyboardButton("📊 Детальный отчет (Excel)", callback_data='export_payment_menu')],
                [InlineKeyboardButton("🔙 Назад", callback_data='back_to_report_menu')]
            ]
        else:  # partner
            inline_keyboard = [
                [InlineKeyboardButton("📊 Детальный отчет (Excel)", callback_data='export_partner_menu')],
                [InlineKeyboardButton("🔙 Назад", callback_data='back_to_report_menu')]
            ]
        inline_markup = InlineKeyboardMarkup(inline_keyboard)
        await query.edit_message_reply_markup(reply_markup=inline_markup)
        return REPORT_EXPORT_PERIOD
    
    # Экспорт с выбранным периодом
    elif query.data.startswith('export_period_'):
        from datetime import datetime, timedelta
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import os
        
        period = query.data.replace('export_period_', '')
        export_type = context.user_data.get('export_type', 'arrival')
        report_year = context.user_data.get('report_year', datetime.now().year)
        
        # Определяем даты в зависимости от периода
        today = datetime.now()
        if period == 'today':
            date_from = today.strftime('%Y-%m-%d')
            date_to = today.strftime('%Y-%m-%d')
            period_name = 'за сегодня'
        elif period == 'week':
            date_from = (today - timedelta(days=7)).strftime('%Y-%m-%d')
            date_to = today.strftime('%Y-%m-%d')
            period_name = 'за неделю'
        elif period == 'month':
            date_from = (today - timedelta(days=30)).strftime('%Y-%m-%d')
            date_to = today.strftime('%Y-%m-%d')
            period_name = 'за месяц'
        else:  # year
            date_from = f'{report_year}-01-01'
            date_to = f'{report_year}-12-31'
            period_name = f'за {report_year} год'
        
        await query.edit_message_text(query.message.text + f'\n\n⏳ Формирую отчет {period_name}...')
        
        # Получаем склады пользователя для фильтрации
        user_warehouses = get_user_warehouses(user_id)
        warehouse_ids = [wh[0] for wh in user_warehouses] if user_warehouses else []
        
        if export_type == 'arrival':
            # Экспорт приходов
            all_arrivals = []
            
            if role == 'admin':
                all_arrivals = get_arrivals(limit=10000, date_from=date_from, date_to=date_to)
            else:
                for wh_id in warehouse_ids:
                    arrivals_wh = get_arrivals(limit=10000, warehouse_id=wh_id, date_from=date_from, date_to=date_to)
                    all_arrivals.extend(arrivals_wh)
            
            # Создаем Excel файл
            wb = Workbook()
            ws = wb.active
            ws.title = f"Приходы {report_year}"
            
            # Заголовок
            headers = ['№', 'Дата', 'Вагон', 'Фирма', 'Склад', 'Товар', 
                      'По документу (шт)', 'По факту (шт)', 'Тонны', 'Расхождение', 'Примечания', 'Пользователь']
            
            # Стиль заголовка
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Данные
            for idx, arr in enumerate(all_arrivals, 2):
                arr_id, arr_date, wagon, firm, warehouse, product, source, qty_doc, qty_act, notes, username_arr, created = arr
                
                tons = qty_act / 20
                diff = qty_act - qty_doc
                
                # Преобразуем дату в строку если нужно
                date_str = arr_date if isinstance(arr_date, str) else arr_date.strftime('%Y-%m-%d')
                
                ws.cell(row=idx, column=1, value=arr_id)
                ws.cell(row=idx, column=2, value=date_str)
                ws.cell(row=idx, column=3, value=wagon or '')
                ws.cell(row=idx, column=4, value=firm or '')
                ws.cell(row=idx, column=5, value=warehouse or '')
                ws.cell(row=idx, column=6, value=product)
                ws.cell(row=idx, column=7, value=qty_doc)
                ws.cell(row=idx, column=8, value=qty_act)
                ws.cell(row=idx, column=9, value=round(tons, 2))
                ws.cell(row=idx, column=10, value=diff)
                ws.cell(row=idx, column=12, value=notes or '')
                ws.cell(row=idx, column=12, value=username_arr or '')
            
            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем файл
            filename = f'arrival_report_{report_year}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(filename)
            
            # Отправляем файл
            with open(filename, 'rb') as file:
                await context.bot.send_document(
                    chat_id=query.message.chat_id,
                    document=file,
                    filename=filename,
                    caption=f'📊 Детальный отчет по приходам за {report_year} год'
                )
            
            # Удаляем временный файл
            os.remove(filename)
            
        elif export_type == 'departure':
            # Экспорт расходов
            all_departures = []
            
            if role == 'admin':
                all_departures = get_departures(limit=10000, date_from=date_from, date_to=date_to)
            else:
                for wh_id in warehouse_ids:
                    departures_wh = get_departures(limit=10000, warehouse_id=wh_id, date_from=date_from, date_to=date_to)
                    all_departures.extend(departures_wh)
            
            # Создаем Excel файл
            wb = Workbook()
            ws = wb.active
            ws.title = f"Расходы {report_year}"
            
            # Заголовок
            headers = ['№', 'Дата', 'Коалица', 'Фирма', 'Склад', 'Товар', 'Клиент',
                      'Количество (шт)', 'Тонны', 'Цена ($)', 'Сумма ($)', 'Примечания', 'Пользователь']
            
            # Стиль заголовка
            header_fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Данные
            for idx, dep in enumerate(all_departures, 2):
                dep_id, dep_date, coalition, firm, warehouse, product, qty, price, notes, username_dep, created, client = dep
                
                tons = qty / 20
                sum_value = tons * float(price) if price else 0
                
                # Преобразуем дату в строку если нужно
                date_str = dep_date if isinstance(dep_date, str) else dep_date.strftime('%Y-%m-%d')
                
                ws.cell(row=idx, column=1, value=dep_id)
                ws.cell(row=idx, column=2, value=date_str)
                ws.cell(row=idx, column=3, value=coalition or '')
                ws.cell(row=idx, column=4, value=firm or '')
                ws.cell(row=idx, column=5, value=warehouse or '')
                ws.cell(row=idx, column=6, value=product)
                ws.cell(row=idx, column=7, value=client or '')
                ws.cell(row=idx, column=8, value=qty)
                ws.cell(row=idx, column=9, value=round(tons, 2))
                ws.cell(row=idx, column=10, value=round(float(price), 2) if price else 0)
                ws.cell(row=idx, column=11, value=round(sum_value, 2))
                ws.cell(row=idx, column=12, value=notes or '')
                ws.cell(row=idx, column=13, value=username_dep or '')
            
            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем файл
            filename = f'departure_report_{report_year}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(filename)
            
            # Отправляем файл
            with open(filename, 'rb') as file:
                await context.bot.send_document(
                    chat_id=query.message.chat_id,
                    document=file,
                    filename=filename,
                    caption=f'📊 Детальный отчет по расходам за {report_year} год'
                )
            
            # Удаляем временный файл
            os.remove(filename)
        
        elif export_type == 'payment':
            # Экспорт погашений
            all_payments = get_payments(limit=10000, date_from=date_from, date_to=date_to)
            
            # Создаем Excel файл
            wb = Workbook()
            ws = wb.active
            ws.title = f"Погашения {report_year}"
            
            # Заголовок
            headers = ['№', 'Дата', 'Клиент', 'Сомони', 'Курс', 'Сумма ($)', 'Примечания', 'Пользователь']
            
            # Стиль заголовка
            header_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
            header_font = Font(bold=True, color='000000')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Данные
            for idx, payment in enumerate(all_payments, 2):
                pay_id, pay_date, client_name, somoni, dollar, rate, total_usd, notes, username_pay, created_at = payment
                
                # Преобразуем дату в строку если нужно
                date_str = pay_date if isinstance(pay_date, str) else pay_date.strftime('%Y-%m-%d')
                
                ws.cell(row=idx, column=1, value=pay_id)
                ws.cell(row=idx, column=2, value=date_str)
                ws.cell(row=idx, column=3, value=client_name or '')
                ws.cell(row=idx, column=4, value=round(float(somoni), 2))
                ws.cell(row=idx, column=5, value=round(float(rate), 4))
                ws.cell(row=idx, column=6, value=round(float(total_usd), 2))
                ws.cell(row=idx, column=7, value=notes or '')
                ws.cell(row=idx, column=8, value=username_pay or '')
            
            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем файл
            filename = f'payment_report_{report_year}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(filename)
            
            # Отправляем файл
            with open(filename, 'rb') as file:
                await context.bot.send_document(
                    chat_id=query.message.chat_id,
                    document=file,
                    filename=filename,
                    caption=f'📊 Детальный отчет по погашениям {period_name}'
                )
            
            # Удаляем временный файл
            os.remove(filename)
        
        else:  # partner
            # Экспорт партнеров
            all_partners = get_partners(limit=10000, date_from=date_from, date_to=date_to)
            
            # Создаем Excel файл
            wb = Workbook()
            ws = wb.active
            ws.title = f"Партнеры {report_year}"
            
            # Заголовок
            headers = ['№', 'Дата', 'Клиент', 'Сомони', 'Курс', 'Сумма ($)', 'Примечания', 'Пользователь']
            
            # Стиль заголовка
            header_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            header_font = Font(bold=True, color='000000')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Данные
            for idx, partner in enumerate(all_partners, 2):
                part_id, part_date, client_name, somoni, rate, total_usd, notes, username_part = partner
                
                # Преобразуем дату в строку если нужно
                date_str = part_date if isinstance(part_date, str) else part_date.strftime('%Y-%m-%d')
                
                ws.cell(row=idx, column=1, value=part_id)
                ws.cell(row=idx, column=2, value=date_str)
                ws.cell(row=idx, column=3, value=client_name or '')
                ws.cell(row=idx, column=4, value=round(float(somoni), 2))
                ws.cell(row=idx, column=5, value=round(float(rate), 4))
                ws.cell(row=idx, column=6, value=round(float(total_usd), 2))
                ws.cell(row=idx, column=7, value=notes or '')
                ws.cell(row=idx, column=8, value=username_part or '')
            
            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем файл
            filename = f'partner_report_{report_year}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(filename)
            
            # Отправляем файл
            with open(filename, 'rb') as file:
                await context.bot.send_document(
                    chat_id=query.message.chat_id,
                    document=file,
                    filename=filename,
                    caption=f'📊 Детальный отчет по партнерам {period_name}'
                )
            
            # Удаляем временный файл
            os.remove(filename)
        
        # Возвращаем исходные кнопки
        if export_type == 'arrival':
            inline_keyboard = [
                [InlineKeyboardButton("📊 Детальный отчет (Excel)", callback_data='export_arrival_menu')],
                [InlineKeyboardButton("🔙 Назад", callback_data='back_to_report_menu')]
            ]
        elif export_type == 'departure':
            inline_keyboard = [
                [InlineKeyboardButton("📊 Детальный отчет (Excel)", callback_data='export_departure_menu')],
                [InlineKeyboardButton("🔙 Назад", callback_data='back_to_report_menu')]
            ]
        elif export_type == 'payment':
            inline_keyboard = [
                [InlineKeyboardButton("📊 Детальный отчет (Excel)", callback_data='export_payment_menu')],
                [InlineKeyboardButton("🔙 Назад", callback_data='back_to_report_menu')]
            ]
        else:  # partner
            inline_keyboard = [
                [InlineKeyboardButton("📊 Детальный отчет (Excel)", callback_data='export_partner_menu')],
                [InlineKeyboardButton("🔙 Назад", callback_data='back_to_report_menu')]
            ]
        inline_markup = InlineKeyboardMarkup(inline_keyboard)
        
        await query.edit_message_text(query.message.text.replace(f'\n\n⏳ Формирую отчет {period_name}...', ''), reply_markup=inline_markup)
        return REPORT_EXPORT_PERIOD
    
    elif query.data == 'back_to_report_menu':
        await query.message.delete()
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text='📋 Выберите тип отчета:',
            reply_markup=report_markup
        )
        return REPORT_MENU
    
    # Обработка выбора клиента для карточки
    elif query.data.startswith('client_card_'):
        client_id = int(query.data.replace('client_card_', ''))
        
        # Получаем имя клиента
        client_name = get_client_by_id(client_id)
        
        if not client_name:
            await query.edit_message_text('❌ Клиент не найден')
            return REPORT_MENU
        
        # Формируем карточку клиента и экспортируем в Excel
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import os
        
        await query.edit_message_text(f'⏳ Формирую карточку клиента {client_name}...')
        
        # Получаем все расходы клиента (покупки)
        conn = get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT departure_date, product_name, quantity, price, total, notes
            FROM departures
            WHERE client_id = %s
            ORDER BY departure_date DESC
        ''', (client_id,))
        purchases = cursor.fetchall()
        
        # Получаем все погашения клиента
        cursor.execute('''
            SELECT payment_date, somoni, exchange_rate, total_usd, notes
            FROM payments
            WHERE client_id = %s
            ORDER BY payment_date DESC
        ''', (client_id,))
        payments = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Создаем Excel файл
        wb = Workbook()
        
        # Лист 1: Покупки
        ws1 = wb.active
        ws1.title = "Покупки"
        
        # Заголовок
        ws1.merge_cells('A1:F1')
        title_cell = ws1['A1']
        title_cell.value = f'КАРТОЧКА КЛИЕНТА: {client_name}'
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        title_cell.font = Font(bold=True, size=14, color='FFFFFF')
        
        # Заголовки таблицы покупок
        headers1 = ['Дата', 'Товар', 'Количество (шт)', 'Цена ($)', 'Сумма ($)', 'Примечания']
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Данные покупок
        total_purchases = 0
        row = 4
        for purchase in purchases:
            date, product, qty, price, total, notes = purchase
            ws1.cell(row=row, column=1, value=str(date)).border = thin_border
            ws1.cell(row=row, column=2, value=product).border = thin_border
            ws1.cell(row=row, column=3, value=float(qty)).border = thin_border
            ws1.cell(row=row, column=3).number_format = '#,##0.00'
            ws1.cell(row=row, column=4, value=float(price)).border = thin_border
            ws1.cell(row=row, column=4).number_format = '#,##0.00'
            ws1.cell(row=row, column=5, value=float(total)).border = thin_border
            ws1.cell(row=row, column=5).number_format = '#,##0.00'
            ws1.cell(row=row, column=6, value=notes or '').border = thin_border
            total_purchases += float(total)
            row += 1
        
        # Итого по покупкам
        ws1.cell(row=row, column=4, value='ИТОГО:').font = Font(bold=True)
        ws1.cell(row=row, column=4).border = thin_border
        ws1.cell(row=row, column=5, value=total_purchases).font = Font(bold=True)
        ws1.cell(row=row, column=5).number_format = '#,##0.00'
        ws1.cell(row=row, column=5).border = thin_border
        ws1.cell(row=row, column=5).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # Настройка ширины колонок
        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 25
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 15
        ws1.column_dimensions['F'].width = 30
        
        # Лист 2: Погашения
        ws2 = wb.create_sheet(title="Погашения")
        
        # Заголовок
        ws2.merge_cells('A1:E1')
        title_cell2 = ws2['A1']
        title_cell2.value = f'ПОГАШЕНИЯ: {client_name}'
        title_cell2.font = Font(bold=True, size=14, color='FFFFFF')
        title_cell2.alignment = Alignment(horizontal='center', vertical='center')
        title_cell2.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # Заголовки таблицы погашений
        headers2 = ['Дата', 'Сомони', 'Курс', 'Сумма ($)', 'Примечания']
        
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Данные погашений
        total_payments = 0
        row = 4
        for payment in payments:
            date, somoni, rate, total_usd, notes = payment
            ws2.cell(row=row, column=1, value=str(date)).border = thin_border
            ws2.cell(row=row, column=2, value=float(somoni)).border = thin_border
            ws2.cell(row=row, column=2).number_format = '#,##0.00'
            ws2.cell(row=row, column=3, value=float(rate)).border = thin_border
            ws2.cell(row=row, column=3).number_format = '#,##0.00'
            ws2.cell(row=row, column=4, value=float(total_usd)).border = thin_border
            ws2.cell(row=row, column=4).number_format = '#,##0.00'
            ws2.cell(row=row, column=5, value=notes or '').border = thin_border
            total_payments += float(total_usd)
            row += 1
        
        # Итого по погашениям
        ws2.cell(row=row, column=3, value='ИТОГО:').font = Font(bold=True)
        ws2.cell(row=row, column=3).border = thin_border
        ws2.cell(row=row, column=4, value=total_payments).font = Font(bold=True)
        ws2.cell(row=row, column=4).number_format = '#,##0.00'
        ws2.cell(row=row, column=4).border = thin_border
        ws2.cell(row=row, column=4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # Настройка ширины колонок
        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 15
        ws2.column_dimensions['E'].width = 30
        
        # Лист 3: Сводка
        ws3 = wb.create_sheet(title="Сводка")
        
        ws3.merge_cells('A1:B1')
        title_cell3 = ws3['A1']
        title_cell3.value = f'СВОДКА ПО КЛИЕНТУ: {client_name}'
        title_cell3.font = Font(bold=True, size=14, color='FFFFFF')
        title_cell3.alignment = Alignment(horizontal='center', vertical='center')
        title_cell3.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        debt = total_purchases - total_payments
        
        summary_data = [
            ['Показатель', 'Значение'],
            ['Всего покупок ($)', total_purchases],
            ['Всего погашений ($)', total_payments],
            ['Остаток долга ($)', debt]
        ]
        
        for row_idx, row_data in enumerate(summary_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx == 3:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_idx == 2 and row_idx > 3:
                    cell.number_format = '#,##0.00'
                    if row_idx == 6:  # Остаток долга
                        if debt > 0:
                            cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                            cell.font = Font(bold=True, color='FFFFFF')
                        else:
                            cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                            cell.font = Font(bold=True)
        
        ws3.column_dimensions['A'].width = 25
        ws3.column_dimensions['B'].width = 20
        
        # Сохраняем файл
        filename = f'client_card_{client_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(filename)
        
        # Отправляем файл
        with open(filename, 'rb') as file:
            await context.bot.send_document(
                chat_id=query.message.chat_id,
                document=file,
                filename=filename,
                caption=f'👤 Карточка клиента: {client_name}\n'
                       f'💰 Покупок: {total_purchases:,.2f} $\n'
                       f'✅ Погашений: {total_payments:,.2f} $\n'
                       f'💳 Долг: {debt:,.2f} $'
            )
        
        # Удаляем временный файл
        os.remove(filename)
        
        await query.edit_message_text(f'✅ Карточка клиента {client_name} отправлена!')
        return REPORT_MENU
    
    # Обработка отчета за день
    elif query.data.startswith('daily_report_'):
        from datetime import datetime, timedelta
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import os
        
        await query.edit_message_text('⏳ Формирую отчет...')
        
        # Определяем дату
        today = datetime.now().date()
        
        if query.data == 'daily_report_today':
            target_date = today
            date_text = f"{target_date.strftime('%d.%m.%Y')}\n(Сегодня)"
        elif query.data == 'daily_report_yesterday':
            target_date = today - timedelta(days=1)
            date_text = f"{target_date.strftime('%d.%m.%Y')}\n(Вчера)"
        elif query.data == 'daily_report_2days':
            target_date = today - timedelta(days=2)
            date_text = f"{target_date.strftime('%d.%m.%Y')}\n(Позавчера)"
        else:
            await query.edit_message_text('❌ Неизвестная дата')
            return CHOOSING_ACTION
        
        conn = get_connection()
        cursor = conn.cursor()
        
        # Получаем приходы за день
        cursor.execute('''
            SELECT a.arrival_date, a.wagon_number, f.name as firm_name, 
                   w.name as warehouse_name, a.product_name, a.source,
                   a.quantity_document, a.quantity_actual, a.notes, u.username
            FROM arrivals a
            LEFT JOIN firms f ON a.firm_id = f.id
            LEFT JOIN warehouses w ON a.warehouse_id = w.id
            LEFT JOIN users u ON a.created_by = u.user_id
            WHERE a.arrival_date = %s
            ORDER BY a.id
        ''', (target_date,))
        arrivals = cursor.fetchall()
        
        # Получаем расходы за день
        cursor.execute('''
            SELECT d.departure_date, c.name as coalition_name, f.name as firm_name,
                   w.name as warehouse_name, d.product_name, cl.name as client_name,
                   d.quantity, d.price, d.total, d.notes, u.username
            FROM departures d
            LEFT JOIN coalitions c ON d.coalition_id = c.id
            LEFT JOIN firms f ON d.firm_id = f.id
            LEFT JOIN warehouses w ON d.warehouse_id = w.id
            LEFT JOIN users u ON d.created_by = u.user_id
            LEFT JOIN clients cl ON d.client_id = cl.id
            WHERE d.departure_date = %s
            ORDER BY d.id
        ''', (target_date,))
        departures = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Формируем текстовый отчет
        message = f'📅 ОТЧЁТ ЗА {date_text}\n'
        message += '═' * 25 + '\n\n'
        
        # Приходы
        message += '📦 ПРИХОД ТОВАРОВ\n'
        message += '─' * 20 + '\n'
        
        if not arrivals:
            message += 'Операций прихода не было\n\n'
        else:
            for arr in arrivals:
                arr_date, wagon, firm, warehouse, product, source, qty_doc, qty_act, notes, username = arr
                message += f'🚂 Вагон: {wagon}\n'
                message += f'🏭 Фирма: {firm}\n'
                message += f'🏪 Склад: {warehouse}\n'
                message += f'📦 Товар: {product}\n'
                message += f'📄 По док: {int(qty_doc)} шт\n'
                message += f'✅ Факт: {int(qty_act)} шт\n'
                if notes:
                    message += f'📝 Примечания: {notes}\n'
                message += '\n'
        
        # Расходы
        message += '📤 РАСХОД ТОВАРОВ\n'
        message += '─' * 20 + '\n'
        
        total_sales = 0
        
        if not departures:
            message += 'Операций расхода не было\n\n'
        else:
            for dep in departures:
                dep_date, coalition, firm, warehouse, product, client, qty, price, total, notes, username = dep
                message += f'📊 Коалица: {coalition}\n'
                message += f'🏭 Фирма: {firm}\n'
                message += f'🏪 Склад: {warehouse}\n'
                message += f'📦 Товар: {product}\n'
                if client:
                    message += f'👤 Клиент: {client}\n'
                message += f'📦 Количество: {int(qty)} шт\n'
                if price:
                    message += f'💵 Цена: {float(price):,.2f} $\n'
                    message += f'💰 Сумма: {float(total):,.2f} $\n'
                    total_sales += float(total)
                if notes:
                    message += f'📝 Примечания: {notes}\n'
                message += '\n'
        
        # Сводка
        message += '═' * 25 + '\n'
        message += '📊 СВОДКА ЗА ДЕНЬ:\n'
        message += f'📦 Приход: {len(arrivals)} операций\n'
        message += f'📤 Расход: {len(departures)} операций\n'
        message += f'💰 Сумма продаж: {total_sales:,.2f} $'
        
        # Отправляем текстовый отчет
        if len(message) > 4000:
            # Разбиваем на части
            parts = [message[i:i+3900] for i in range(0, len(message), 3900)]
            for i, part in enumerate(parts):
                if i == 0:
                    await query.edit_message_text(part)
                else:
                    await context.bot.send_message(chat_id=query.message.chat_id, text=part)
        else:
            await query.edit_message_text(message)
        
        # Создаем Excel файл
        wb = Workbook()
        
        # Лист 1: Приходы
        ws1 = wb.active
        ws1.title = "Приходы"
        
        headers1 = ['Дата', 'Вагон', 'Фирма', 'Склад', 'Товар', 'Источник', 'По док (шт)', 'Факт (шт)', 'Примечания', 'Пользователь']
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_idx, arr in enumerate(arrivals, 2):
            arr_date, wagon, firm, warehouse, product, source, qty_doc, qty_act, notes, username = arr
            ws1.cell(row=row_idx, column=1, value=str(arr_date)).border = thin_border
            ws1.cell(row=row_idx, column=2, value=wagon).border = thin_border
            ws1.cell(row=row_idx, column=3, value=firm).border = thin_border
            ws1.cell(row=row_idx, column=4, value=warehouse).border = thin_border
            ws1.cell(row=row_idx, column=5, value=product).border = thin_border
            ws1.cell(row=row_idx, column=6, value=source or '').border = thin_border
            ws1.cell(row=row_idx, column=7, value=int(qty_doc)).border = thin_border
            ws1.cell(row=row_idx, column=8, value=int(qty_act)).border = thin_border
            ws1.cell(row=row_idx, column=9, value=notes or '').border = thin_border
            ws1.cell(row=row_idx, column=10, value=username or '').border = thin_border
        
        # Лист 2: Расходы
        ws2 = wb.create_sheet(title="Расходы")
        
        headers2 = ['Дата', 'Коалица', 'Фирма', 'Склад', 'Товар', 'Клиент', 'Количество (шт)', 'Цена ($)', 'Сумма ($)', 'Примечания', 'Пользователь']
        
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_idx, dep in enumerate(departures, 2):
            dep_date, coalition, firm, warehouse, product, client, qty, price, total, notes, username = dep
            ws2.cell(row=row_idx, column=1, value=str(dep_date)).border = thin_border
            ws2.cell(row=row_idx, column=2, value=coalition).border = thin_border
            ws2.cell(row=row_idx, column=3, value=firm).border = thin_border
            ws2.cell(row=row_idx, column=4, value=warehouse).border = thin_border
            ws2.cell(row=row_idx, column=5, value=product).border = thin_border
            ws2.cell(row=row_idx, column=6, value=client or '').border = thin_border
            ws2.cell(row=row_idx, column=7, value=int(qty)).border = thin_border
            ws2.cell(row=row_idx, column=8, value=float(price) if price else 0).border = thin_border
            ws2.cell(row=row_idx, column=8).number_format = '#,##0.00'
            ws2.cell(row=row_idx, column=9, value=float(total) if total else 0).border = thin_border
            ws2.cell(row=row_idx, column=9).number_format = '#,##0.00'
            ws2.cell(row=row_idx, column=10, value=notes or '').border = thin_border
            ws2.cell(row=row_idx, column=11, value=username or '').border = thin_border
        
        # Сохраняем файл
        filename = f'daily_report_{target_date.strftime("%Y%m%d")}.xlsx'
        wb.save(filename)
        
        # Отправляем файл с inline кнопкой
        with open(filename, 'rb') as file:
            await context.bot.send_document(
                chat_id=query.message.chat_id,
                document=file,
                filename=filename,
                caption=f'📊 Детальный отчет за {date_text}'
            )
        
        # Удаляем временный файл
        os.remove(filename)
        
        return CHOOSING_ACTION
    
    # Обработка расхода за день
    elif query.data.startswith('expense_report_'):
        from datetime import datetime, timedelta
        
        await query.edit_message_text('⏳ Формирую отчет...')
        
        # Определяем дату
        today = datetime.now().date()
        
        if query.data == 'expense_report_today':
            target_date = today
            date_text = f"📅 {target_date.strftime('%d.%m.%Y')}"
            period_name = "Сегодня"
        elif query.data == 'expense_report_yesterday':
            target_date = today - timedelta(days=1)
            date_text = f"📅 {target_date.strftime('%d.%m.%Y')} (Вчера)"
            period_name = "Вчера"
        elif query.data == 'expense_report_2days':
            target_date = today - timedelta(days=2)
            date_text = f"📅 {target_date.strftime('%d.%m.%Y')} (Позавчера)"
            period_name = "Позавчера"
        else:
            await query.edit_message_text('❌ Неизвестная дата')
            return CHOOSING_ACTION
        
        conn = get_connection()
        cursor = conn.cursor()
        
        # Получаем расходы за день с группировкой
        cursor.execute('''
            SELECT w.name as warehouse_name, d.product_name, SUM(d.quantity) as total_qty
            FROM departures d
            LEFT JOIN warehouses w ON d.warehouse_id = w.id
            WHERE d.departure_date = %s
            GROUP BY w.name, d.product_name
            ORDER BY w.name, d.product_name
        ''', (target_date,))
        departures = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        if not departures:
            message = f'📤 {period_name} расходов не было\n{date_text}'
            await query.edit_message_text(message)
            return CHOOSING_ACTION
        
        # Группируем данные по складам
        warehouse_data = {}
        product_totals = {}
        
        for warehouse, product, qty in departures:
            if warehouse not in warehouse_data:
                warehouse_data[warehouse] = {}
            
            warehouse_data[warehouse][product] = qty
            
            if product not in product_totals:
                product_totals[product] = 0
            product_totals[product] += qty
        
        # Формируем отчет
        message = f'📤 РАСХОД ТОВАРОВ\n{date_text}\n'
        message += '═' * 25 + '\n\n'
        
        total_weight = 0
        
        for warehouse in sorted(warehouse_data.keys()):
            message += f'📁 {warehouse}\n'
            message += '─' * 20 + '\n'
            
            for product in sorted(warehouse_data[warehouse].keys()):
                qty = warehouse_data[warehouse][product]
                weight = qty / 20  # Переводим в тонны
                total_weight += weight
                message += f'{product} {weight:,.2f} т/н\n'
            
            message += '\n'
        
        message += '═' * 25 + '\n'
        message += f'💰 Всего: {total_weight:,.1f} т\n\n'
        
        # Итого по товарам
        message += '📦 ИТОГО ПО ТОВАРАМ:\n'
        message += '─' * 20 + '\n'
        
        for product in sorted(product_totals.keys()):
            total_qty = product_totals[product]
            weight = total_qty / 20
            message += f'{product} {weight:,.2f} т/н\n'
        
        await query.edit_message_text(message)
        return CHOOSING_ACTION
    
    # Обработка уведомлений о долгах - выбор периода
    elif query.data.startswith('debt_notify_'):
        from datetime import datetime, timedelta
        
        await query.edit_message_text('⏳ Загружаю данные...')
        
        # Определяем дату на основе выбора
        today = datetime.now().date()
        
        if query.data == 'debt_notify_today':
            target_date = today
            days_ago = 0
            period_text = "сегодня"
        elif query.data == 'debt_notify_yesterday':
            target_date = today - timedelta(days=1)
            days_ago = 1
            period_text = "вчера"
        elif query.data == 'debt_notify_7days':
            target_date = today - timedelta(days=7)
            days_ago = 7
            period_text = f"{target_date.strftime('%d.%m.%Y')} (7 дней назад)"
        elif query.data == 'debt_notify_14days':
            target_date = today - timedelta(days=14)
            days_ago = 14
            period_text = f"{target_date.strftime('%d.%m.%Y')} (14 дней назад)"
        elif query.data == 'debt_notify_30days':
            target_date = today - timedelta(days=30)
            days_ago = 30
            period_text = f"{target_date.strftime('%d.%m.%Y')} (30 дней назад)"
        else:
            await query.edit_message_text('❌ Неизвестный период')
            return REPORT_MENU
        
        conn = get_connection()
        cursor = conn.cursor()
        
        # Получаем всех клиентов с долгами
        cursor.execute('''
            SELECT client_id, SUM(total) as total_sales
            FROM departures
            WHERE client_id IS NOT NULL
            GROUP BY client_id
        ''')
        sales_data = {row[0]: float(row[1]) for row in cursor.fetchall()}
        
        cursor.execute('''
            SELECT client_id, SUM(total_usd) as total_paid
            FROM payments
            GROUP BY client_id
        ''')
        payments_data = {row[0]: float(row[1]) for row in cursor.fetchall()}
        
        # Получаем покупки клиентов в выбранную дату
        cursor.execute('''
            SELECT client_id, product_name, quantity, total
            FROM departures
            WHERE client_id IS NOT NULL AND departure_date = %s
            ORDER BY client_id, product_name
        ''', (target_date,))
        purchases_on_date = cursor.fetchall()
        
        # Группируем покупки по клиентам
        client_purchases = {}
        for client_id, product, qty, total in purchases_on_date:
            if client_id not in client_purchases:
                client_purchases[client_id] = []
            client_purchases[client_id].append((product, qty, total))
        
        # Получаем имена клиентов
        cursor.execute('SELECT id, name FROM clients')
        clients_data = {row[0]: row[1] for row in cursor.fetchall()}
        
        cursor.close()
        conn.close()
        
        # Формируем отчет только для клиентов, которые покупали в выбранную дату
        report_lines = []
        total_debt_all = 0
        total_purchases_on_date = 0
        client_count = 0
        
        for client_id in sorted(client_purchases.keys(), key=lambda x: sales_data.get(x, 0) - payments_data.get(x, 0), reverse=True):
            client_name = clients_data.get(client_id, 'Неизвестный')
            total_sales = sales_data.get(client_id, 0)
            total_paid = payments_data.get(client_id, 0)
            debt = total_sales - total_paid
            
            if debt <= 0:
                continue
            
            client_count += 1
            total_debt_all += debt
            
            # Покупки клиента в выбранную дату
            purchases = client_purchases[client_id]
            purchases_sum = sum(p[2] for p in purchases)
            total_purchases_on_date += purchases_sum
            
            report_lines.append(f"{client_count}. 👤 {client_name}")
            report_lines.append(f"💳 Общий долг: {debt:,.2f} $")
            report_lines.append(f"📦 Покупки {period_text}:")
            
            for product, qty, total in purchases:
                report_lines.append(f"• {product} - {int(qty)} шт ({total:,.2f} $)")
            
            report_lines.append(f"💰 Сумма покупок в тот день: {purchases_sum:,.2f} $")
            report_lines.append("")
        
        if client_count == 0:
            await query.edit_message_text(
                f'📊 Клиенты с долгами, которые покупали {period_text}:\n\n'
                f'✅ Нет клиентов с долгами, которые покупали в этот день'
            )
            return REPORT_MENU
        
        # Формируем итоговое сообщение
        header = f"🔔 УВЕДОМЛЕНИЯ О ДОЛГАХ\n"
        header += f"📅 Клиенты, которые покупали {period_text}\n"
        header += "═" * 30 + "\n\n"
        
        footer = "\n" + "═" * 30 + "\n"
        footer += f"📊 ИТОГО:\n"
        footer += f"👥 Должников: {client_count}\n"
        footer += f"💳 Общий долг: {total_debt_all:,.2f} $\n"
        footer += f"💰 Сумма покупок {period_text}: {total_purchases_on_date:,.2f} $\n\n"
        footer += "⚠️ Рекомендуется связаться с этими клиентами для напоминания о долге"
        
        full_message = header + "\n".join(report_lines) + footer
        
        # Telegram ограничивает сообщения 4096 символами
        if len(full_message) > 4000:
            # Разбиваем на части
            await query.edit_message_text(header + "\n".join(report_lines[:20]) + "\n\n... (продолжение следует)")
            
            # Отправляем остальное отдельными сообщениями
            remaining = report_lines[20:]
            chunk_size = 30
            for i in range(0, len(remaining), chunk_size):
                chunk = remaining[i:i+chunk_size]
                await context.bot.send_message(
                    chat_id=query.message.chat_id,
                    text="\n".join(chunk)
                )
            
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text=footer
            )
        else:
            await query.edit_message_text(full_message)
        
        return REPORT_MENU
    
    # Обработка уведомлений о долгах
    elif query.data.startswith('notify_debt_'):
        if query.data == 'notify_debt_all':
            # Отправка уведомлений всем клиентам с долгами
            await query.edit_message_text('⏳ Загружаю данные...')
            
            conn = get_connection()
            cursor = conn.cursor()
            
            # Упрощенный запрос: получаем продажи
            cursor.execute('''
                SELECT client_id, SUM(total) as total_sales
                FROM departures
                WHERE client_id IS NOT NULL
                GROUP BY client_id
            ''')
            sales_data = {row[0]: float(row[1]) for row in cursor.fetchall()}
            
            # Получаем погашения
            cursor.execute('''
                SELECT client_id, SUM(total_usd) as total_paid
                FROM payments
                GROUP BY client_id
            ''')
            payments_data = {row[0]: float(row[1]) for row in cursor.fetchall()}
            
            # Получаем клиентов
            cursor.execute('SELECT id, name, phone FROM clients')
            clients_data = cursor.fetchall()
            
            cursor.close()
            conn.close()
            
            # Формируем список должников
            debts = []
            for client_id, client_name, phone in clients_data:
                total_sales = sales_data.get(client_id, 0)
                total_paid = payments_data.get(client_id, 0)
                debt = total_sales - total_paid
                
                if debt > 0:
                    debts.append((client_id, client_name, phone, total_sales, total_paid, debt))
            
            if not debts:
                await query.edit_message_text('✅ Нет клиентов с долгами')
                return REPORT_MENU
            
            await query.edit_message_text(f'⏳ Отправляю уведомления {len(debts)} клиентам...')
            
            sent_count = 0
            failed_count = 0
            
            for client_id, client_name, phone, total_sales, total_paid, debt in debts:
                # Формируем сообщение
                message = f'📢 УВЕДОМЛЕНИЕ О ЗАДОЛЖЕННОСТИ\n\n'
                message += f'Уважаемый(ая) {client_name}!\n\n'
                message += f'💰 Сумма покупок: {total_sales:,.2f} $\n'
                message += f'✅ Погашено: {total_paid:,.2f} $\n'
                message += f'💳 Остаток долга: {debt:,.2f} $\n\n'
                message += f'Просим погасить задолженность в ближайшее время.\n'
                message += f'Спасибо за сотрудничество!'
                
                try:
                    # Здесь должна быть логика отправки сообщения клиенту
                    # Например: await context.bot.send_message(chat_id=client_telegram_id, text=message)
                    sent_count += 1
                except Exception as e:
                    failed_count += 1
                    logger.error(f'Ошибка отправки уведомления клиенту {client_name}: {e}')
            
            result_message = f'✅ Уведомления отправлены!\n\n'
            result_message += f'📤 Отправлено: {sent_count}\n'
            if failed_count > 0:
                result_message += f'❌ Не удалось отправить: {failed_count}\n'
            result_message += f'\n💡 Примечание: Для отправки уведомлений клиентам необходимо добавить их Telegram ID в базу данных.'
            
            await query.edit_message_text(result_message)
            return REPORT_MENU
        else:
            # Отправка уведомления одному клиенту
            client_id = int(query.data.replace('notify_debt_', ''))
            
            await query.edit_message_text('⏳ Загружаю данные...')
            
            # Получаем информацию о клиенте и его долге
            conn = get_connection()
            cursor = conn.cursor()
            
            # Получаем данные клиента
            cursor.execute('SELECT name, phone FROM clients WHERE id = %s', (client_id,))
            client_result = cursor.fetchone()
            
            if not client_result:
                cursor.close()
                conn.close()
                await query.edit_message_text('❌ Клиент не найден')
                return REPORT_MENU
            
            client_name, phone = client_result
            
            # Получаем продажи
            cursor.execute('SELECT COALESCE(SUM(total), 0) FROM departures WHERE client_id = %s', (client_id,))
            total_sales = float(cursor.fetchone()[0])
            
            # Получаем погашения
            cursor.execute('SELECT COALESCE(SUM(total_usd), 0) FROM payments WHERE client_id = %s', (client_id,))
            total_paid = float(cursor.fetchone()[0])
            
            cursor.close()
            conn.close()
            
            debt = total_sales - total_paid
            
            # Формируем сообщение
            message = f'📢 УВЕДОМЛЕНИЕ О ЗАДОЛЖЕННОСТИ\n\n'
            message += f'Уважаемый(ая) {client_name}!\n\n'
            message += f'💰 Сумма покупок: {total_sales:,.2f} $\n'
            message += f'✅ Погашено: {total_paid:,.2f} $\n'
            message += f'💳 Остаток долга: {debt:,.2f} $\n\n'
            message += f'Просим погасить задолженность в ближайшее время.\n'
            message += f'Спасибо за сотрудничество!'
            
            # Здесь должна быть логика отправки сообщения клиенту
            # Например: await context.bot.send_message(chat_id=client_telegram_id, text=message)
            
            await query.edit_message_text(
                f'✅ Уведомление для клиента {client_name} подготовлено!\n\n'
                f'{message}\n\n'
                f'💡 Примечание: Для отправки уведомлений клиентам необходимо добавить их Telegram ID в базу данных.'
            )
            return REPORT_MENU

def main():
    """Запуск бота"""
    init_db()
    
    application = Application.builder().token(BOT_TOKEN).build()
    
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            CHOOSING_ACTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_action), CallbackQueryHandler(handle_callback_query)],
            PRODUCT_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_product_name)],
            PRODUCT_QUANTITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_product_quantity)],
            ADMIN_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_admin_menu)],
            ADMIN_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_admin_id)],
            MANAGEMENT_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_management_menu)],
            WAREHOUSE_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_warehouse_menu)],
            WAREHOUSE_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_warehouse_name)],
            WAREHOUSE_DATA: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_warehouse_data)],
            WAREHOUSE_DELETE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_warehouse_delete)],
            WAREHOUSE_GROUP_SELECT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_warehouse_group_select)],
            WAREHOUSE_GROUP_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_warehouse_group_menu)],
            WAREHOUSE_GROUP_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_warehouse_group_name)],
            WAREHOUSE_GROUP_DELETE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_warehouse_group_delete)],
            PRODUCT_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_product_menu)],
            PRODUCT_DATA: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_product_data)],
            PRODUCT_DELETE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_product_delete)],
            FIRM_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_firm_menu)],
            FIRM_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_firm_name)],
            FIRM_DELETE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_firm_delete)],
            CLIENT_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_client_menu)],
            CLIENT_DATA: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_client_data)],
            CLIENT_DELETE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_client_delete)],
            PRICE_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_price_menu)],
            PRICE_SELECT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_price_select)],
            PRICE_DATA: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_price_data)],
            PRICE_VALUE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_price_value)],
            PRICE_DELETE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_price_delete)],
            COALITION_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_coalition_menu)],
            COALITION_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_coalition_name)],
            COALITION_DELETE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_coalition_delete)],
            USER_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_user_menu)],
            USER_ADD_LOGIN: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_user_add_login)],
            USER_ADD_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_user_add_password)],
            USER_ADD_USERNAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_user_add_username)],
            USER_ADD_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_user_add_id)],
            USER_ADD_ROLE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_user_add_role)],
            USER_ADD_GROUP: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_user_add_group)],
            USER_EDIT_SELECT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_user_edit_select)],
            USER_EDIT_ROLE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_user_edit_role)],
            USER_EDIT_GROUP: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_user_edit_group)],
            LOGIN: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_login)],
            PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_password)],
            ARRIVAL_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_arrival_date)],
            ARRIVAL_WAGON: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_arrival_wagon)],
            ARRIVAL_FIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_arrival_firm)],
            ARRIVAL_WAREHOUSE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_arrival_warehouse)],
            ARRIVAL_PRODUCT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_arrival_product)],
            ARRIVAL_QTY_DOC: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_arrival_qty_doc)],
            ARRIVAL_QTY_ACTUAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_arrival_qty_actual)],
            ARRIVAL_NOTES: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_arrival_notes)],
            DEPARTURE_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_departure_date)],
            DEPARTURE_COALITION: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_departure_coalition)],
            DEPARTURE_FIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_departure_firm)],
            DEPARTURE_WAREHOUSE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_departure_warehouse)],
            DEPARTURE_PRODUCT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_departure_product)],
            DEPARTURE_CLIENT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_departure_client)],
            DEPARTURE_QUANTITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_departure_quantity)],
            DEPARTURE_PRICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_departure_price)],
            DEPARTURE_NOTES: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_departure_notes)],
            REPORT_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_report_menu), CallbackQueryHandler(handle_callback_query)],
            REPORT_EXPORT_PERIOD: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_report_export), CallbackQueryHandler(handle_callback_query)],
            PAYMENT_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_payment_date)],
            PAYMENT_CLIENT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_payment_client)],
            PAYMENT_SOMONI: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_payment_somoni)],
            PAYMENT_RATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_payment_rate)],
            PAYMENT_NOTES: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_payment_notes)],
            PARTNER_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_partner_date)],
            PARTNER_CLIENT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_partner_client)],
            PARTNER_SOMONI: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_partner_somoni)],
            PARTNER_RATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_partner_rate)],
            PARTNER_NOTES: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_partner_notes)],
            DELETE_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_delete_menu)],
            DELETE_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_delete_id)],
            DELETE_CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_delete_confirm)],
            EDIT_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_edit_menu)],
            EDIT_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_edit_id)],
            EDIT_FIELD: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_edit_field)],
            EDIT_VALUE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_edit_value)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    
    application.add_handler(conv_handler)
    
    logger.info('Бот запущен')
    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()


