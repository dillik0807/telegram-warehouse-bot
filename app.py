from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from functools import wraps
from datetime import datetime, timedelta
import hashlib
from database import *
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this-in-production'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)

# Добавляем функции в контекст шаблонов
@app.context_processor
def utility_processor():
    def to_float(value):
        """Преобразует Decimal в float для использования в шаблонах"""
        if value is None:
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0
    
    return dict(now=datetime.now, to_float=to_float)

# Декоратор для проверки авторизации
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user_id = session.get('user_id')
        if user_id is None or not isinstance(user_id, int):
            session.clear()
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Декоратор для проверки прав администратора
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user_id = session.get('user_id')
        if user_id is None or not isinstance(user_id, int):
            session.clear()
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Доступ запрещен. Только для администраторов.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
@login_required
def index():
    """Главная страница"""
    # Проверка типа user_id для отладки
    user_id = session.get('user_id')
    if not isinstance(user_id, int):
        # Если user_id не integer, очищаем сессию и перенаправляем на логин
        session.clear()
        flash('Сессия устарела. Пожалуйста, войдите снова.', 'warning')
        return redirect(url_for('login'))
    
    return render_template('index.html', 
                         username=session.get('username'),
                         role=session.get('role'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Страница входа"""
    # Очищаем сессию при открытии страницы логина
    session.clear()
    
    if request.method == 'POST':
        login_name = request.form.get('login')
        password = request.form.get('password')
        
        result = verify_login(login_name, password)
        
        if result[0]:  # Если первый элемент True, значит логин успешен
            success, user_id, username, role, warehouse_group_id = result
            
            # Проверяем типы данных перед сохранением
            if not isinstance(user_id, int):
                flash(f'Ошибка: неверный тип user_id ({type(user_id)})', 'error')
                return render_template('login.html')
            
            session.permanent = True
            session['user_id'] = int(user_id)  # Явное приведение к int
            session['username'] = str(username)
            session['role'] = str(role)
            session['warehouse_group_id'] = int(warehouse_group_id) if warehouse_group_id else None
            
            return redirect(url_for('index'))
        else:
            flash('Неверный логин или пароль', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Выход из системы"""
    session.clear()
    return redirect(url_for('login'))

# ============= ПРИХОДЫ =============

@app.route('/arrivals')
@login_required
def arrivals():
    """Список приходов"""
    user_id = session.get('user_id')
    role = session.get('role')
    warehouse_group_id = session.get('warehouse_group_id')
    
    # Получаем склады пользователя
    user_warehouses = get_user_warehouses(user_id)
    warehouse_ids = [wh[0] for wh in user_warehouses] if user_warehouses else []
    
    # Получаем приходы
    if role == 'admin':
        arrivals_list = get_arrivals(limit=100)
    else:
        arrivals_list = []
        for wh_id in warehouse_ids:
            arrivals_list.extend(get_arrivals(limit=100, warehouse_id=wh_id))
    
    return render_template('arrivals.html', arrivals=arrivals_list)

@app.route('/arrivals/add', methods=['GET', 'POST'])
@login_required
def add_arrival_page():
    """Добавить приход"""
    if request.method == 'POST':
        arrival_date = request.form.get('arrival_date')
        wagon_number = request.form.get('wagon_number')
        firm_id = request.form.get('firm_id')
        warehouse_id = request.form.get('warehouse_id')
        product_name = request.form.get('product_name')
        source = request.form.get('source')
        qty_doc = int(request.form.get('qty_doc'))
        qty_actual = int(request.form.get('qty_actual'))
        notes = request.form.get('notes')
        
        user_id = session.get('user_id')
        
        # Отладочный вывод
        print(f"DEBUG: user_id = {user_id}, type = {type(user_id)}")
        
        # Проверяем тип
        if not isinstance(user_id, int):
            flash(f'Ошибка сессии: user_id имеет неверный тип ({type(user_id)}). Пожалуйста, выйдите и войдите снова.', 'error')
            return redirect(url_for('logout'))
        
        success, result = add_arrival(
            arrival_date, wagon_number, firm_id, warehouse_id,
            product_name, source, qty_doc, qty_actual, notes,
            user_id
        )
        
        if success:
            flash('Приход успешно добавлен', 'success')
            return redirect(url_for('arrivals'))
        else:
            flash(f'Ошибка: {result}', 'error')
    
    firms = get_firms()
    warehouses = get_warehouses()
    
    return render_template('add_arrival.html', firms=firms, warehouses=warehouses)

# ============= РАСХОДЫ =============

@app.route('/departures')
@login_required
def departures():
    """Список расходов"""
    user_id = session.get('user_id')
    role = session.get('role')
    
    user_warehouses = get_user_warehouses(user_id)
    warehouse_ids = [wh[0] for wh in user_warehouses] if user_warehouses else []
    
    if role == 'admin':
        departures_list = get_departures(limit=100)
    else:
        departures_list = []
        for wh_id in warehouse_ids:
            departures_list.extend(get_departures(limit=100, warehouse_id=wh_id))
    
    return render_template('departures.html', departures=departures_list)

@app.route('/departures/add', methods=['GET', 'POST'])
@login_required
def add_departure_page():
    """Добавить расход"""
    if request.method == 'POST':
        departure_date = request.form.get('departure_date')
        coalition_id = request.form.get('coalition_id')
        firm_id = request.form.get('firm_id')
        warehouse_id = request.form.get('warehouse_id')
        product_name = request.form.get('product_name')
        client_id = request.form.get('client_id') or None
        quantity = int(request.form.get('quantity'))
        price = float(request.form.get('price'))
        notes = request.form.get('notes')
        
        user_id = session.get('user_id')
        
        # Проверяем тип
        if not isinstance(user_id, int):
            flash(f'Ошибка сессии: user_id имеет неверный тип. Пожалуйста, выйдите и войдите снова.', 'error')
            return redirect(url_for('logout'))
        
        success, result = add_departure(
            departure_date, coalition_id, firm_id, warehouse_id,
            product_name, client_id, quantity, price, notes,
            user_id
        )
        
        if success:
            flash('Расход успешно добавлен', 'success')
            return redirect(url_for('departures'))
        else:
            flash(f'Ошибка: {result}', 'error')
    
    coalitions = get_coalitions()
    firms = get_firms()
    warehouses = get_warehouses()
    clients = get_clients()
    
    return render_template('add_departure.html', 
                         coalitions=coalitions,
                         firms=firms,
                         warehouses=warehouses,
                         clients=clients)

# ============= ПОГАШЕНИЯ =============

@app.route('/payments')
@login_required
def payments():
    """Список погашений"""
    payments_list = get_payments(limit=100)
    return render_template('payments.html', payments=payments_list)

@app.route('/payments/add', methods=['GET', 'POST'])
@login_required
def add_payment_page():
    """Добавить погашение"""
    if request.method == 'POST':
        payment_date = request.form.get('payment_date')
        client_id = request.form.get('client_id')
        somoni = float(request.form.get('somoni'))
        rate = float(request.form.get('rate'))
        notes = request.form.get('notes')
        
        user_id = session.get('user_id')
        
        # Проверяем тип
        if not isinstance(user_id, int):
            flash(f'Ошибка сессии: user_id имеет неверный тип. Пожалуйста, выйдите и войдите снова.', 'error')
            return redirect(url_for('logout'))
        
        success, result = add_payment(
            payment_date, client_id, somoni, 0, rate, notes,
            user_id
        )
        
        if success:
            flash('Погашение успешно добавлено', 'success')
            return redirect(url_for('payments'))
        else:
            flash(f'Ошибка: {result}', 'error')
    
    clients = get_clients()
    return render_template('add_payment.html', clients=clients)

# ============= ПАРТНЕРЫ =============

@app.route('/partners')
@login_required
def partners():
    """Список партнеров"""
    partners_list = get_partners(limit=100)
    return render_template('partners.html', partners=partners_list)

@app.route('/partners/add', methods=['GET', 'POST'])
@login_required
def add_partner_page():
    """Добавить партнера"""
    if request.method == 'POST':
        partner_date = request.form.get('partner_date')
        client_id = int(request.form.get('client_id'))
        somoni = float(request.form.get('somoni'))
        rate = float(request.form.get('rate'))
        notes = request.form.get('notes')
        user_id = session.get('user_id')
        
        # Вычисляем сумму в долларах
        total_usd = somoni / rate
        
        # Добавляем партнера
        success = add_partner(partner_date, client_id, somoni, rate, total_usd, notes, user_id)
        
        if success:
            flash('Партнер успешно добавлен', 'success')
            return redirect(url_for('partners'))
        else:
            flash('Ошибка при добавлении партнера', 'error')
    
    # GET запрос - показываем форму
    clients = get_clients()
    return render_template('add_partner.html', clients=clients)

# ============= ОТЧЕТЫ =============

@app.route('/reports')
@login_required
def reports():
    """Страница отчетов"""
    return render_template('reports.html')

@app.route('/reports/daily', methods=['GET'])
@login_required
def daily_report():
    """Отчет за день"""
    date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    
    user_id = session.get('user_id')
    role = session.get('role')
    
    # Получаем склады пользователя
    user_warehouses = get_user_warehouses(user_id)
    warehouse_ids = [wh[0] for wh in user_warehouses] if user_warehouses else []
    
    # Получаем приходы за день
    if role == 'admin':
        arrivals_list = get_arrivals(limit=1000, date_from=date, date_to=date)
    else:
        arrivals_list = []
        for wh_id in warehouse_ids:
            arrivals_list.extend(get_arrivals(limit=1000, warehouse_id=wh_id, date_from=date, date_to=date))
    
    # Получаем расходы за день
    if role == 'admin':
        departures_list = get_departures(limit=1000, date_from=date, date_to=date)
    else:
        departures_list = []
        for wh_id in warehouse_ids:
            departures_list.extend(get_departures(limit=1000, warehouse_id=wh_id, date_from=date, date_to=date))
    
    return render_template('daily_report.html', 
                         date=date,
                         arrivals=arrivals_list,
                         departures=departures_list)

@app.route('/reports/daily_expense', methods=['GET'])
@login_required
def daily_expense():
    """Расход за день"""
    date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    
    user_id = session.get('user_id')
    role = session.get('role')
    
    # Получаем склады пользователя
    user_warehouses = get_user_warehouses(user_id)
    warehouse_ids = [wh[0] for wh in user_warehouses] if user_warehouses else []
    
    # Получаем расходы за день
    if role == 'admin':
        departures_list = get_departures(limit=1000, date_from=date, date_to=date)
    else:
        departures_list = []
        for wh_id in warehouse_ids:
            departures_list.extend(get_departures(limit=1000, warehouse_id=wh_id, date_from=date, date_to=date))
    
    # Группируем по складам и товарам
    warehouse_data = {}
    product_totals = {}
    
    for dep in departures_list:
        dep_id, dep_date, coalition, firm, warehouse, product, qty, price, notes, username_dep, created, client, total = dep
        
        warehouse_name = warehouse if warehouse else 'Неизвестно'
        product_key = f"{firm} {product}"
        
        if warehouse_name not in warehouse_data:
            warehouse_data[warehouse_name] = {}
        
        if product_key not in warehouse_data[warehouse_name]:
            warehouse_data[warehouse_name][product_key] = 0
        
        warehouse_data[warehouse_name][product_key] += qty / 20  # В тоннах
        
        # Общий итог по товарам
        if product_key not in product_totals:
            product_totals[product_key] = 0
        product_totals[product_key] += qty / 20
    
    return render_template('daily_expense.html',
                         date=date,
                         warehouse_data=warehouse_data,
                         product_totals=product_totals)

@app.route('/reports/actual_balance', methods=['GET'])
@login_required
def actual_balance():
    """Фактический остаток"""
    user_id = session.get('user_id')
    role = session.get('role')
    
    conn = get_connection()
    cursor = conn.cursor()
    
    # Получаем склады пользователя
    user_warehouses = get_user_warehouses(user_id)
    warehouse_ids = [wh[0] for wh in user_warehouses] if user_warehouses else []
    
    # Формируем условие для фильтрации складов
    if role != 'admin' and warehouse_ids:
        warehouse_filter = f"AND warehouse_id IN ({','.join(map(str, warehouse_ids))})"
    else:
        warehouse_filter = ""
    
    # Получаем остатки с группировкой по коалициям, складам и товарам
    query = f'''
        WITH balance AS (
            SELECT 
                w.group_id,
                wg.name as coalition_name,
                a.warehouse_id,
                w.name as warehouse_name,
                a.firm_id,
                f.name as firm_name,
                a.product_name,
                SUM(a.quantity_actual) as total_qty
            FROM arrivals a
            JOIN warehouses w ON a.warehouse_id = w.id
            LEFT JOIN warehouse_groups wg ON w.group_id = wg.id
            JOIN firms f ON a.firm_id = f.id
            WHERE 1=1 {warehouse_filter}
            GROUP BY w.group_id, wg.name, a.warehouse_id, w.name, a.firm_id, f.name, a.product_name
            
            UNION ALL
            
            SELECT 
                w.group_id,
                wg.name as coalition_name,
                d.warehouse_id,
                w.name as warehouse_name,
                d.firm_id,
                f.name as firm_name,
                d.product_name,
                -SUM(d.quantity) as total_qty
            FROM departures d
            JOIN warehouses w ON d.warehouse_id = w.id
            LEFT JOIN warehouse_groups wg ON w.group_id = wg.id
            JOIN firms f ON d.firm_id = f.id
            WHERE 1=1 {warehouse_filter}
            GROUP BY w.group_id, wg.name, d.warehouse_id, w.name, d.firm_id, f.name, d.product_name
        )
        SELECT 
            group_id,
            coalition_name,
            warehouse_id,
            warehouse_name,
            firm_name,
            product_name,
            SUM(total_qty) as balance_qty
        FROM balance
        GROUP BY group_id, coalition_name, warehouse_id, warehouse_name, firm_name, product_name
        HAVING SUM(total_qty) > 0
        ORDER BY coalition_name, warehouse_name, firm_name, product_name
    '''
    
    cursor.execute(query)
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    
    # Группируем данные
    coalition_data = {}
    
    for row in results:
        group_id, coalition_name, warehouse_id, warehouse_name, firm_name, product_name, balance_qty = row
        
        coalition_name = coalition_name or 'Без коалиции'
        
        if coalition_name not in coalition_data:
            coalition_data[coalition_name] = {}
        
        if warehouse_name not in coalition_data[coalition_name]:
            coalition_data[coalition_name][warehouse_name] = {}
        
        product_key = f"{firm_name} {product_name}"
        
        if product_key not in coalition_data[coalition_name][warehouse_name]:
            coalition_data[coalition_name][warehouse_name][product_key] = 0
        
        coalition_data[coalition_name][warehouse_name][product_key] += balance_qty / 20  # В тоннах
    
    return render_template('actual_balance.html', coalition_data=coalition_data)

@app.route('/reports/client_debts')
@login_required
def client_debts():
    """Долги клиентов"""
    conn = get_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT client_id, SUM(total) as total_sales
        FROM departures
        WHERE client_id IS NOT NULL
        GROUP BY client_id
    ''')
    sales_data = {row[0]: float(row[1]) for row in cursor.fetchall()}
    
    cursor.execute('''
        SELECT client_id, SUM(total_usd) as total_paid
        FROM payments
        GROUP BY client_id
    ''')
    payments_data = {row[0]: float(row[1]) for row in cursor.fetchall()}
    
    cursor.execute('SELECT id, name FROM clients')
    clients_data = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    debts = []
    for client_id, client_name in clients_data:
        total_sales = sales_data.get(client_id, 0)
        total_paid = payments_data.get(client_id, 0)
        debt = total_sales - total_paid
        
        if debt > 0:
            debts.append({
                'client_id': client_id,
                'client_name': client_name,
                'total_sales': total_sales,
                'total_paid': total_paid,
                'debt': debt
            })
    
    debts.sort(key=lambda x: x['debt'], reverse=True)
    
    return render_template('client_debts.html', debts=debts)

# ============= УПРАВЛЕНИЕ =============

@app.route('/management')
@admin_required
def management():
    """Страница управления"""
    return render_template('management.html')

@app.route('/management/warehouses')
@admin_required
def manage_warehouses():
    """Управление складами"""
    warehouses = get_warehouses()
    groups = get_warehouse_groups()
    return render_template('manage_warehouses.html', warehouses=warehouses, groups=groups)

@app.route('/management/clients')
@admin_required
def manage_clients():
    """Управление клиентами"""
    clients = get_clients()
    return render_template('manage_clients.html', clients=clients)

@app.route('/management/clients/add', methods=['POST'])
@admin_required
def add_client_route():
    """Добавить клиента"""
    client_name = request.form.get('client_name')
    phone = request.form.get('phone')
    
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO clients (name, phone) VALUES (%s, %s)", (client_name, phone))
        conn.commit()
        flash(f'Клиент "{client_name}" успешно добавлен', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Ошибка при добавлении клиента: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('manage_clients'))

@app.route('/management/clients/delete/<int:client_id>')
@admin_required
def delete_client_route(client_id):
    """Удалить клиента"""
    conn = get_connection()
    cursor = conn.cursor()
    try:
        # Проверяем, используется ли клиент
        cursor.execute("SELECT COUNT(*) FROM departures WHERE client_id = %s", (client_id,))
        count = cursor.fetchone()[0]
        if count > 0:
            flash(f'Невозможно удалить клиента. Он используется в {count} записях.', 'error')
        else:
            cursor.execute("DELETE FROM clients WHERE id = %s", (client_id,))
            conn.commit()
            flash('Клиент успешно удален', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Ошибка при удалении клиента: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('manage_clients'))

# ============= УПРАВЛЕНИЕ СКЛАДАМИ =============

@app.route('/management/warehouses/add', methods=['POST'])
@admin_required
def add_warehouse_route():
    """Добавить склад"""
    warehouse_name = request.form.get('warehouse_name')
    address = request.form.get('address')
    group_id = request.form.get('group_id')
    
    if group_id:
        group_id = int(group_id)
    else:
        group_id = None
    
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO warehouses (name, address, group_id) VALUES (%s, %s, %s)", 
                      (warehouse_name, address, group_id))
        conn.commit()
        flash(f'Склад "{warehouse_name}" успешно добавлен', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Ошибка при добавлении склада: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('manage_warehouses'))

@app.route('/management/warehouses/delete/<int:warehouse_id>')
@admin_required
def delete_warehouse_route(warehouse_id):
    """Удалить склад"""
    conn = get_connection()
    cursor = conn.cursor()
    try:
        # Проверяем, используется ли склад
        cursor.execute("SELECT COUNT(*) FROM arrivals WHERE warehouse_id = %s", (warehouse_id,))
        count = cursor.fetchone()[0]
        if count > 0:
            flash(f'Невозможно удалить склад. Он используется в {count} записях.', 'error')
        else:
            cursor.execute("DELETE FROM warehouses WHERE id = %s", (warehouse_id,))
            conn.commit()
            flash('Склад успешно удален', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Ошибка при удалении склада: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('manage_warehouses'))

@app.route('/management/warehouse_groups/add', methods=['POST'])
@admin_required
def add_warehouse_group_route():
    """Добавить подгруппу складов"""
    group_name = request.form.get('group_name')
    
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO warehouse_groups (name) VALUES (%s)", (group_name,))
        conn.commit()
        flash(f'Подгруппа "{group_name}" успешно добавлена', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Ошибка при добавлении подгруппы: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('manage_warehouses'))

@app.route('/management/warehouse_groups/delete/<int:group_id>')
@admin_required
def delete_warehouse_group_route(group_id):
    """Удалить подгруппу складов"""
    conn = get_connection()
    cursor = conn.cursor()
    try:
        # Проверяем, используется ли подгруппа
        cursor.execute("SELECT COUNT(*) FROM warehouses WHERE group_id = %s", (group_id,))
        count = cursor.fetchone()[0]
        if count > 0:
            flash(f'Невозможно удалить подгруппу. Она используется в {count} складах.', 'error')
        else:
            cursor.execute("DELETE FROM warehouse_groups WHERE id = %s", (group_id,))
            conn.commit()
            flash('Подгруппа успешно удалена', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Ошибка при удалении подгруппы: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('manage_warehouses'))

@app.route('/management/users')
@admin_required
def manage_users():
    """Управление пользователями"""
    users = get_all_users()
    return render_template('manage_users.html', users=users)

@app.route('/management/users/add', methods=['POST'])
@admin_required
def add_user_route():
    """Добавить пользователя"""
    user_id = int(request.form.get('user_id'))
    username = request.form.get('username')
    login_name = request.form.get('login')
    password = request.form.get('password')
    role = request.form.get('role')
    warehouse_group_id = request.form.get('warehouse_group_id')
    
    if warehouse_group_id:
        warehouse_group_id = int(warehouse_group_id)
    else:
        warehouse_group_id = None
    
    # Создаем пользователя
    success, message = create_login(user_id, login_name, password, username, role, warehouse_group_id)
    
    if success:
        flash(f'Пользователь {username} успешно добавлен', 'success')
    else:
        flash(f'Ошибка: {message}', 'error')
    
    return redirect(url_for('manage_users'))

@app.route('/management/users/activate/<int:user_id>')
@admin_required
def activate_user_route(user_id):
    """Активировать пользователя"""
    activate_user(user_id)
    flash('Пользователь активирован', 'success')
    return redirect(url_for('manage_users'))

@app.route('/management/users/deactivate/<int:user_id>')
@admin_required
def deactivate_user_route(user_id):
    """Деактивировать пользователя"""
    deactivate_user(user_id)
    flash('Пользователь деактивирован', 'warning')
    return redirect(url_for('manage_users'))

@app.route('/management/users/delete/<int:user_id>')
@admin_required
def delete_user_route(user_id):
    """Удалить пользователя"""
    delete_user(user_id)
    flash('Пользователь удален', 'success')
    return redirect(url_for('manage_users'))

# ============= УПРАВЛЕНИЕ ТОВАРАМИ =============

@app.route('/management/products')
@admin_required
def manage_products():
    """Управление товарами"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT product_name FROM arrivals ORDER BY product_name")
    products = cursor.fetchall()
    conn.close()
    return render_template('manage_products.html', products=products)

@app.route('/management/products/add', methods=['POST'])
@admin_required
def add_product_route():
    """Добавить товар"""
    product_name = request.form.get('product_name')
    
    conn = get_connection()
    cursor = conn.cursor()
    try:
        # Проверяем, существует ли уже такой товар
        cursor.execute("SELECT COUNT(*) FROM arrivals WHERE product_name = %s", (product_name,))
        count = cursor.fetchone()[0]
        
        if count > 0:
            flash(f'Товар "{product_name}" уже существует в системе', 'warning')
        else:
            # Добавляем товар через фиктивную запись в arrivals с нулевым количеством
            # Это позволит товару появиться в списках выбора
            cursor.execute("""
                INSERT INTO arrivals 
                (arrival_date, wagon_number, firm_id, warehouse_id, product_name, source, quantity_doc, quantity_actual, notes, created_by)
                VALUES (CURRENT_DATE, 'SYSTEM', 1, 1, %s, 'Добавлено вручную', 0, 0, 'Товар добавлен через веб-интерфейс', %s)
            """, (product_name, session.get('user_id')))
            conn.commit()
            flash(f'Товар "{product_name}" успешно добавлен', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Ошибка при добавлении товара: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('manage_products'))

@app.route('/management/products/delete/<product_name>')
@admin_required
def delete_product_route(product_name):
    """Удалить товар"""
    conn = get_connection()
    cursor = conn.cursor()
    try:
        # Проверяем, используется ли товар в реальных записях (не системных)
        cursor.execute("""
            SELECT COUNT(*) FROM arrivals 
            WHERE product_name = %s AND wagon_number != 'SYSTEM'
        """, (product_name,))
        count = cursor.fetchone()[0]
        
        if count > 0:
            flash(f'Невозможно удалить товар "{product_name}". Он используется в {count} записях.', 'error')
        else:
            # Удаляем все записи товара (включая системные)
            cursor.execute("DELETE FROM arrivals WHERE product_name = %s", (product_name,))
            conn.commit()
            flash(f'Товар "{product_name}" успешно удален', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Ошибка при удалении товара: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('manage_products'))

# ============= УПРАВЛЕНИЕ ФИРМАМИ =============

@app.route('/management/firms')
@admin_required
def manage_firms():
    """Управление фирмами"""
    firms = get_firms()
    return render_template('manage_firms.html', firms=firms)

@app.route('/management/firms/add', methods=['POST'])
@admin_required
def add_firm_route():
    """Добавить фирму"""
    firm_name = request.form.get('firm_name')
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO firms (firm_name) VALUES (%s)", (firm_name,))
        conn.commit()
        flash(f'Фирма "{firm_name}" успешно добавлена', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Ошибка при добавлении фирмы: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('manage_firms'))

@app.route('/management/firms/delete/<int:firm_id>')
@admin_required
def delete_firm_route(firm_id):
    """Удалить фирму"""
    conn = get_connection()
    cursor = conn.cursor()
    try:
        # Проверяем, используется ли фирма
        cursor.execute("SELECT COUNT(*) FROM arrivals WHERE firm_id = %s", (firm_id,))
        count = cursor.fetchone()[0]
        if count > 0:
            flash(f'Невозможно удалить фирму. Она используется в {count} записях.', 'error')
        else:
            cursor.execute("DELETE FROM firms WHERE firm_id = %s", (firm_id,))
            conn.commit()
            flash('Фирма успешно удалена', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Ошибка при удалении фирмы: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('manage_firms'))

# ============= УПРАВЛЕНИЕ ЦЕНАМИ =============

@app.route('/management/prices')
@admin_required
def manage_prices():
    """Управление ценами"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, product_name, price, updated_at
        FROM prices
        ORDER BY updated_at DESC
    """)
    prices = cursor.fetchall()
    
    # Получаем список товаров
    cursor.execute("SELECT DISTINCT product_name FROM arrivals ORDER BY product_name")
    products = cursor.fetchall()
    
    conn.close()
    return render_template('manage_prices.html', prices=prices, products=products)

@app.route('/management/prices/add', methods=['POST'])
@admin_required
def add_price_route():
    """Установить цену"""
    product_name = request.form.get('product_name')
    price = float(request.form.get('price'))
    
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO prices (product_name, price) 
            VALUES (%s, %s) 
            ON CONFLICT (product_name) 
            DO UPDATE SET price = %s, updated_at = CURRENT_TIMESTAMP
        """, (product_name, price, price))
        conn.commit()
        flash(f'Цена для "{product_name}" установлена: ${price}', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Ошибка при установке цены: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('manage_prices'))

@app.route('/management/prices/delete/<int:price_id>')
@admin_required
def delete_price_route(price_id):
    """Удалить цену"""
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM prices WHERE id = %s", (price_id,))
        conn.commit()
        flash('Цена успешно удалена', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Ошибка при удалении цены: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('manage_prices'))

# ============= УПРАВЛЕНИЕ КОАЛИЦЕЙ =============

@app.route('/management/coalitions')
@admin_required
def manage_coalitions():
    """Управление коалицей"""
    coalitions = get_coalitions()
    return render_template('manage_coalitions.html', coalitions=coalitions)

@app.route('/management/coalitions/add', methods=['POST'])
@admin_required
def add_coalition_route():
    """Добавить коалицу"""
    coalition_name = request.form.get('coalition_name')
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO coalitions (name) VALUES (%s)", (coalition_name,))
        conn.commit()
        flash(f'Коалица "{coalition_name}" успешно добавлена', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Ошибка при добавлении коалицы: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('manage_coalitions'))

@app.route('/management/coalitions/delete/<int:coalition_id>')
@admin_required
def delete_coalition_route(coalition_id):
    """Удалить коалицу"""
    conn = get_connection()
    cursor = conn.cursor()
    try:
        # Проверяем, используется ли коалица
        cursor.execute("SELECT COUNT(*) FROM departures WHERE coalition_id = %s", (coalition_id,))
        count = cursor.fetchone()[0]
        if count > 0:
            flash(f'Невозможно удалить коалицу. Она используется в {count} записях.', 'error')
        else:
            cursor.execute("DELETE FROM coalitions WHERE id = %s", (coalition_id,))
            conn.commit()
            flash('Коалица успешно удалена', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Ошибка при удалении коалицы: {str(e)}', 'error')
    finally:
        conn.close()
    return redirect(url_for('manage_coalitions'))

# ============= УДАЛЕНИЕ ЗАПИСЕЙ (ТОЛЬКО ДЛЯ АДМИНОВ) =============

@app.route('/arrivals/edit/<int:arrival_id>', methods=['GET', 'POST'])
@admin_required
def edit_arrival_page(arrival_id):
    """Редактировать приход"""
    if request.method == 'POST':
        arrival_date = request.form.get('arrival_date')
        wagon_number = request.form.get('wagon_number')
        firm_id = int(request.form.get('firm_id'))
        warehouse_id = int(request.form.get('warehouse_id'))
        product_name = request.form.get('product_name')
        source = request.form.get('source')
        qty_doc = int(request.form.get('qty_doc'))
        qty_actual = int(request.form.get('qty_actual'))
        notes = request.form.get('notes')
        
        conn = get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                UPDATE arrivals 
                SET arrival_date=%s, wagon_number=%s, firm_id=%s, warehouse_id=%s, 
                    product_name=%s, source=%s, quantity_doc=%s, quantity_actual=%s, notes=%s
                WHERE id=%s
            """, (arrival_date, wagon_number, firm_id, warehouse_id, product_name, 
                  source, qty_doc, qty_actual, notes, arrival_id))
            conn.commit()
            flash('Приход успешно обновлен', 'success')
            return redirect(url_for('arrivals'))
        except Exception as e:
            conn.rollback()
            flash(f'Ошибка: {str(e)}', 'error')
        finally:
            conn.close()
    
    # GET - показываем форму с данными
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM arrivals WHERE id = %s", (arrival_id,))
    arrival = cursor.fetchone()
    conn.close()
    
    if not arrival:
        flash('Приход не найден', 'error')
        return redirect(url_for('arrivals'))
    
    firms = get_firms()
    warehouses = get_warehouses()
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT product_name FROM arrivals ORDER BY product_name")
    products = cursor.fetchall()
    conn.close()
    
    return render_template('edit_arrival.html', arrival=arrival, firms=firms, 
                         warehouses=warehouses, products=products)

@app.route('/arrivals/delete/<int:arrival_id>')
@admin_required
def delete_arrival_route(arrival_id):
    """Удалить приход"""
    success, message = delete_arrival(arrival_id)
    if success:
        flash('Приход успешно удален', 'success')
    else:
        flash(f'Ошибка: {message}', 'error')
    return redirect(url_for('arrivals'))

@app.route('/departures/edit/<int:departure_id>', methods=['GET', 'POST'])
@admin_required
def edit_departure_page(departure_id):
    """Редактировать расход"""
    if request.method == 'POST':
        departure_date = request.form.get('departure_date')
        coalition_id = int(request.form.get('coalition_id'))
        firm_id = int(request.form.get('firm_id'))
        warehouse_id = int(request.form.get('warehouse_id'))
        product_name = request.form.get('product_name')
        client_id = request.form.get('client_id')
        if client_id:
            client_id = int(client_id)
        else:
            client_id = None
        quantity = int(request.form.get('quantity'))
        price = float(request.form.get('price'))
        notes = request.form.get('notes')
        total = (quantity / 20) * price
        
        conn = get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                UPDATE departures 
                SET departure_date=%s, coalition_id=%s, firm_id=%s, warehouse_id=%s, 
                    product_name=%s, quantity=%s, price=%s, notes=%s, client_id=%s, total=%s
                WHERE id=%s
            """, (departure_date, coalition_id, firm_id, warehouse_id, product_name, 
                  quantity, price, notes, client_id, total, departure_id))
            conn.commit()
            flash('Расход успешно обновлен', 'success')
            return redirect(url_for('departures'))
        except Exception as e:
            conn.rollback()
            flash(f'Ошибка: {str(e)}', 'error')
        finally:
            conn.close()
    
    # GET
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM departures WHERE id = %s", (departure_id,))
    departure = cursor.fetchone()
    conn.close()
    
    if not departure:
        flash('Расход не найден', 'error')
        return redirect(url_for('departures'))
    
    coalitions = get_coalitions()
    firms = get_firms()
    warehouses = get_warehouses()
    clients = get_clients()
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT product_name FROM arrivals ORDER BY product_name")
    products = cursor.fetchall()
    conn.close()
    
    return render_template('edit_departure.html', departure=departure, coalitions=coalitions,
                         firms=firms, warehouses=warehouses, clients=clients, products=products)

@app.route('/departures/delete/<int:departure_id>')
@admin_required
def delete_departure_route(departure_id):
    """Удалить расход"""
    success, message = delete_departure(departure_id)
    if success:
        flash('Расход успешно удален', 'success')
    else:
        flash(f'Ошибка: {message}', 'error')
    return redirect(url_for('departures'))

@app.route('/payments/edit/<int:payment_id>', methods=['GET', 'POST'])
@admin_required
def edit_payment_page(payment_id):
    """Редактировать погашение"""
    if request.method == 'POST':
        payment_date = request.form.get('payment_date')
        client_id = int(request.form.get('client_id'))
        somoni = float(request.form.get('somoni'))
        rate = float(request.form.get('rate'))
        notes = request.form.get('notes')
        total_usd = somoni / rate
        
        conn = get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                UPDATE payments 
                SET payment_date=%s, client_id=%s, somoni=%s, exchange_rate=%s, total_usd=%s, notes=%s
                WHERE id=%s
            """, (payment_date, client_id, somoni, rate, total_usd, notes, payment_id))
            conn.commit()
            flash('Погашение успешно обновлено', 'success')
            return redirect(url_for('payments'))
        except Exception as e:
            conn.rollback()
            flash(f'Ошибка: {str(e)}', 'error')
        finally:
            conn.close()
    
    # GET
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM payments WHERE id = %s", (payment_id,))
    payment = cursor.fetchone()
    conn.close()
    
    if not payment:
        flash('Погашение не найдено', 'error')
        return redirect(url_for('payments'))
    
    clients = get_clients()
    return render_template('edit_payment.html', payment=payment, clients=clients)

@app.route('/payments/delete/<int:payment_id>')
@admin_required
def delete_payment_route(payment_id):
    """Удалить погашение"""
    success, message = delete_payment(payment_id)
    if success:
        flash('Погашение успешно удалено', 'success')
    else:
        flash(f'Ошибка: {message}', 'error')
    return redirect(url_for('payments'))

@app.route('/partners/edit/<int:partner_id>', methods=['GET', 'POST'])
@admin_required
def edit_partner_page(partner_id):
    """Редактировать партнера"""
    if request.method == 'POST':
        partner_date = request.form.get('partner_date')
        client_id = int(request.form.get('client_id'))
        somoni = float(request.form.get('somoni'))
        rate = float(request.form.get('rate'))
        notes = request.form.get('notes')
        total_usd = somoni / rate
        
        conn = get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                UPDATE partners 
                SET partner_date=%s, client_id=%s, somoni=%s, exchange_rate=%s, total_usd=%s, notes=%s
                WHERE id=%s
            """, (partner_date, client_id, somoni, rate, total_usd, notes, partner_id))
            conn.commit()
            flash('Партнер успешно обновлен', 'success')
            return redirect(url_for('partners'))
        except Exception as e:
            conn.rollback()
            flash(f'Ошибка: {str(e)}', 'error')
        finally:
            conn.close()
    
    # GET
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM partners WHERE id = %s", (partner_id,))
    partner = cursor.fetchone()
    conn.close()
    
    if not partner:
        flash('Партнер не найден', 'error')
        return redirect(url_for('partners'))
    
    clients = get_clients()
    return render_template('edit_partner.html', partner=partner, clients=clients)

@app.route('/partners/delete/<int:partner_id>')
@admin_required
def delete_partner_route(partner_id):
    """Удалить партнера"""
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('DELETE FROM partners WHERE id = %s', (partner_id,))
        conn.commit()
        cursor.close()
        conn.close()
        flash('Партнер успешно удален', 'success')
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'Ошибка: {str(e)}', 'error')
    return redirect(url_for('partners'))

# ============= ДОПОЛНИТЕЛЬНЫЕ ОТЧЕТЫ =============

@app.route('/reports/arrivals_monthly')
@login_required
def arrivals_monthly():
    """Отчет по приходам с группировкой по месяцам"""
    from datetime import datetime
    
    user_id = session.get('user_id')
    role = session.get('role')
    current_year = datetime.now().year
    
    user_warehouses = get_user_warehouses(user_id)
    warehouse_ids = [wh[0] for wh in user_warehouses] if user_warehouses else []
    
    date_from = f'{current_year}-01-01'
    date_to = f'{current_year}-12-31'
    
    all_arrivals = []
    if role == 'admin':
        all_arrivals = get_arrivals(limit=10000, date_from=date_from, date_to=date_to)
    else:
        for wh_id in warehouse_ids:
            arrivals_wh = get_arrivals(limit=10000, warehouse_id=wh_id, date_from=date_from, date_to=date_to)
            all_arrivals.extend(arrivals_wh)
    
    # Группируем по месяцам
    months_data = {}
    month_names = {
        1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
        5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
        9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
    }
    
    for arr in all_arrivals:
        arr_date = arr[1]
        qty_act = arr[8]
        
        if isinstance(arr_date, str):
            month = int(arr_date.split('-')[1])
        else:
            month = arr_date.month
        
        if month not in months_data:
            months_data[month] = {'count': 0, 'total_tons': 0}
        
        months_data[month]['count'] += 1
        months_data[month]['total_tons'] += qty_act / 20
    
    return render_template('arrivals_monthly.html', 
                         months_data=months_data,
                         month_names=month_names,
                         current_year=current_year)

@app.route('/reports/departures_monthly')
@login_required
def departures_monthly():
    """Отчет по расходам с группировкой по месяцам"""
    from datetime import datetime
    
    user_id = session.get('user_id')
    role = session.get('role')
    current_year = datetime.now().year
    
    user_warehouses = get_user_warehouses(user_id)
    warehouse_ids = [wh[0] for wh in user_warehouses] if user_warehouses else []
    
    date_from = f'{current_year}-01-01'
    date_to = f'{current_year}-12-31'
    
    all_departures = []
    if role == 'admin':
        all_departures = get_departures(limit=10000, date_from=date_from, date_to=date_to)
    else:
        for wh_id in warehouse_ids:
            departures_wh = get_departures(limit=10000, warehouse_id=wh_id, date_from=date_from, date_to=date_to)
            all_departures.extend(departures_wh)
    
    # Группируем по месяцам
    months_data = {}
    month_names = {
        1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
        5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
        9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
    }
    
    for dep in all_departures:
        dep_date = dep[1]
        total = dep[12]
        
        if isinstance(dep_date, str):
            month = int(dep_date.split('-')[1])
        else:
            month = dep_date.month
        
        if month not in months_data:
            months_data[month] = {'count': 0, 'total_usd': 0}
        
        months_data[month]['count'] += 1
        months_data[month]['total_usd'] += float(total) if total else 0
    
    return render_template('departures_monthly.html',
                         months_data=months_data,
                         month_names=month_names,
                         current_year=current_year)

@app.route('/reports/client_card')
@login_required
def client_card_select():
    """Выбор клиента для карточки"""
    clients = get_clients()
    return render_template('client_card_select.html', clients=clients)

@app.route('/reports/client_card/<int:client_id>')
@login_required
def client_card(client_id):
    """Карточка клиента с детальной информацией"""
    conn = get_connection()
    cursor = conn.cursor()
    
    # Получаем имя клиента
    cursor.execute('SELECT name FROM clients WHERE id = %s', (client_id,))
    client_result = cursor.fetchone()
    if not client_result:
        flash('Клиент не найден', 'error')
        return redirect(url_for('client_card_select'))
    
    client_name = client_result[0]
    
    # Получаем покупки
    cursor.execute('''
        SELECT d.departure_date, w.name, d.product_name, d.quantity, d.price, d.total
        FROM departures d
        LEFT JOIN warehouses w ON d.warehouse_id = w.id
        WHERE d.client_id = %s
        ORDER BY d.departure_date DESC
    ''', (client_id,))
    purchases = cursor.fetchall()
    
    # Получаем погашения
    cursor.execute('''
        SELECT p.payment_date, p.somoni, p.exchange_rate, p.total_usd, p.notes
        FROM payments p
        WHERE p.client_id = %s
        ORDER BY p.payment_date DESC
    ''', (client_id,))
    payments_list = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    # Считаем итоги
    total_purchases = sum(float(p[5]) if p[5] else 0 for p in purchases)
    total_payments = sum(float(p[3]) if p[3] else 0 for p in payments_list)
    balance = total_purchases - total_payments
    
    return render_template('client_card.html',
                         client_name=client_name,
                         client_id=client_id,
                         purchases=purchases,
                         payments=payments_list,
                         total_purchases=total_purchases,
                         total_payments=total_payments,
                         balance=balance)

@app.route('/reports/debt_notifications')
@login_required
def debt_notifications():
    """Уведомления о долгах"""
    from datetime import datetime, timedelta
    
    period = request.args.get('period', 'today')
    
    # Определяем дату
    if period == 'today':
        target_date = datetime.now().date()
        period_name = 'Сегодня'
    elif period == 'yesterday':
        target_date = (datetime.now() - timedelta(days=1)).date()
        period_name = 'Вчера'
    elif period == '7days':
        target_date = (datetime.now() - timedelta(days=7)).date()
        period_name = '7 дней назад'
    elif period == '14days':
        target_date = (datetime.now() - timedelta(days=14)).date()
        period_name = '14 дней назад'
    elif period == '30days':
        target_date = (datetime.now() - timedelta(days=30)).date()
        period_name = '30 дней назад'
    else:
        target_date = datetime.now().date()
        period_name = 'Сегодня'
    
    conn = get_connection()
    cursor = conn.cursor()
    
    # Получаем покупки на выбранную дату
    cursor.execute('''
        SELECT c.name, SUM(d.total) as total_purchases
        FROM departures d
        JOIN clients c ON d.client_id = c.id
        WHERE d.departure_date = %s AND d.client_id IS NOT NULL
        GROUP BY c.id, c.name
        ORDER BY c.name
    ''', (target_date,))
    
    purchases_on_date = cursor.fetchall()
    
    # Получаем общие долги клиентов
    cursor.execute('''
        SELECT c.id, c.name,
               COALESCE(SUM(d.total), 0) as total_sales,
               COALESCE((SELECT SUM(p.total_usd) FROM payments p WHERE p.client_id = c.id), 0) as total_paid
        FROM clients c
        LEFT JOIN departures d ON c.id = d.client_id
        GROUP BY c.id, c.name
        HAVING COALESCE(SUM(d.total), 0) - COALESCE((SELECT SUM(p.total_usd) FROM payments p WHERE p.client_id = c.id), 0) > 0
        ORDER BY c.name
    ''')
    
    debts = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('debt_notifications.html',
                         period=period,
                         period_name=period_name,
                         target_date=target_date,
                         purchases_on_date=purchases_on_date,
                         debts=debts)

@app.route('/reports/wagon_summary')
@login_required
def wagon_summary():
    """Итоги вагонов"""
    user_id = session.get('user_id')
    role = session.get('role')
    
    user_warehouses = get_user_warehouses(user_id)
    warehouse_ids = [wh[0] for wh in user_warehouses] if user_warehouses else []
    
    all_arrivals = []
    if role == 'admin':
        all_arrivals = get_arrivals(limit=10000)
    else:
        for wh_id in warehouse_ids:
            arrivals_wh = get_arrivals(limit=10000, warehouse_id=wh_id)
            all_arrivals.extend(arrivals_wh)
    
    # Группируем по складам и товарам
    warehouse_data = {}
    
    for arr in all_arrivals:
        warehouse = arr[4]
        firm = arr[3]
        product = arr[5]
        qty_doc = arr[7]
        qty_act = arr[8]
        
        if warehouse not in warehouse_data:
            warehouse_data[warehouse] = {}
        
        key = f"{firm} {product}"
        
        if key not in warehouse_data[warehouse]:
            warehouse_data[warehouse][key] = {
                'wagons': 0,
                'qty_doc': 0,
                'qty_act': 0
            }
        
        warehouse_data[warehouse][key]['wagons'] += 1
        warehouse_data[warehouse][key]['qty_doc'] += qty_doc
        warehouse_data[warehouse][key]['qty_act'] += qty_act
    
    return render_template('wagon_summary.html', warehouse_data=warehouse_data)

# ============= ЭКСПОРТ В EXCEL =============

@app.route('/export/client_debts')
@login_required
def export_client_debts():
    """Экспорт долгов клиентов в Excel"""
    from io import BytesIO
    
    conn = get_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT client_id, SUM(total) as total_sales
        FROM departures
        WHERE client_id IS NOT NULL
        GROUP BY client_id
    ''')
    sales_data = {row[0]: float(row[1]) for row in cursor.fetchall()}
    
    cursor.execute('''
        SELECT client_id, SUM(total_usd) as total_paid
        FROM payments
        GROUP BY client_id
    ''')
    payments_data = {row[0]: float(row[1]) for row in cursor.fetchall()}
    
    cursor.execute('SELECT id, name FROM clients')
    clients_data = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    debts = []
    for client_id, client_name in clients_data:
        total_sales = sales_data.get(client_id, 0)
        total_paid = payments_data.get(client_id, 0)
        debt = total_sales - total_paid
        
        if debt > 0:
            debts.append({
                'client_name': client_name,
                'total_sales': total_sales,
                'total_paid': total_paid,
                'debt': debt
            })
    
    debts.sort(key=lambda x: x['debt'], reverse=True)
    
    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Долги клиентов"
    
    # Заголовки
    headers = ['№', 'Клиент', 'Сумма продаж ($)', 'Погашения ($)', 'Остаток долга ($)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Данные
    for idx, debt in enumerate(debts, 2):
        ws.cell(row=idx, column=1, value=idx-1)
        ws.cell(row=idx, column=2, value=debt['client_name'])
        ws.cell(row=idx, column=3, value=debt['total_sales'])
        ws.cell(row=idx, column=4, value=debt['total_paid'])
        ws.cell(row=idx, column=5, value=debt['debt'])
    
    # Итоговая строка
    total_row = len(debts) + 2
    ws.cell(row=total_row, column=1, value='ИТОГО:')
    ws.cell(row=total_row, column=2, value='')
    ws.cell(row=total_row, column=3, value=sum(d['total_sales'] for d in debts))
    ws.cell(row=total_row, column=4, value=sum(d['total_paid'] for d in debts))
    ws.cell(row=total_row, column=5, value=sum(d['debt'] for d in debts))
    
    for col in range(1, 6):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # Форматирование
    for row in ws.iter_rows(min_row=2, max_row=total_row, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = '#,##0.00'
    
    # Ширина колонок
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    
    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'client_debts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/export/client_card/<int:client_id>')
@login_required
def export_client_card(client_id):
    """Экспорт карточки клиента в Excel"""
    from io import BytesIO
    
    conn = get_connection()
    cursor = conn.cursor()
    
    # Получаем имя клиента
    cursor.execute('SELECT name FROM clients WHERE id = %s', (client_id,))
    client_result = cursor.fetchone()
    if not client_result:
        flash('Клиент не найден', 'error')
        return redirect(url_for('client_card_select'))
    
    client_name = client_result[0]
    
    # Получаем покупки
    cursor.execute('''
        SELECT d.departure_date, w.name, d.product_name, d.quantity, d.price, d.total
        FROM departures d
        LEFT JOIN warehouses w ON d.warehouse_id = w.id
        WHERE d.client_id = %s
        ORDER BY d.departure_date DESC
    ''', (client_id,))
    purchases = cursor.fetchall()
    
    # Получаем погашения
    cursor.execute('''
        SELECT p.payment_date, p.somoni, p.exchange_rate, p.total_usd, p.notes
        FROM payments p
        WHERE p.client_id = %s
        ORDER BY p.payment_date DESC
    ''', (client_id,))
    payments_list = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    # Создаем Excel файл
    wb = Workbook()
    
    # Лист 1: Покупки
    ws1 = wb.active
    ws1.title = "Покупки"
    
    headers1 = ['Дата', 'Склад', 'Товар', 'Кол-во', 'Цена', 'Сумма']
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    for idx, purchase in enumerate(purchases, 2):
        ws1.cell(row=idx, column=1, value=str(purchase[0]))
        ws1.cell(row=idx, column=2, value=purchase[1])
        ws1.cell(row=idx, column=3, value=purchase[2])
        ws1.cell(row=idx, column=4, value=purchase[3])
        ws1.cell(row=idx, column=5, value=float(purchase[4]) if purchase[4] else 0)
        ws1.cell(row=idx, column=6, value=float(purchase[5]) if purchase[5] else 0)
    
    # Лист 2: Погашения
    ws2 = wb.create_sheet("Погашения")
    
    headers2 = ['Дата', 'Сомони', 'Курс', 'Сумма ($)', 'Примечания']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    for idx, payment in enumerate(payments_list, 2):
        ws2.cell(row=idx, column=1, value=str(payment[0]))
        ws2.cell(row=idx, column=2, value=float(payment[1]) if payment[1] else 0)
        ws2.cell(row=idx, column=3, value=float(payment[2]) if payment[2] else 0)
        ws2.cell(row=idx, column=4, value=float(payment[3]) if payment[3] else 0)
        ws2.cell(row=idx, column=5, value=payment[4] or '')
    
    # Лист 3: Сводка
    ws3 = wb.create_sheet("Сводка")
    
    total_purchases = sum(float(p[5]) if p[5] else 0 for p in purchases)
    total_payments = sum(float(p[3]) if p[3] else 0 for p in payments_list)
    balance = total_purchases - total_payments
    
    ws3.cell(row=1, column=1, value='Клиент:')
    ws3.cell(row=1, column=2, value=client_name)
    ws3.cell(row=2, column=1, value='Сумма продаж:')
    ws3.cell(row=2, column=2, value=total_purchases)
    ws3.cell(row=3, column=1, value='Погашения:')
    ws3.cell(row=3, column=2, value=total_payments)
    ws3.cell(row=4, column=1, value='Остаток долга:')
    ws3.cell(row=4, column=2, value=balance)
    
    for row in range(1, 5):
        ws3.cell(row=row, column=1).font = Font(bold=True)
    
    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'client_card_{client_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/export/daily_report')
@login_required
def export_daily_report():
    """Экспорт отчета за день в Excel"""
    from io import BytesIO
    
    date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    
    user_id = session.get('user_id')
    role = session.get('role')
    
    user_warehouses = get_user_warehouses(user_id)
    warehouse_ids = [wh[0] for wh in user_warehouses] if user_warehouses else []
    
    # Получаем приходы
    if role == 'admin':
        arrivals_list = get_arrivals(limit=1000, date_from=date, date_to=date)
    else:
        arrivals_list = []
        for wh_id in warehouse_ids:
            arrivals_list.extend(get_arrivals(limit=1000, warehouse_id=wh_id, date_from=date, date_to=date))
    
    # Получаем расходы
    if role == 'admin':
        departures_list = get_departures(limit=1000, date_from=date, date_to=date)
    else:
        departures_list = []
        for wh_id in warehouse_ids:
            departures_list.extend(get_departures(limit=1000, warehouse_id=wh_id, date_from=date, date_to=date))
    
    # Создаем Excel файл
    wb = Workbook()
    
    # Лист 1: Приходы
    ws1 = wb.active
    ws1.title = "Приходы"
    
    headers1 = ['Дата', 'Вагон', 'Фирма', 'Склад', 'Товар', 'Источник', 'По док.', 'Факт.']
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    for idx, arrival in enumerate(arrivals_list, 2):
        ws1.cell(row=idx, column=1, value=str(arrival[1]))
        ws1.cell(row=idx, column=2, value=arrival[2])
        ws1.cell(row=idx, column=3, value=arrival[3])
        ws1.cell(row=idx, column=4, value=arrival[4])
        ws1.cell(row=idx, column=5, value=arrival[5])
        ws1.cell(row=idx, column=6, value=arrival[6])
        ws1.cell(row=idx, column=7, value=arrival[7])
        ws1.cell(row=idx, column=8, value=arrival[8])
    
    # Лист 2: Расходы
    ws2 = wb.create_sheet("Расходы")
    
    headers2 = ['Дата', 'Коалиция', 'Фирма', 'Склад', 'Товар', 'Клиент', 'Кол-во', 'Цена', 'Сумма']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    for idx, dep in enumerate(departures_list, 2):
        ws2.cell(row=idx, column=1, value=str(dep[1]))
        ws2.cell(row=idx, column=2, value=dep[2])
        ws2.cell(row=idx, column=3, value=dep[3])
        ws2.cell(row=idx, column=4, value=dep[4])
        ws2.cell(row=idx, column=5, value=dep[5])
        ws2.cell(row=idx, column=6, value=dep[11] or '-')
        ws2.cell(row=idx, column=7, value=dep[6])
        ws2.cell(row=idx, column=8, value=float(dep[7]) if dep[7] else 0)
        ws2.cell(row=idx, column=9, value=float(dep[12]) if dep[12] else 0)
    
    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'daily_report_{date}.xlsx'
    )

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
